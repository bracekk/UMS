from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from ums_core import db_compat as sqlite3
from datetime import datetime, timedelta
import calendar
import math
import os
import json
import secrets
import hashlib
from io import BytesIO
from email.message import EmailMessage
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
import resend

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from ums_core.bootstrap import bootstrap_database


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-key-change-me")


load_dotenv()



DB_PATH = os.environ.get("DATABASE_PATH", "database.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row

    if not getattr(sqlite3, "USE_POSTGRES", False):
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute("PRAGMA busy_timeout = 30000;")

    return conn


def _table_exists(cursor, table_name):
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    )
    return cursor.fetchone() is not None


def _column_exists(cursor, table_name, column_name):
    cursor.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
          AND column_name = %s
        LIMIT 1
        """,
        (table_name, column_name),
    )
    return cursor.fetchone() is not None


def _safe_add_column(cursor, table_name, column_sql):
    try:
        cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_sql}")
    except Exception:
        pass


def ensure_render_safe_schema():
    init_db()



def init_db():
    conn = get_connection()
    cursor = conn.cursor()

    # Core schema based on the working legacy SQLite database.
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT,
            company TEXT,
            email TEXT UNIQUE,
            password TEXT,
            company_id INTEGER,
            role TEXT NOT NULL DEFAULT 'admin',
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS user_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            company_id INTEGER,
            permission_key TEXT NOT NULL,
            allowed INTEGER NOT NULL DEFAULT 1,
            UNIQUE(user_id, permission_key)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            supplier_code TEXT,
            contact_person TEXT,
            email TEXT,
            phone TEXT,
            address TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL,
            item_name TEXT NOT NULL,
            description TEXT,
            measurement_unit TEXT NOT NULL DEFAULT 'pcs',
            unit TEXT,
            unit_price REAL NOT NULL DEFAULT 0,
            stock_quantity REAL NOT NULL DEFAULT 0,
            min_stock REAL NOT NULL DEFAULT 0,
            reserved_quantity REAL NOT NULL DEFAULT 0,
            available_quantity REAL NOT NULL DEFAULT 0,
            supplier_id INTEGER,
            company_id INTEGER,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS product_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(company_id, name),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_code TEXT NOT NULL,
            product_name TEXT NOT NULL,
            description TEXT,
            measurement_unit TEXT NOT NULL DEFAULT 'pcs',
            stock_quantity REAL NOT NULL DEFAULT 0,
            time_per_unit REAL NOT NULL DEFAULT 0,
            company_id INTEGER,
            group_id INTEGER,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (group_id) REFERENCES product_groups(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS workstation_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            color TEXT NOT NULL DEFAULT '#6366f1',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(company_id, name),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS workstations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            hours_per_shift REAL NOT NULL DEFAULT 8,
            shifts_per_day INTEGER NOT NULL DEFAULT 1,
            working_days_per_month INTEGER NOT NULL DEFAULT 20,
            color TEXT NOT NULL DEFAULT '#3b82f6',
            company_id INTEGER,
            cost_per_hour REAL NOT NULL DEFAULT 0,
            group_id INTEGER,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (group_id) REFERENCES workstation_groups(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT,
            customer TEXT,
            status TEXT,
            due_date TEXT,
            priority TEXT,
            product_id INTEGER,
            quantity REAL DEFAULT 1,
            materials_reserved INTEGER NOT NULL DEFAULT 0,
            materials_consumed INTEGER NOT NULL DEFAULT 0,
            finished_stock_added INTEGER NOT NULL DEFAULT 0,
            company_id INTEGER,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS order_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            batch_number TEXT,
            name TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            notes TEXT,
            created_by INTEGER,
            launched_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_deleted INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS batch_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            company_id INTEGER NOT NULL,
            order_number TEXT NOT NULL,
            product_id INTEGER NOT NULL,
            quantity REAL NOT NULL DEFAULT 1,
            due_date TEXT,
            priority TEXT DEFAULT 'Medium',
            notes TEXT,
            launched_order_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (batch_id) REFERENCES order_batches(id),
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (launched_order_id) REFERENCES orders(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bom (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL DEFAULT 0,
            quantity REAL NOT NULL,
            component_type TEXT NOT NULL DEFAULT 'item',
            child_product_id INTEGER,
            company_id INTEGER,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (item_id) REFERENCES items(id),
            FOREIGN KEY (child_product_id) REFERENCES products(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS product_job_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            workstation_id INTEGER NOT NULL,
            job_name TEXT NOT NULL,
            sequence INTEGER NOT NULL DEFAULT 1,
            estimated_hours REAL NOT NULL DEFAULT 0,
            company_id INTEGER,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (workstation_id) REFERENCES workstations(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS order_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            product_job_template_id INTEGER,
            job_product_id INTEGER,
            workstation_id INTEGER NOT NULL,
            job_name TEXT NOT NULL,
            sequence INTEGER NOT NULL,
            planned_quantity REAL NOT NULL,
            completed_quantity REAL NOT NULL DEFAULT 0,
            estimated_hours REAL NOT NULL,
            status TEXT NOT NULL DEFAULT 'Waiting',
            planned_start TEXT,
            planned_end TEXT,
            parent_job_id INTEGER,
            is_split_child INTEGER NOT NULL DEFAULT 0,
            company_id INTEGER,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (job_product_id) REFERENCES products(id),
            FOREIGN KEY (workstation_id) REFERENCES workstations(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS purchase_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            request_number TEXT,
            item_id INTEGER,
            supplier_id INTEGER,
            order_id INTEGER,
            title TEXT NOT NULL,
            description TEXT,
            quantity REAL NOT NULL,
            unit TEXT,
            status TEXT DEFAULT 'draft',
            priority TEXT DEFAULT 'normal',
            needed_by DATE,
            requested_by INTEGER,
            approved_by INTEGER,
            ordered_by INTEGER,
            notes TEXT,
            source_type TEXT DEFAULT 'manual',
            source_batch_id INTEGER,
            source_batch_order_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY (item_id) REFERENCES items(id),
            FOREIGN KEY (requested_by) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS shortages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            order_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            required_qty REAL NOT NULL DEFAULT 0,
            reserved_qty REAL NOT NULL DEFAULT 0,
            missing_qty REAL NOT NULL DEFAULT 0,
            covered INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (item_id) REFERENCES items(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS stock_destinations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            destination_code TEXT,
            notes TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS product_transfers_out (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            destination_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            transfer_date TEXT NOT NULL,
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (destination_id) REFERENCES stock_destinations(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS production_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            job_id INTEGER NOT NULL,
            order_id INTEGER,
            product_id INTEGER,
            workstation_id INTEGER,
            report_type TEXT NOT NULL,
            quantity REAL NOT NULL DEFAULT 0,
            notes TEXT,
            reported_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            unit TEXT DEFAULT 'pcs',
            FOREIGN KEY (company_id) REFERENCES companies(id),
            FOREIGN KEY (job_id) REFERENCES order_jobs(id),
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (workstation_id) REFERENCES workstations(id),
            FOREIGN KEY (reported_by) REFERENCES users(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dashboard_layouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            company_id INTEGER NOT NULL,
            page_key TEXT NOT NULL DEFAULT 'dashboard',
            layout_json TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, company_id, page_key),
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            used_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # Safe incremental column upgrades.
    desired_columns = {
        "users": [
            "company TEXT",
            "company_id INTEGER",
            "role TEXT NOT NULL DEFAULT 'admin'",
            "is_active INTEGER DEFAULT 1",
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        ],
        "user_permissions": ["company_id INTEGER", "allowed INTEGER NOT NULL DEFAULT 1"],
        "items": [
            "description TEXT",
            "measurement_unit TEXT DEFAULT 'pcs'",
            "unit TEXT",
            "supplier_id INTEGER",
            "company_id INTEGER",
            "reserved_quantity REAL NOT NULL DEFAULT 0",
            "available_quantity REAL NOT NULL DEFAULT 0",
        ],
        "products": [
            "description TEXT",
            "measurement_unit TEXT NOT NULL DEFAULT 'pcs'",
            "time_per_unit REAL NOT NULL DEFAULT 0",
            "company_id INTEGER",
            "group_id INTEGER",
        ],
        "workstations": [
            "description TEXT",
            "hours_per_shift REAL NOT NULL DEFAULT 8",
            "shifts_per_day INTEGER NOT NULL DEFAULT 1",
            "working_days_per_month INTEGER NOT NULL DEFAULT 20",
            "company_id INTEGER",
            "cost_per_hour REAL NOT NULL DEFAULT 0",
            "group_id INTEGER",
        ],
        "orders": [
            "customer TEXT",
            "product_id INTEGER",
            "quantity REAL DEFAULT 1",
            "materials_reserved INTEGER NOT NULL DEFAULT 0",
            "materials_consumed INTEGER NOT NULL DEFAULT 0",
            "finished_stock_added INTEGER NOT NULL DEFAULT 0",
            "company_id INTEGER",
        ],
        "order_batches": [
            "batch_number TEXT",
            "notes TEXT",
            "created_by INTEGER",
            "launched_at TEXT",
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "is_deleted INTEGER NOT NULL DEFAULT 0",
        ],
        "batch_orders": [
            "notes TEXT",
            "launched_order_id INTEGER",
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        ],
        "bom": ["component_type TEXT NOT NULL DEFAULT 'item'", "child_product_id INTEGER", "company_id INTEGER"],
        "product_job_templates": ["company_id INTEGER"],
        "order_jobs": [
            "product_job_template_id INTEGER",
            "job_product_id INTEGER",
            "planned_start TEXT",
            "planned_end TEXT",
            "parent_job_id INTEGER",
            "is_split_child INTEGER NOT NULL DEFAULT 0",
            "company_id INTEGER",
        ],
        "purchase_requests": [
            "request_number TEXT",
            "source_type TEXT DEFAULT 'manual'",
            "source_batch_id INTEGER",
            "source_batch_order_id INTEGER",
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "updated_at TIMESTAMP",
        ],
        "shortages": ["covered INTEGER NOT NULL DEFAULT 0", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"],
        "stock_destinations": ["destination_code TEXT", "notes TEXT", "is_active INTEGER NOT NULL DEFAULT 1", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"],
        "product_transfers_out": ["notes TEXT", "created_by INTEGER", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"],
        "production_reports": [
            "company_id INTEGER",
            "job_id INTEGER",
            "order_id INTEGER",
            "product_id INTEGER",
            "workstation_id INTEGER",
            "report_type TEXT",
            "quantity REAL NOT NULL DEFAULT 0",
            "notes TEXT",
            "reported_by INTEGER",
            "unit TEXT DEFAULT 'pcs'",
        ],
        "dashboard_layouts": [
            "page_key TEXT NOT NULL DEFAULT 'dashboard'",
            "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        ],
        "password_reset_tokens": ["used_at TEXT", "created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP"],
    }

    for table_name, column_sql_list in desired_columns.items():
        if not _table_exists(cursor, table_name):
            continue
        for column_sql in column_sql_list:
            column_name = column_sql.split()[0]
            if not _column_exists(cursor, table_name, column_name):
                _safe_add_column(cursor, table_name, column_sql)

    if _table_exists(cursor, 'items'):
        if _column_exists(cursor, 'items', 'measurement_unit') and _column_exists(cursor, 'items', 'unit'):
            cursor.execute("UPDATE items SET measurement_unit = COALESCE(measurement_unit, unit, 'pcs')")
            cursor.execute("UPDATE items SET unit = COALESCE(unit, measurement_unit, 'pcs')")
        elif _column_exists(cursor, 'items', 'measurement_unit'):
            cursor.execute("UPDATE items SET measurement_unit = COALESCE(measurement_unit, 'pcs')")
        elif _column_exists(cursor, 'items', 'unit'):
            cursor.execute("UPDATE items SET unit = COALESCE(unit, 'pcs')")

    index_statements = [
        "CREATE INDEX IF NOT EXISTS idx_products_company_group ON products(company_id, group_id)",
        "CREATE INDEX IF NOT EXISTS idx_batch_orders_batch_company ON batch_orders(batch_id, company_id)",
        "CREATE INDEX IF NOT EXISTS idx_pjt_product_company ON product_job_templates(product_id, company_id)",
        "CREATE INDEX IF NOT EXISTS idx_bom_product_company ON bom(product_id, company_id)",
        "CREATE INDEX IF NOT EXISTS idx_order_jobs_company_workstation ON order_jobs(company_id, workstation_id)",
        "CREATE INDEX IF NOT EXISTS idx_orders_company_status ON orders(company_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_purchase_requests_company_status ON purchase_requests(company_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_shortages_company_item_order ON shortages(company_id, item_id, order_id)",
    ]

    for sql in index_statements:
        try:
            cursor.execute(sql)
        except Exception:
            pass

    conn.commit()
    conn.close()


def ensure_users_is_active_schema():
    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT 1 FROM users LIMIT 1")
    except Exception:
        conn.close()
        return

    try:
        cursor.execute("""
            ALTER TABLE users
            ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1
        """)
    except Exception:
        pass

    conn.commit()
    conn.close()









def ensure_workstation_groups_and_batch_delete_schema():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS workstation_groups (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            color TEXT NOT NULL DEFAULT '#6366f1',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(company_id, name),
            FOREIGN KEY (company_id) REFERENCES companies(id)
        )
    """)

    alter_statements = [
        "ALTER TABLE workstations ADD COLUMN group_id INTEGER",
        "ALTER TABLE order_batches ADD COLUMN is_deleted INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE workstations ADD COLUMN cost_per_hour REAL NOT NULL DEFAULT 0",
    ]

    for sql in alter_statements:
        try:
            cursor.execute(sql)
        except Exception:
            pass

    conn.commit()
    conn.close()

def seed_data():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM orders")
    count = cursor.fetchone()[0]

    if count == 0:
        cursor.execute("SELECT COUNT(*) FROM companies")
        company_count = cursor.fetchone()[0]

        if company_count == 1:
            cursor.execute("SELECT id FROM companies LIMIT 1")
            company_id = cursor.fetchone()[0]

            cursor.execute("""
                INSERT INTO orders (order_number, customer, status, due_date, priority, company_id)
                VALUES
                ('ORD-1001', 'NordSteel', 'In Progress', '2026-03-25', 'High', ?),
                ('ORD-1002', 'Baltic Frame', 'Waiting', '2026-03-28', 'Medium', ?),
                ('ORD-1003', 'MetalWorks LT', 'Completed', '2026-03-20', 'Low', ?)
            """, (company_id, company_id, company_id))

    conn.commit()
    conn.close()


def ensure_database_ready():
    init_db()
    

ensure_database_ready()
# ensure_render_safe_schema()
# ensure_users_is_active_schema()
# ensure_workstation_groups_and_batch_delete_schema()


def is_logged_in():
    return "user_id" in session

def get_company_id():
    company_id = session.get("company_id")
    if not company_id:
        raise ValueError("Missing company_id in session.")
    return company_id

def get_user_role():
    return session.get("user_role", "user")


def ensure_logged_in():
    if not is_logged_in():
        return redirect(url_for("login"))
    return None


def fetch_one(cursor, query, params=()):
    cursor.execute(query, params)
    return cursor.fetchone()


def company_scope_condition(alias=None):
    column = "company_id" if not alias else f"{alias}.company_id"
    return f"{column} = ?"


def company_params():
    company_id = get_company_id()
    if not company_id:
        raise ValueError("Missing company_id in session.")
    return (company_id,)


def record_belongs_to_company(cursor, table_name, record_id, company_id):
    query = f"SELECT 1 FROM {table_name} WHERE id = ? AND company_id = ?"
    cursor.execute(query, (record_id, company_id))
    return cursor.fetchone() is not None


def require_company_record(cursor, table_name, record_id, company_id, not_found_message="Record not found."):
    if not record_belongs_to_company(cursor, table_name, record_id, company_id):
        raise ValueError(not_found_message)


def redirect_back(default_endpoint="jobs"):
    return redirect(request.referrer or url_for(default_endpoint))

def calculate_job_total_hours(estimated_hours, planned_quantity, completed_quantity=0):
    estimated_hours = float(estimated_hours or 0)
    planned_quantity = float(planned_quantity or 0)
    completed_quantity = float(completed_quantity or 0)

    remaining_quantity = planned_quantity - completed_quantity
    if remaining_quantity < 0:
        remaining_quantity = 0

    return estimated_hours * remaining_quantity

def calculate_job_duration_days(total_job_hours, hours_per_shift, shifts_per_day):
    daily_capacity = float(hours_per_shift or 0) * float(shifts_per_day or 0)
    if daily_capacity <= 0:
        daily_capacity = 8
    if total_job_hours <= 0:
        return 1
    return max(1, math.ceil(total_job_hours / daily_capacity))








def recalculate_job_dates(cursor, job_id, planned_start=None):
    cursor.execute("""
        SELECT
            oj.id,
            oj.estimated_hours,
            oj.planned_quantity,
            oj.completed_quantity,
            oj.planned_start,
            w.hours_per_shift,
            w.shifts_per_day
        FROM order_jobs oj
        JOIN workstations w ON oj.workstation_id = w.id
        WHERE oj.id = ?
    """, (job_id,))
    row = cursor.fetchone()

    if row is None:
        return

    estimated_hours = float(row[1] or 0)
    planned_quantity = float(row[2] or 0)
    completed_quantity = float(row[3] or 0)
    current_planned_start = row[4]
    hours_per_shift = float(row[5] or 0)
    shifts_per_day = float(row[6] or 0)

    start_value = planned_start if planned_start is not None else current_planned_start

    if not start_value:
        cursor.execute("""
            UPDATE order_jobs
            SET planned_start = NULL, planned_end = NULL
            WHERE id = ?
        """, (job_id,))
        return

    total_job_hours = calculate_job_total_hours(
        estimated_hours,
        planned_quantity,
        completed_quantity
    )

    duration_days = calculate_job_duration_days(
        total_job_hours,
        hours_per_shift,
        shifts_per_day
    )

    start_date = datetime.strptime(start_value, "%Y-%m-%d").date()
    end_date = start_date + timedelta(days=duration_days - 1)

    cursor.execute("""
        UPDATE order_jobs
        SET planned_start = ?, planned_end = ?
        WHERE id = ?
    """, (
        start_date.isoformat(),
        end_date.isoformat(),
        job_id
    ))

def recalculate_company_planner_dates(cursor, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT id, planned_start
        FROM order_jobs
        WHERE company_id = ?
          AND planned_start IS NOT NULL
        ORDER BY id ASC
    """, (company_id,))
    rows = cursor.fetchall()

    for job_id, planned_start in rows:
        recalculate_job_dates(cursor, job_id, planned_start)


  


def build_month_days(year, month):
    days_in_month = calendar.monthrange(year, month)[1]
    month_days = []

    for day in range(1, days_in_month + 1):
        current_date = datetime(year, month, day).date()
        month_days.append({
            "day": day,
            "date": current_date.isoformat(),
            "weekday": current_date.strftime("%a")
        })

    return month_days







def generate_order_jobs_recursive(cursor, order_id, current_product_id, current_quantity, planned_date=None, path=None, company_id=None):
    if path is None:
        path = []

    if company_id is None:
        raise ValueError("company_id is required for job generation.")

    if current_product_id in path:
        raise ValueError("Circular BOM detected.")

    current_path = path + [current_product_id]

    cursor.execute("""
        SELECT child_product_id, quantity
        FROM bom
        WHERE product_id = ?
          AND company_id = ?
          AND component_type = 'product'
          AND child_product_id IS NOT NULL
        ORDER BY id ASC
    """, (current_product_id, company_id))
    child_rows = cursor.fetchall()

    for child_product_id, bom_quantity in child_rows:
        child_required_quantity = float(current_quantity) * float(bom_quantity)
        generate_order_jobs_recursive(
            cursor,
            order_id,
            child_product_id,
            child_required_quantity,
            planned_date,
            current_path,
            company_id
        )

    cursor.execute("""
        SELECT id, workstation_id, job_name, sequence, estimated_hours
        FROM product_job_templates
        WHERE product_id = ?
          AND company_id = ?
        ORDER BY sequence ASC, id ASC
    """, (current_product_id, company_id))
    templates = cursor.fetchall()

    for template_id, workstation_id, job_name, sequence, estimated_hours in templates:
        cursor.execute("""
            INSERT INTO order_jobs (
                order_id,
                product_job_template_id,
                job_product_id,
                workstation_id,
                job_name,
                sequence,
                planned_quantity,
                completed_quantity,
                estimated_hours,
                status,
                planned_start,
                planned_end,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            order_id,
            template_id,
            current_product_id,
            workstation_id,
            job_name,
            sequence,
            current_quantity,
            0,
            estimated_hours,
            "Waiting",
            planned_date,
            planned_date,
            company_id
        ))

        new_job_id = cursor.lastrowid
        recalculate_job_dates(cursor, new_job_id, planned_date)

def is_float_equal(a, b, tolerance=0.0001):
    return abs(float(a) - float(b)) <= tolerance


def job_has_split_children(cursor, job_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT 1
        FROM order_jobs
        WHERE parent_job_id = ?
          AND is_split_child = 1
          AND company_id = ?
        LIMIT 1
    """, (job_id, company_id))
    return cursor.fetchone() is not None



def sync_parent_job_status(cursor, parent_job_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT id, company_id
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (parent_job_id, company_id))
    parent = cursor.fetchone()

    if parent is None:
        return

    cursor.execute("""
        SELECT status, completed_quantity, planned_quantity
        FROM order_jobs
        WHERE parent_job_id = ?
          AND is_split_child = 1
          AND company_id = ?
    """, (parent_job_id, company_id))
    child_rows = cursor.fetchall()

    if not child_rows:
        return

    statuses = [row[0] for row in child_rows]
    all_done = all(status == "Done" for status in statuses)
    any_ongoing = any(status == "Ongoing" for status in statuses)
    any_paused = any(status == "Paused" for status in statuses)
    any_waiting = any(status == "Waiting" for status in statuses)

    total_completed = sum(float(row[1] or 0) for row in child_rows)
    total_planned = sum(float(row[2] or 0) for row in child_rows)

    if all_done:
        new_status = "Done"
    elif any_ongoing:
        new_status = "Ongoing"
    elif any_paused:
        new_status = "Paused"
    elif any_waiting:
        new_status = "Paused"
    else:
        new_status = "Paused"

    cursor.execute("""
        UPDATE order_jobs
        SET completed_quantity = ?, planned_quantity = ?, status = ?
        WHERE id = ?
          AND company_id = ?
    """, (total_completed, total_planned, new_status, parent_job_id, company_id))


def can_start_job(cursor, job_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT order_id, sequence, parent_job_id, is_split_child
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        return False

    order_id, sequence, parent_job_id, is_split_child = row

    if parent_job_id:
        cursor.execute("""
            SELECT status
            FROM order_jobs
            WHERE id = ?
              AND company_id = ?
        """, (parent_job_id, company_id))
        parent_row = cursor.fetchone()

        if parent_row and parent_row[0] not in ("Paused", "Ongoing", "Done"):
            return False

    cursor.execute("""
        SELECT COUNT(*)
        FROM order_jobs prev
        WHERE prev.order_id = ?
          AND prev.company_id = ?
          AND prev.sequence < ?
          AND prev.status != 'Done'
          AND (
                prev.is_split_child = 1
                OR NOT EXISTS (
                    SELECT 1
                    FROM order_jobs child
                    WHERE child.parent_job_id = prev.id
                      AND child.is_split_child = 1
                      AND child.company_id = prev.company_id
                )
          )
    """, (order_id, company_id, sequence))
    remaining = cursor.fetchone()[0]

    return remaining == 0


def sync_order_status(cursor, order_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT status
        FROM order_jobs
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))
    rows = cursor.fetchall()

    if not rows:
        cursor.execute("""
            UPDATE orders
            SET status = 'Waiting'
            WHERE id = ?
              AND company_id = ?
        """, (order_id, company_id))
        return

    statuses = [row[0] for row in rows]

    if all(status == "Done" for status in statuses):
        order_status = "Completed"
    elif any(status in ("Ongoing", "Paused") for status in statuses):
        order_status = "In Progress"
    else:
        order_status = "Waiting"

    cursor.execute("""
        UPDATE orders
        SET status = ?
        WHERE id = ?
          AND company_id = ?
    """, (order_status, order_id, company_id))


def create_split_children(cursor, parent_job_id, split_rows, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT
            id,
            order_id,
            product_job_template_id,
            job_product_id,
            workstation_id,
            job_name,
            sequence,
            planned_quantity,
            completed_quantity,
            estimated_hours,
            status,
            planned_start,
            planned_end,
            parent_job_id,
            is_split_child
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (parent_job_id, company_id))
    parent = cursor.fetchone()

    if parent is None:
        raise ValueError("Parent job not found.")

    if int(parent[14] or 0) == 1:
        raise ValueError("Split child job cannot be split again.")

    if float(parent[8] or 0) > 0:
        raise ValueError("Cannot split job that already has completed quantity.")

    if job_has_split_children(cursor, parent_job_id, company_id=company_id):
        raise ValueError("Job is already split.")

    parent_planned_quantity = float(parent[7] or 0)
    total_split_quantity = sum(float(row["quantity"]) for row in split_rows)

    if parent_planned_quantity <= 0:
        raise ValueError("Parent job quantity must be greater than 0.")

    if not is_float_equal(total_split_quantity, parent_planned_quantity):
        raise ValueError("Split quantities must match original planned quantity.")

    parent_planned_start = parent[11]
    parent_estimated_hours = float(parent[9] or 0)

    for row in split_rows:
        workstation_id = int(row["workstation_id"])
        split_quantity = float(row["quantity"])

        require_company_record(cursor, "workstations", workstation_id, company_id, "Workstation not found.")

        split_ratio = split_quantity / parent_planned_quantity
        child_estimated_hours = parent_estimated_hours * split_ratio

        cursor.execute("""
            INSERT INTO order_jobs (
                order_id,
                product_job_template_id,
                job_product_id,
                workstation_id,
                job_name,
                sequence,
                planned_quantity,
                completed_quantity,
                estimated_hours,
                status,
                planned_start,
                planned_end,
                parent_job_id,
                is_split_child,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            parent[1],
            parent[2],
            parent[3],
            workstation_id,
            parent[5],
            parent[6],
            split_quantity,
            0,
            child_estimated_hours,
            "Waiting",
            parent_planned_start,
            parent_planned_start,
            parent_job_id,
            1,
            company_id
        ))

        new_child_id = cursor.lastrowid
        recalculate_job_dates(cursor, new_child_id, parent_planned_start)

    cursor.execute("""
        UPDATE order_jobs
        SET status = 'Paused',
            planned_start = NULL,
            planned_end = NULL
        WHERE id = ?
          AND company_id = ?
    """, (parent_job_id, company_id))

    sync_parent_job_status(cursor, parent_job_id, company_id=company_id)


def rebuild_order_jobs(cursor, order_id, root_product_id, root_quantity, planned_date=None, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        DELETE FROM order_jobs
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))

    generate_order_jobs_recursive(
        cursor,
        order_id,
        root_product_id,
        root_quantity,
        planned_date,
        company_id=company_id
    )

def get_active_jobs_for_reports(cursor, company_id):
    cursor.execute("""
        SELECT
            oj.id,
            o.order_number,
            COALESCE(p.product_name, '-') AS product_name,
            oj.job_name,
            COALESCE(w.name, '-') AS workstation_name,
            oj.status,
            oj.planned_quantity,
            oj.completed_quantity
        FROM order_jobs oj
        JOIN orders o
          ON oj.order_id = o.id
         AND o.company_id = oj.company_id
        LEFT JOIN products p
          ON oj.job_product_id = p.id
         AND p.company_id = oj.company_id
        LEFT JOIN workstations w
          ON oj.workstation_id = w.id
         AND w.company_id = oj.company_id
        WHERE oj.company_id = ?
          AND oj.status IN ('Waiting', 'Ongoing', 'Paused', 'Delayed')
          AND (
                oj.is_split_child = 1
                OR NOT EXISTS (
                    SELECT 1
                    FROM order_jobs child
                    WHERE child.parent_job_id = oj.id
                      AND child.is_split_child = 1
                      AND child.company_id = oj.company_id
                )
          )
        ORDER BY o.order_number ASC, oj.sequence ASC, oj.id ASC
    """, (company_id,))

    rows = cursor.fetchall()

    active_jobs = []
    for row in rows:
        active_jobs.append({
            "id": row[0],
            "order_number": row[1],
            "product_name": row[2],
            "job_name": row[3],
            "workstation_name": row[4],
            "status": row[5],
            "planned_quantity": float(row[6] or 0),
            "completed_quantity": float(row[7] or 0),
        })

    return active_jobs


def hash_reset_token(raw_token):
    return hashlib.sha256(raw_token.encode("utf-8")).hexdigest()


def build_reset_link(raw_token):
    base_url = os.environ.get("APP_BASE_URL", "").strip()

    if base_url:
        return f"{base_url.rstrip('/')}{url_for('reset_password', token=raw_token)}"

    return url_for("reset_password", token=raw_token, _external=True)


import resend

resend.api_key = os.environ.get("RESEND_API_KEY")


def send_password_reset_email(recipient_email, reset_link):
    resend.Emails.send({
        "from": "UMS <onboarding@resend.dev>",
        "to": recipient_email,
        "subject": "Reset your UMS password",
        "html": f"""
        <h2>Reset your password</h2>
        <p>Click below:</p>
        <a href="{reset_link}">Reset Password</a>
        <p>Expires in 45 minutes.</p>
        """
    })
    

BACKUP_DIR = os.path.join(os.path.dirname(__file__), "backups")
BACKUP_RETENTION = 20


def ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


def prune_old_backups(retention=BACKUP_RETENTION):
    ensure_backup_dir()

    backup_files = sorted(
        [
            os.path.join(BACKUP_DIR, name)
            for name in os.listdir(BACKUP_DIR)
            if name.endswith(".db") or name.endswith(".sql")
        ],
        key=lambda path: os.path.getmtime(path),
        reverse=True
    )

    for old_file in backup_files[retention:]:
        try:
            os.remove(old_file)
        except OSError:
            pass


def create_database_backup(reason="manual"):
    ensure_backup_dir()
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if getattr(sqlite3, "USE_POSTGRES", False):
        backup_path = os.path.join(BACKUP_DIR, f"ums_{reason}_{timestamp}.sql")
        conn = get_connection()
        cursor = conn.cursor()
        table_names = [
            "companies", "users", "user_permissions", "suppliers", "items", "product_groups",
            "products", "workstation_groups", "workstations", "orders", "order_batches",
            "batch_orders", "bom", "product_job_templates", "order_jobs", "purchase_requests",
            "shortages", "stock_destinations", "product_transfers_out", "production_reports",
            "dashboard_layouts", "password_reset_tokens",
        ]
        statements = [
            "-- UMS PostgreSQL logical backup",
            f"-- generated_at={datetime.utcnow().isoformat()}Z",
            "BEGIN;",
        ]
        try:
            for table_name in table_names:
                if not _table_exists(cursor, table_name):
                    continue
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()
                columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in columns_info]
                if not columns:
                    continue
                quoted_columns = ", ".join(f'"{col}"' for col in columns)
                cursor.execute(f"SELECT {quoted_columns} FROM {table_name}")
                rows = cursor.fetchall()
                for row in rows:
                    values = []
                    for value in row:
                        if value is None:
                            values.append("NULL")
                        elif isinstance(value, bool):
                            values.append("TRUE" if value else "FALSE")
                        elif isinstance(value, (int, float)):
                            values.append(str(value))
                        else:
                            escaped = str(value).replace("'", "''")
                            values.append(f"'{escaped}'")
                    statements.append(
                        f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({", ".join(values)});'
                    )
            statements.append("COMMIT;")
            with open(backup_path, "w", encoding="utf-8") as handle:
                handle.write("\n".join(statements) + "\n")
        finally:
            conn.close()
        prune_old_backups()
        return backup_path

    backup_path = os.path.join(BACKUP_DIR, f"ums_{reason}_{timestamp}.db")
    source_conn = get_connection()
    backup_conn = sqlite3.connect(backup_path)
    try:
        source_conn.backup(backup_conn)
    finally:
        backup_conn.close()
        source_conn.close()
    prune_old_backups()
    return backup_path


def create_sqlite_backup(reason="manual"):
    return create_database_backup(reason)


def _format_export_value(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 2)
    return value


def build_excel_bytes(sheet_definitions):
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="E5ECFF")
    thin = Side(style="thin", color="334155")

    for sheet_name, columns, rows in sheet_definitions:
        ws = workbook.create_sheet(title=sheet_name[:31])
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)

        for row_idx, row_values in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_format_export_value(value))
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx, column_name in enumerate(columns, start=1):
            max_len = len(str(column_name))
            for row in rows:
                value = row[col_idx - 1] if col_idx - 1 < len(row) else ""
                max_len = max(max_len, len(str(_format_export_value(value))))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max(max_len + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def excel_response(filename, sheet_definitions):
    output = build_excel_bytes(sheet_definitions)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def pdf_response(filename, title, columns, rows, subtitle=""):
    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm
    )

    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"])]

    if subtitle:
        story.append(Paragraph(subtitle, styles["BodyText"]))
        story.append(Spacer(1, 6))

    table_data = [columns]
    for row in rows:
        table_data.append([str(_format_export_value(value)) for value in row])

    available_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
    col_width = available_width / max(1, len(columns))

    table = Table(table_data, repeatRows=1, colWidths=[col_width] * len(columns))
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2A44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F1F5F9")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(table)
    document.build(story)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


def fetch_orders_export_rows(company_id, args):
    order_number = args.get("order_number", "").strip()
    product_name = args.get("product_name", "").strip()
    statuses = args.getlist("status")
    priority = args.get("priority", "").strip()
    due_date_from = args.get("due_date_from", "").strip()
    due_date_to = args.get("due_date_to", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            o.order_number,
            COALESCE(p.product_name, '-'),
            o.quantity,
            o.status,
            o.priority,
            o.due_date
        FROM orders o
        LEFT JOIN products p
          ON o.product_id = p.id
         AND p.company_id = o.company_id
        WHERE o.company_id = ?
    """
    params = [company_id]

    if order_number:
        query += " AND o.order_number LIKE ?"
        params.append(f"%{order_number}%")

    if product_name:
        query += " AND p.product_name LIKE ?"
        params.append(f"%{product_name}%")

    if statuses and "All" not in statuses:
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND o.status IN ({placeholders})"
        params.extend(statuses)

    if priority:
        query += " AND o.priority = ?"
        params.append(priority)

    if due_date_from:
        query += " AND o.due_date >= ?"
        params.append(due_date_from)

    if due_date_to:
        query += " AND o.due_date <= ?"
        params.append(due_date_to)

    query += " ORDER BY o.id DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_jobs_export_rows(company_id, args):
    order_number = args.get("order_number", "").strip()
    product_name = args.get("product_name", "").strip()
    job_name = args.get("job_name", "").strip()
    workstation = args.get("workstation", "").strip()
    workstation_text = args.get("workstation_text", "").strip()
    due_date_from = args.get("due_date_from", "").strip()
    due_date_to = args.get("due_date_to", "").strip()

    statuses = [s.strip() for s in args.getlist("status") if s.strip()]
    if "All" in statuses:
        statuses = []

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            o.order_number,
            COALESCE(jp.product_name, '-'),
            oj.job_name,
            w.name,
            oj.sequence,
            oj.planned_quantity,
            oj.completed_quantity,
            oj.estimated_hours,
            oj.status,
            o.due_date,
            oj.planned_start,
            oj.planned_end
        FROM order_jobs oj
        JOIN orders o
          ON oj.order_id = o.id
         AND o.company_id = oj.company_id
        LEFT JOIN products jp
          ON oj.job_product_id = jp.id
         AND jp.company_id = oj.company_id
        JOIN workstations w
          ON oj.workstation_id = w.id
         AND w.company_id = oj.company_id
        WHERE oj.company_id = ?
    """
    params = [company_id]

    if order_number:
        query += " AND o.order_number LIKE ?"
        params.append(f"%{order_number}%")

    if product_name:
        query += " AND jp.product_name LIKE ?"
        params.append(f"%{product_name}%")

    if job_name:
        query += " AND oj.job_name LIKE ?"
        params.append(f"%{job_name}%")

    if workstation:
        query += " AND oj.workstation_id = ?"
        params.append(workstation)

    if workstation_text:
        query += " AND w.name LIKE ?"
        params.append(f"%{workstation_text}%")

    if statuses:
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND oj.status IN ({placeholders})"
        params.extend(statuses)

    if due_date_from:
        query += " AND o.due_date >= ?"
        params.append(due_date_from)

    if due_date_to:
        query += " AND o.due_date <= ?"
        params.append(due_date_to)

    query += " ORDER BY o.due_date ASC, o.order_number ASC, oj.sequence ASC, oj.id ASC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_suppliers_export_rows(company_id, args):
    search = args.get("search", "").strip()
    status_filter = args.get("status", "").strip().lower()

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT
            name,
            supplier_code,
            contact_person,
            email,
            phone,
            address,
            CASE WHEN COALESCE(is_active, 0) = 1 THEN 'Active' ELSE 'Inactive' END,
            created_at
        FROM suppliers
        WHERE company_id = ?
    """
    params = [company_id]

    if search:
        like_value = f"%{search.lower()}%"
        sql += """
          AND (
                LOWER(COALESCE(name, '')) LIKE ?
             OR LOWER(COALESCE(supplier_code, '')) LIKE ?
             OR LOWER(COALESCE(contact_person, '')) LIKE ?
             OR LOWER(COALESCE(email, '')) LIKE ?
             OR LOWER(COALESCE(phone, '')) LIKE ?
          )
        """
        params.extend([like_value] * 5)

    if status_filter == "active":
        sql += " AND COALESCE(is_active, 0) = 1"
    elif status_filter == "inactive":
        sql += " AND COALESCE(is_active, 0) = 0"

    sql += " ORDER BY created_at DESC, id DESC"
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_shortage_export_rows(company_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            i.item_code,
            i.item_name,
            i.measurement_unit,
            COALESCE(i.stock_quantity, 0),
            COALESCE(i.min_stock, 0),
            MAX(0, COALESCE(i.min_stock, 0) - COALESCE(i.stock_quantity, 0)),
            COALESCE(s.name, '')
        FROM items i
        LEFT JOIN suppliers s
          ON i.supplier_id = s.id
         AND s.company_id = i.company_id
        WHERE i.company_id = ?
          AND COALESCE(i.stock_quantity, 0) < COALESCE(i.min_stock, 0)
        ORDER BY i.item_name ASC
    """, (company_id,))

    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_reports_export_rows(company_id, args):
    report_type = args.get("report_type", "").strip()
    job_search = args.get("job_search", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            pr.report_type,
            pr.quantity,
            COALESCE(pr.unit, 'pcs'),
            pr.notes,
            pr.created_at,
            oj.job_name,
            oj.status,
            COALESCE(o.order_number, '-'),
            COALESCE(p.product_name, '-'),
            COALESCE(w.name, '-'),
            COALESCE(u.full_name, 'System')
        FROM production_reports pr
        JOIN order_jobs oj
          ON pr.job_id = oj.id
         AND oj.company_id = pr.company_id
        LEFT JOIN orders o
          ON pr.order_id = o.id
         AND o.company_id = pr.company_id
        LEFT JOIN products p
          ON pr.product_id = p.id
         AND p.company_id = pr.company_id
        LEFT JOIN workstations w
          ON pr.workstation_id = w.id
         AND w.company_id = pr.company_id
        LEFT JOIN users u
          ON pr.reported_by = u.id
        WHERE pr.company_id = ?
    """
    params = [company_id]

    if report_type:
        query += " AND pr.report_type = ?"
        params.append(report_type)

    if job_search:
        like_value = f"%{job_search}%"
        query += """
          AND (
                o.order_number LIKE ?
             OR oj.job_name LIKE ?
             OR p.product_name LIKE ?
             OR w.name LIKE ?
          )
        """
        params.extend([like_value, like_value, like_value, like_value])

    query += " ORDER BY pr.created_at DESC, pr.id DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_purchase_request_export_rows(company_id, args, history=False):
    search_key = "history_search" if history else "search"
    status_key = "history_status" if history else "status"
    priority_key = "history_priority" if history else "priority"
    supplier_key = "history_supplier_id" if history else "supplier_id"

    search = args.get(search_key, "").strip()
    status_filter = args.get(status_key, "").strip().lower()
    priority_filter = args.get(priority_key, "").strip().lower()
    supplier_filter = args.get(supplier_key, "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT
            COALESCE(pr.request_number, 'PR-' || pr.id),
            pr.title,
            COALESCE(i.item_name, ''),
            COALESCE(s.name, ''),
            pr.quantity,
            COALESCE(pr.unit, ''),
            COALESCE(pr.priority, 'normal'),
            COALESCE(pr.status, 'draft'),
            COALESCE(pr.needed_by, ''),
            COALESCE(u.full_name, ''),
            COALESCE(pr.created_at, ''),
            COALESCE(pr.updated_at, ''),
            COALESCE(pr.notes, '')
        FROM purchase_requests pr
        LEFT JOIN suppliers s
          ON pr.supplier_id = s.id
         AND s.company_id = pr.company_id
        LEFT JOIN users u
          ON pr.requested_by = u.id
         AND u.company_id = pr.company_id
        LEFT JOIN items i
          ON pr.item_id = i.id
         AND i.company_id = pr.company_id
        WHERE pr.company_id = ?
    """
    params = [company_id]

    if history:
        sql += " AND COALESCE(pr.status, 'draft') IN ('received', 'cancelled', 'rejected')"
    else:
        sql += " AND COALESCE(pr.status, 'draft') NOT IN ('received', 'cancelled', 'rejected')"

    if search:
        like_value = f"%{search.lower()}%"
        sql += """
          AND (
                LOWER(COALESCE(pr.request_number, '')) LIKE ?
             OR LOWER(COALESCE(pr.title, '')) LIKE ?
             OR LOWER(COALESCE(pr.description, '')) LIKE ?
             OR LOWER(COALESCE(s.name, '')) LIKE ?
             OR LOWER(COALESCE(i.item_name, '')) LIKE ?
          )
        """
        params.extend([like_value, like_value, like_value, like_value, like_value])

    valid_statuses = {"draft", "submitted", "approved", "ordered"} if not history else {"received", "cancelled", "rejected"}
    if status_filter in valid_statuses:
        sql += " AND LOWER(COALESCE(pr.status, 'draft')) = ?"
        params.append(status_filter)

    if priority_filter in ("low", "normal", "high"):
        sql += " AND LOWER(COALESCE(pr.priority, 'normal')) = ?"
        params.append(priority_filter)

    if supplier_filter:
        try:
            sql += " AND pr.supplier_id = ?"
            params.append(int(supplier_filter))
        except ValueError:
            pass

    sql += " ORDER BY COALESCE(pr.updated_at, pr.created_at) DESC, pr.id DESC"
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_inventory_export_rows(company_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            'Item',
            item_code,
            item_name,
            measurement_unit,
            stock_quantity,
            unit_price,
            COALESCE(stock_quantity, 0) * COALESCE(unit_price, 0)
        FROM items
        WHERE company_id = ?

        UNION ALL

        SELECT
            'Product',
            product_code,
            product_name,
            measurement_unit,
            stock_quantity,
            0,
            0
        FROM products
        WHERE company_id = ?

        ORDER BY 1, 3
    """, (company_id, company_id))

    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_products_export_rows(company_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            product_code,
            product_name,
            description,
            measurement_unit,
            time_per_unit,
            stock_quantity
        FROM products
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,))

    rows = cursor.fetchall()
    conn.close()
    return rows


def fetch_items_export_rows(company_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            item_code,
            item_name,
            description,
            measurement_unit,
            unit_price,
            stock_quantity,
            min_stock
        FROM items
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,))

    rows = cursor.fetchall()
    conn.close()
    return rows


def get_order_root_product_and_quantity(cursor, order_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT product_id, quantity, materials_consumed, finished_stock_added
        FROM orders
        WHERE id = ?
          AND company_id = ?
    """, (order_id, company_id))
    row = cursor.fetchone()

    if row is None:
        return None

    return {
        "product_id": row[0],
        "quantity": float(row[1] or 0),
        "materials_consumed": int(row[2] or 0),
        "finished_stock_added": int(row[3] or 0)
    }

def fetch_order_material_rows(company_id, order_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            o.order_number,
            o.quantity,
            p.id,
            p.product_name
        FROM orders o
        JOIN products p
          ON o.product_id = p.id
         AND p.company_id = o.company_id
        WHERE o.id = ? AND o.company_id = ?
    """, (order_id, company_id))

    order = cursor.fetchone()

    if not order:
        conn.close()
        return None, []

    product_id = order[2]
    order_quantity = float(order[1] or 0)

    per_unit_exploded = explode_bom_items_recursive(cursor, product_id, 1, company_id=company_id)
    total_exploded = explode_bom_items_recursive(cursor, product_id, order_quantity, company_id=company_id)

    rows = []

    if total_exploded:
        item_ids = list(total_exploded.keys())
        placeholders = ",".join(["?"] * len(item_ids))

        cursor.execute(f"""
            SELECT id, item_name, item_code, measurement_unit, unit_price
            FROM items
            WHERE id IN ({placeholders}) AND company_id = ?
            ORDER BY item_name ASC
        """, item_ids + [company_id])

        for item_id, item_name, item_code, measurement_unit, unit_price in cursor.fetchall():
            bom_quantity = float(per_unit_exploded.get(item_id, 0) or 0)
            total_quantity = float(total_exploded.get(item_id, 0) or 0)
            unit_price = float(unit_price or 0)

            rows.append((
                item_code,
                item_name,
                measurement_unit,
                bom_quantity,
                total_quantity,
                unit_price,
                total_quantity * unit_price
            ))

    conn.close()

    return {
        "order_number": order[0],
        "quantity": order_quantity,
        "product_name": order[3]
    }, rows


def fetch_planner_export_rows(company_id, year, month):
    month_start = f"{year:04d}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    month_end = f"{year:04d}-{month:02d}-{last_day:02d}"

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            w.id,
            w.name
        FROM workstations w
        WHERE w.company_id = ?
        ORDER BY w.name ASC
    """, (company_id,))
    workstation_map = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.execute("""
        SELECT
            oj.id,
            oj.workstation_id,
            oj.job_name,
            oj.status,
            oj.planned_start,
            oj.planned_end,
            oj.completed_quantity,
            oj.planned_quantity,
            oj.estimated_hours,
            o.order_number,
            COALESCE(p.product_name, '-')
        FROM order_jobs oj
        JOIN orders o
          ON oj.order_id = o.id
         AND o.company_id = oj.company_id
        LEFT JOIN products p
          ON oj.job_product_id = p.id
         AND p.company_id = oj.company_id
        WHERE oj.company_id = ?
          AND (
                oj.is_split_child = 1
                OR NOT EXISTS (
                    SELECT 1
                    FROM order_jobs child
                    WHERE child.parent_job_id = oj.id
                      AND child.is_split_child = 1
                      AND child.company_id = oj.company_id
                )
          )
        ORDER BY o.order_number ASC, oj.sequence ASC, oj.id ASC
    """, (company_id,))

    all_rows = cursor.fetchall()
    conn.close()

    scheduled = []
    unscheduled = []

    for row in all_rows:
        export_row = (
            row[9],
            row[10],
            row[2],
            workstation_map.get(row[1], ""),
            row[3],
            row[4] or "",
            row[5] or "",
            row[7],
            row[6],
            row[8]
        )

        if not row[4] or not row[5]:
            unscheduled.append(export_row)
            continue

        if row[5] < month_start or row[4] > month_end:
            continue

        scheduled.append(export_row)

    return scheduled, unscheduled


def explode_bom_items_recursive(cursor, product_id, required_quantity, collected=None, path=None, company_id=None):
    if collected is None:
        collected = {}

    if path is None:
        path = []

    if company_id is None:
        raise ValueError("company_id is required.")

    if product_id in path:
        raise ValueError("Circular BOM detected.")

    current_path = path + [product_id]

    cursor.execute("""
        SELECT
            component_type,
            item_id,
            child_product_id,
            quantity
        FROM bom
        WHERE product_id = ?
          AND company_id = ?
        ORDER BY id ASC
    """, (product_id, company_id))
    rows = cursor.fetchall()

    for component_type, item_id, child_product_id, bom_quantity in rows:
        bom_quantity = float(bom_quantity or 0)
        total_required = float(required_quantity) * bom_quantity

        if component_type == "product" and child_product_id:
            explode_bom_items_recursive(
                cursor,
                child_product_id,
                total_required,
                collected,
                current_path,
                company_id
            )
        else:
            if item_id not in collected:
                collected[item_id] = 0
            collected[item_id] += total_required

    return collected


def generate_batch_number(cursor, company_id):
    cursor.execute("""
        SELECT COUNT(*)
        FROM order_batches
        WHERE company_id = ?
    """, (company_id,))
    count = int(cursor.fetchone()[0] or 0) + 1
    return f"BAT-{count:05d}"


def fetch_batch_with_stats(cursor, batch_id, company_id):
    cursor.execute("""
        SELECT
            b.id,
            b.batch_number,
            b.name,
            b.status,
            b.notes,
            b.created_at,
            b.launched_at,
            COALESCE(COUNT(bo.id), 0) AS order_count,
            COALESCE(SUM(bo.quantity), 0) AS total_quantity
        FROM order_batches b
        LEFT JOIN batch_orders bo
          ON bo.batch_id = b.id
         AND bo.company_id = b.company_id
        WHERE b.id = ?
          AND b.company_id = ?
          AND COALESCE(b.is_deleted, 0) = 0
        GROUP BY b.id, b.batch_number, b.name, b.status, b.notes, b.created_at, b.launched_at
    """, (batch_id, company_id))
    row = cursor.fetchone()

    if row is None:
        return None

    return {
        "id": row[0],
        "batch_number": row[1] or f"BAT-{row[0]:05d}",
        "name": row[2],
        "status": row[3],
        "notes": row[4] or "",
        "created_at": row[5],
        "launched_at": row[6],
        "order_count": int(row[7] or 0),
        "total_quantity": float(row[8] or 0),
    }


def get_batch_material_requirements(cursor, batch_id, company_id):
    cursor.execute("""
        SELECT
            bo.id,
            bo.product_id,
            bo.quantity
        FROM batch_orders bo
        JOIN order_batches b
          ON b.id = bo.batch_id
         AND b.company_id = bo.company_id
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        ORDER BY bo.id ASC
    """, (batch_id, company_id))
    batch_orders_rows = cursor.fetchall()

    total_requirements = {}

    for batch_order_id, product_id, quantity in batch_orders_rows:
        exploded = explode_bom_items_recursive(
            cursor,
            product_id,
            float(quantity or 0),
            company_id=company_id
        )

        for item_id, required_qty in exploded.items():
            if item_id not in total_requirements:
                total_requirements[item_id] = 0
            total_requirements[item_id] += float(required_qty or 0)

    if not total_requirements:
        return []

    item_ids = list(total_requirements.keys())
    placeholders = ",".join(["?"] * len(item_ids))

    cursor.execute(f"""
        SELECT
            i.id,
            i.item_code,
            i.item_name,
            i.measurement_unit,
            COALESCE(i.stock_quantity, 0),
            COALESCE(i.reserved_quantity, 0),
            COALESCE(i.available_quantity, COALESCE(i.stock_quantity, 0) - COALESCE(i.reserved_quantity, 0)),
            i.supplier_id,
            COALESCE(s.name, '')
        FROM items i
        LEFT JOIN suppliers s
          ON s.id = i.supplier_id
         AND s.company_id = i.company_id
        WHERE i.id IN ({placeholders})
          AND i.company_id = ?
        ORDER BY i.item_name ASC
    """, item_ids + [company_id])

    item_rows = cursor.fetchall()
    materials = []

    for row in item_rows:
        item_id = row[0]
        required_qty = float(total_requirements.get(item_id, 0) or 0)
        stock_qty = float(row[4] or 0)
        reserved_qty = float(row[5] or 0)
        available_qty = float(row[6] or 0)
        shortage_qty = max(0, required_qty - available_qty)

        materials.append({
            "item_id": item_id,
            "item_code": row[1],
            "item_name": row[2],
            "measurement_unit": row[3] or "pcs",
            "required_qty": required_qty,
            "stock_qty": stock_qty,
            "reserved_qty": reserved_qty,
            "available_qty": available_qty,
            "shortage_qty": shortage_qty,
            "supplier_id": row[7],
            "supplier_name": row[8] or "",
        })

    return materials


def create_batch_purchase_requests(cursor, batch_id, company_id, user_id):
    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        raise ValueError("Batch not found.")

    if batch["status"] != "draft":
        raise ValueError("Only draft batches can create purchase requests.")

    materials = get_batch_material_requirements(cursor, batch_id, company_id)
    created_count = 0

    for material in materials:
        shortage_qty = float(material["shortage_qty"] or 0)
        if shortage_qty <= 0:
            continue

        cursor.execute("""
            SELECT id
            FROM purchase_requests
            WHERE company_id = ?
              AND item_id = ?
              AND source_type = 'batch'
              AND source_batch_id = ?
              AND COALESCE(status, 'draft') IN ('draft', 'ordered', 'waiting_receiving')
            LIMIT 1
        """, (company_id, material["item_id"], batch_id))
        existing = cursor.fetchone()

        if existing:
            continue

        title = f"Batch material need — {material['item_name']}"
        description = (
            f"Created from batch {batch['batch_number']} / {batch['name']}. "
            f"Required for planned batch execution."
        )

        cursor.execute("""
            INSERT INTO purchase_requests (
                company_id,
                item_id,
                supplier_id,
                title,
                description,
                quantity,
                unit,
                status,
                priority,
                requested_by,
                notes,
                source_type,
                source_batch_id,
                source_batch_order_id,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            company_id,
            material["item_id"],
            material["supplier_id"],
            title,
            description,
            shortage_qty,
            material["measurement_unit"],
            "draft",
            "high",
            user_id,
            f"Generated from batch {batch['batch_number']}.",
            "batch",
            batch_id,
            None
        ))

        new_request_id = cursor.lastrowid
        request_number = f"{batch['batch_number']}-PR-{new_request_id:03d}"

        cursor.execute("""
            UPDATE purchase_requests
            SET request_number = ?
            WHERE id = ?
              AND company_id = ?
        """, (request_number, new_request_id, company_id))

        created_count += 1

    return created_count


def launch_batch_orders(cursor, batch_id, company_id):
    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        raise ValueError("Batch not found.")

    if batch["status"] != "draft":
        raise ValueError("Only draft batches can be launched.")

    cursor.execute("""
        SELECT
            bo.id,
            bo.order_number,
            bo.product_id,
            bo.quantity,
            bo.due_date,
            bo.priority
        FROM batch_orders bo
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        ORDER BY bo.id ASC
    """, (batch_id, company_id))
    rows = cursor.fetchall()

    if not rows:
        raise ValueError("Batch has no orders to launch.")

    launched_count = 0

    for row in rows:
        batch_order_id, order_number, product_id, quantity, due_date, priority = row
        quantity = float(quantity or 0)

        if quantity <= 0:
            continue

        cursor.execute("""
            INSERT INTO orders (
                order_number,
                product_id,
                quantity,
                status,
                due_date,
                priority,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            order_number,
            product_id,
            quantity,
            "Waiting",
            due_date,
            priority or "Medium",
            company_id
        ))

        order_id = cursor.lastrowid

        create_order_shortages_and_reservations(
            cursor,
            order_id,
            product_id,
            quantity,
            company_id=company_id
        )

        generate_order_jobs_recursive(
            cursor,
            order_id,
            product_id,
            quantity,
            planned_date=None,
            company_id=company_id
        )

        cursor.execute("""
            UPDATE batch_orders
            SET launched_order_id = ?
            WHERE id = ?
              AND company_id = ?
        """, (order_id, batch_order_id, company_id))

        launched_count += 1

    cursor.execute("""
        UPDATE order_batches
        SET status = 'launched',
            launched_at = CURRENT_TIMESTAMP
        WHERE id = ?
          AND company_id = ?
    """, (batch_id, company_id))

    rebuild_company_reserved_quantities(cursor, company_id=company_id)

    return launched_count



def sync_item_available(cursor, item_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        UPDATE items
        SET available_quantity = COALESCE(stock_quantity, 0) - COALESCE(reserved_quantity, 0)
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))






def recalculate_shortages_for_item(cursor, item_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT COALESCE(stock_quantity, 0)
        FROM items
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))
    row = cursor.fetchone()

    if row is None:
        return

    remaining_stock = float(row[0] or 0)

    cursor.execute("""
        SELECT
            s.id,
            COALESCE(s.required_qty, 0)
        FROM shortages s
        JOIN orders o
          ON o.id = s.order_id
         AND o.company_id = s.company_id
        WHERE s.item_id = ?
          AND s.company_id = ?
          AND COALESCE(o.status, '') != 'Completed'
        ORDER BY s.created_at ASC, s.id ASC
    """, (item_id, company_id))
    shortages = cursor.fetchall()

    for shortage_id, required_qty in shortages:
        required_qty = float(required_qty or 0)

        coverable = min(max(remaining_stock, 0), required_qty)
        missing_qty = max(0, required_qty - coverable)
        covered = 1 if missing_qty <= 0 else 0

        cursor.execute("""
            UPDATE shortages
            SET reserved_qty = ?,
                missing_qty = ?,
                covered = ?
            WHERE id = ?
              AND company_id = ?
        """, (
            required_qty,
            missing_qty,
            covered,
            shortage_id,
            company_id
        ))

        remaining_stock -= required_qty

def recalculate_all_shortages(cursor, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT DISTINCT item_id
        FROM shortages
        WHERE company_id = ?
    """, (company_id,))
    item_rows = cursor.fetchall()

    for row in item_rows:
        current_item_id = row[0]
        if current_item_id:
            recalculate_shortages_for_item(cursor, current_item_id, company_id=company_id)


def rebuild_company_reserved_quantities(cursor, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    recalculate_all_shortages(cursor, company_id=company_id)

    cursor.execute("""
        UPDATE items
        SET reserved_quantity = 0
        WHERE company_id = ?
    """, (company_id,))

    cursor.execute("""
        SELECT
            s.item_id,
            COALESCE(SUM(s.required_qty), 0)
        FROM shortages s
        JOIN orders o
          ON o.id = s.order_id
         AND o.company_id = s.company_id
        WHERE s.company_id = ?
          AND COALESCE(o.status, '') != 'Completed'
        GROUP BY s.item_id
    """, (company_id,))
    reserved_rows = cursor.fetchall()

    for reserved_row in reserved_rows:
        current_item_id = reserved_row[0]
        reserved_total = float(reserved_row[1] or 0)

        cursor.execute("""
            UPDATE items
            SET reserved_quantity = ?
            WHERE id = ?
              AND company_id = ?
        """, (reserved_total, current_item_id, company_id))

    cursor.execute("""
        SELECT id
        FROM items
        WHERE company_id = ?
    """, (company_id,))
    all_item_rows = cursor.fetchall()

    for item_row in all_item_rows:
        current_item_id = item_row[0]
        sync_item_available(cursor, current_item_id, company_id=company_id)


def clear_order_shortages(cursor, order_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        DELETE FROM shortages
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))

def release_order_reservations(cursor, order_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT
            item_id,
            COALESCE(required_qty, 0)
        FROM shortages
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))
    rows = cursor.fetchall()

    for row in rows:
        current_item_id = row[0]
        required_qty = float(row[1] or 0)

        if not current_item_id or required_qty <= 0:
            continue

        cursor.execute("""
            UPDATE items
            SET reserved_quantity = MAX(0, COALESCE(reserved_quantity, 0) - ?)
            WHERE id = ?
              AND company_id = ?
        """, (required_qty, current_item_id, company_id))

        sync_item_available(cursor, current_item_id, company_id=company_id)


def reserve_inventory_for_order(cursor, item_id, required_qty, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    required_qty = float(required_qty or 0)

    cursor.execute("""
        SELECT
            COALESCE(stock_quantity, 0),
            COALESCE(reserved_quantity, 0)
        FROM items
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))
    item_row = cursor.fetchone()

    if item_row is None:
        raise ValueError("Item not found while reserving inventory.")

    stock_quantity = float(item_row[0] or 0)
    reserved_quantity = float(item_row[1] or 0)
    available_before = stock_quantity - reserved_quantity

    missing_qty = max(0, required_qty - max(0, available_before))

    cursor.execute("""
        UPDATE items
        SET reserved_quantity = COALESCE(reserved_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (required_qty, item_id, company_id))

    sync_item_available(cursor, item_id, company_id=company_id)

    return required_qty, missing_qty



def create_order_shortages_and_reservations(cursor, order_id, product_id, quantity, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    exploded = explode_bom_items_recursive(
        cursor,
        product_id,
        quantity,
        company_id=company_id
    )

    release_order_reservations(cursor, order_id, company_id=company_id)
    clear_order_shortages(cursor, order_id, company_id=company_id)

    has_any_demand = False

    for current_item_id, required_qty in exploded.items():
        required_qty = float(required_qty or 0)
        if required_qty <= 0:
            continue

        has_any_demand = True

        reserved_qty, missing_qty = reserve_inventory_for_order(
            cursor,
            current_item_id,
            required_qty,
            company_id=company_id
        )

        cursor.execute("""
            INSERT INTO shortages (
                company_id,
                order_id,
                item_id,
                required_qty,
                reserved_qty,
                missing_qty,
                covered
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            company_id,
            order_id,
            current_item_id,
            required_qty,
            reserved_qty,
            missing_qty,
            1 if missing_qty <= 0 else 0
        ))

    cursor.execute("""
        UPDATE orders
        SET materials_reserved = ?
        WHERE id = ?
          AND company_id = ?
    """, (
        1 if has_any_demand else 0,
        order_id,
        company_id
    ))

    rebuild_company_reserved_quantities(cursor, company_id=company_id)


def split_parent_exclusion_sql(alias="oj"):
    return f"""
        NOT EXISTS (
            SELECT 1
            FROM order_jobs child
            WHERE child.parent_job_id = {alias}.id
              AND child.is_split_child = 1
              AND child.company_id = {alias}.company_id
        )
    """





def consume_job_materials(cursor, product_id, produced_quantity, company_id=None):
    produced_quantity = float(produced_quantity or 0)

    if company_id is None:
        raise ValueError("company_id is required.")

    if not product_id or produced_quantity <= 0:
        return

    exploded_items = explode_bom_items_recursive(
        cursor,
        product_id,
        produced_quantity,
        company_id=company_id
    )

    for item_id, required_quantity in exploded_items.items():
        cursor.execute("""
            UPDATE items
            SET stock_quantity = COALESCE(stock_quantity, 0) - ?
            WHERE id = ?
              AND company_id = ?
        """, (float(required_quantity or 0), item_id, company_id))

def add_finished_product_stock(cursor, product_id, produced_quantity, company_id=None):
    produced_quantity = float(produced_quantity or 0)

    if company_id is None:
        raise ValueError("company_id is required.")

    if not product_id or produced_quantity <= 0:
        return

    cursor.execute("""
        UPDATE products
        SET stock_quantity = COALESCE(stock_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (produced_quantity, product_id, company_id))

def is_final_job(cursor, order_id, job_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT MAX(sequence)
        FROM order_jobs
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))
    max_sequence = cursor.fetchone()[0]

    cursor.execute("""
        SELECT sequence
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        return False

    job_sequence = row[0]
    return job_sequence == max_sequence


def reserve_order_materials(cursor, order_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT 1
        FROM orders
        WHERE id = ?
          AND company_id = ?
        LIMIT 1
    """, (order_id, company_id))
    exists = cursor.fetchone()

    if not exists:
        raise ValueError("Order not found.")

    # Orders currently do not reserve stock.
    # Kept for compatibility and future extension.
    return False






def get_product_material_breakdown(cursor, product_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    try:
        exploded_items = explode_bom_items_recursive(
            cursor,
            product_id,
            1,
            company_id=company_id
        )
    except ValueError:
        return []

    if not exploded_items:
        return []

    item_ids = list(exploded_items.keys())
    placeholders = ",".join(["?"] * len(item_ids))
    cursor.execute(f"""
        SELECT id, item_code, item_name, measurement_unit, COALESCE(unit_price, 0)
        FROM items
        WHERE id IN ({placeholders})
          AND company_id = ?
        ORDER BY item_name ASC
    """, item_ids + [company_id])

    materials = []
    for item_id, item_code, item_name, measurement_unit, unit_price in cursor.fetchall():
        quantity_per_unit = float(exploded_items.get(item_id, 0) or 0)
        total_cost = quantity_per_unit * float(unit_price or 0)
        materials.append({
            "item_id": item_id,
            "item_code": item_code,
            "item_name": item_name,
            "unit": measurement_unit or "pcs",
            "quantity_per_unit": quantity_per_unit,
            "unit_price": float(unit_price or 0),
            "total_cost": total_cost,
        })
    return materials


def calculate_product_material_cost(cursor, product_id, company_id=None):
    return sum(row["total_cost"] for row in get_product_material_breakdown(cursor, product_id, company_id=company_id))


def fetch_workstation_groups(cursor, company_id):
    cursor.execute("""
        SELECT
            id,
            name,
            COALESCE(description, ''),
            COALESCE(color, '#6366f1')
        FROM workstation_groups
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    return [
        {
            "id": row[0],
            "name": row[1],
            "description": row[2],
            "color": row[3],
        }
        for row in cursor.fetchall()
    ]


def get_workstation_group_current_used_load(cursor, group_key, company_id):
    if group_key.startswith("g:"):
        group_id = int(group_key.split(":", 1)[1])
        cursor.execute("""
            SELECT COALESCE(SUM(
                oj.estimated_hours *
                CASE
                    WHEN (oj.planned_quantity - oj.completed_quantity) < 0 THEN 0
                    ELSE (oj.planned_quantity - oj.completed_quantity)
                END
            ), 0)
            FROM order_jobs oj
            JOIN workstations w
              ON w.id = oj.workstation_id
             AND w.company_id = oj.company_id
            WHERE w.group_id = ?
              AND oj.company_id = ?
              AND oj.status != 'Done'
              AND (
                    oj.status = 'Waiting'
                    OR oj.status = 'Ongoing'
                    OR oj.status = 'Paused'
                    OR oj.status = 'Delayed'
              )
              AND (
                    oj.is_split_child = 1
                    OR NOT EXISTS (
                        SELECT 1
                        FROM order_jobs child
                        WHERE child.parent_job_id = oj.id
                          AND child.is_split_child = 1
                          AND child.company_id = oj.company_id
                    )
              )
        """, (group_id, company_id))
        row = cursor.fetchone()
        return float((row[0] if row else 0) or 0)

    workstation_id = int(group_key.split(":", 1)[1])
    return get_workstation_current_used_load(cursor, workstation_id, company_id)


def get_batch_planned_workstation_group_load(cursor, batch_id, company_id=None, selected_group_keys=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    selected = {str(value).strip() for value in (selected_group_keys or []) if str(value).strip()}

    cursor.execute("""
        SELECT
            CASE
                WHEN wg.id IS NOT NULL THEN 'g:' || wg.id
                ELSE 'w:' || w.id
            END AS load_key,
            CASE
                WHEN wg.id IS NOT NULL THEN wg.name
                ELSE 'Ungrouped — ' || w.name
            END AS group_name,
            CASE
                WHEN wg.id IS NOT NULL THEN COALESCE(wg.color, w.color, '#6366f1')
                ELSE COALESCE(w.color, '#3b82f6')
            END AS group_color,
            CASE
                WHEN wg.id IS NOT NULL THEN wg.id
                ELSE NULL
            END AS workstation_group_id,
            COALESCE(SUM(COALESCE(pjt.estimated_hours, 0) * COALESCE(bo.quantity, 0)), 0) AS planned_batch_hours,
            COUNT(DISTINCT bo.id) AS batch_order_count,
            COUNT(DISTINCT pjt.id) AS job_template_count
        FROM batch_orders bo
        JOIN product_job_templates pjt
          ON pjt.product_id = bo.product_id
         AND pjt.company_id = bo.company_id
        JOIN workstations w
          ON w.id = pjt.workstation_id
         AND w.company_id = pjt.company_id
        LEFT JOIN workstation_groups wg
          ON wg.id = w.group_id
         AND wg.company_id = w.company_id
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        GROUP BY
            CASE
                WHEN wg.id IS NOT NULL THEN 'g:' || wg.id
                ELSE 'w:' || w.id
            END,
            CASE
                WHEN wg.id IS NOT NULL THEN wg.name
                ELSE 'Ungrouped — ' || w.name
            END,
            CASE
                WHEN wg.id IS NOT NULL THEN COALESCE(wg.color, w.color, '#6366f1')
                ELSE COALESCE(w.color, '#3b82f6')
            END,
            CASE
                WHEN wg.id IS NOT NULL THEN wg.id
                ELSE NULL
            END
        ORDER BY group_name ASC
    """, (batch_id, company_id))

    rows = []
    all_rows = cursor.fetchall()

    for row in all_rows:
        group_key = row[0]
        if selected and group_key not in selected:
            continue

        if group_key.startswith("g:"):
            group_id = int(group_key.split(":", 1)[1])

            cursor.execute("""
                SELECT
                    COALESCE(SUM(
                        COALESCE(w.hours_per_shift, 0) *
                        COALESCE(w.shifts_per_day, 0) *
                        COALESCE(w.working_days_per_month, 0)
                    ), 0),
                    COUNT(*)
                FROM workstations w
                WHERE w.group_id = ?
                  AND w.company_id = ?
            """, (group_id, company_id))
            capacity_row = cursor.fetchone()
            monthly_capacity = float((capacity_row[0] if capacity_row else 0) or 0)
            workstation_count = int((capacity_row[1] if capacity_row else 0) or 0)
        else:
            workstation_id = int(group_key.split(":", 1)[1])
            cursor.execute("""
                SELECT
                    (
                        COALESCE(hours_per_shift, 0) *
                        COALESCE(shifts_per_day, 0) *
                        COALESCE(working_days_per_month, 0)
                    )
                FROM workstations
                WHERE id = ?
                  AND company_id = ?
            """, (workstation_id, company_id))
            capacity_row = cursor.fetchone()
            monthly_capacity = float((capacity_row[0] if capacity_row else 0) or 0)
            workstation_count = 1

        current_used = get_workstation_group_current_used_load(cursor, group_key, company_id)
        planned_batch_hours = float(row[4] or 0)
        preview_total = current_used + planned_batch_hours

        rows.append({
            "group_key": group_key,
            "workstation_group_id": row[3],
            "name": row[1],
            "color": row[2] or "#6366f1",
            "monthly_capacity": monthly_capacity,
            "current_used": current_used,
            "planned_batch_hours": planned_batch_hours,
            "preview_total": preview_total,
            "preview_percent": (preview_total / monthly_capacity * 100) if monthly_capacity > 0 else 0,
            "batch_order_count": int(row[5] or 0),
            "job_template_count": int(row[6] or 0),
            "workstation_count": workstation_count,
        })

    return rows


def get_product_job_cost_breakdown(cursor, product_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT
            pjt.id,
            pjt.job_name,
            pjt.sequence,
            COALESCE(pjt.estimated_hours, 0),
            pjt.workstation_id,
            COALESCE(w.name, '-'),
            COALESCE(w.cost_per_hour, 0)
        FROM product_job_templates pjt
        LEFT JOIN workstations w
          ON w.id = pjt.workstation_id
         AND w.company_id = pjt.company_id
        WHERE pjt.product_id = ?
          AND pjt.company_id = ?
        ORDER BY pjt.sequence ASC, pjt.id ASC
    """, (product_id, company_id))

    jobs = []
    for row in cursor.fetchall():
        estimated_hours = float(row[3] or 0)
        workstation_cost_per_hour = float(row[6] or 0)
        jobs.append({
            "job_id": row[0],
            "job_name": row[1],
            "sequence": row[2],
            "estimated_hours": estimated_hours,
            "workstation_id": row[4],
            "workstation_name": row[5],
            "workstation_cost_per_hour": workstation_cost_per_hour,
            "job_cost": estimated_hours * workstation_cost_per_hour,
        })
    return jobs


def calculate_product_jobs_cost(cursor, product_id, company_id=None):
    return sum(job["job_cost"] for job in get_product_job_cost_breakdown(cursor, product_id, company_id=company_id))


def get_product_cost_snapshot(cursor, product_id, company_id=None):
    materials = get_product_material_breakdown(cursor, product_id, company_id=company_id)
    jobs = get_product_job_cost_breakdown(cursor, product_id, company_id=company_id)
    materials_cost = sum(row["total_cost"] for row in materials)
    jobs_cost = sum(row["job_cost"] for row in jobs)
    return {
        "materials": materials,
        "jobs": jobs,
        "materials_cost": materials_cost,
        "jobs_cost": jobs_cost,
        "total_cost": materials_cost + jobs_cost,
    }


def get_order_cost_snapshot(cursor, product_id, quantity, company_id=None):
    product_cost = get_product_cost_snapshot(cursor, product_id, company_id=company_id)
    quantity = float(quantity or 0)
    return {
        "materials_cost": product_cost["materials_cost"] * quantity,
        "jobs_cost": product_cost["jobs_cost"] * quantity,
        "total_cost": product_cost["total_cost"] * quantity,
    }


def get_batch_cost_snapshot(cursor, batch_id, company_id=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    cursor.execute("""
        SELECT product_id, quantity
        FROM batch_orders
        WHERE batch_id = ?
          AND company_id = ?
    """, (batch_id, company_id))

    totals = {
        "materials_cost": 0.0,
        "jobs_cost": 0.0,
        "total_cost": 0.0,
    }
    for product_id, quantity in cursor.fetchall():
        snapshot = get_order_cost_snapshot(cursor, product_id, quantity, company_id=company_id)
        totals["materials_cost"] += snapshot["materials_cost"]
        totals["jobs_cost"] += snapshot["jobs_cost"]
        totals["total_cost"] += snapshot["total_cost"]
    return totals


def get_workstation_current_used_load(cursor, workstation_id, company_id):
    cursor.execute("""
        SELECT COALESCE(SUM(
            oj.estimated_hours *
            CASE
                WHEN (oj.planned_quantity - oj.completed_quantity) < 0 THEN 0
                ELSE (oj.planned_quantity - oj.completed_quantity)
            END
        ), 0)
        FROM order_jobs oj
        WHERE oj.workstation_id = ?
          AND oj.company_id = ?
          AND oj.status != 'Done'
          AND (
                oj.status = 'Waiting'
                OR oj.status = 'Ongoing'
                OR oj.status = 'Paused'
                OR oj.status = 'Delayed'
          )
          AND (
                oj.is_split_child = 1
                OR NOT EXISTS (
                    SELECT 1
                    FROM order_jobs child
                    WHERE child.parent_job_id = oj.id
                      AND child.is_split_child = 1
                      AND child.company_id = oj.company_id
                )
          )
    """, (workstation_id, company_id))
    row = cursor.fetchone()
    return float((row[0] if row else 0) or 0)


def get_batch_planned_workstation_load(cursor, batch_id, company_id=None, selected_workstation_ids=None):
    if company_id is None:
        raise ValueError("company_id is required.")

    selected = {int(value) for value in (selected_workstation_ids or []) if str(value).strip()}

    cursor.execute("""
        SELECT
            w.id,
            w.name,
            w.color,
            COALESCE(w.cost_per_hour, 0),
            COALESCE(w.hours_per_shift, 0),
            COALESCE(w.shifts_per_day, 0),
            COALESCE(w.working_days_per_month, 0),
            COALESCE(SUM(COALESCE(pjt.estimated_hours, 0) * COALESCE(bo.quantity, 0)), 0) AS planned_batch_hours,
            COUNT(DISTINCT bo.id) AS batch_order_count,
            COUNT(DISTINCT pjt.id) AS job_template_count
        FROM batch_orders bo
        JOIN product_job_templates pjt
          ON pjt.product_id = bo.product_id
         AND pjt.company_id = bo.company_id
        JOIN workstations w
          ON w.id = pjt.workstation_id
         AND w.company_id = pjt.company_id
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        GROUP BY w.id, w.name, w.color, w.cost_per_hour, w.hours_per_shift, w.shifts_per_day, w.working_days_per_month
        ORDER BY w.name ASC
    """, (batch_id, company_id))

    rows = []
    for row in cursor.fetchall():
        workstation_id = int(row[0])
        if selected and workstation_id not in selected:
            continue

        monthly_capacity = float(row[4] or 0) * float(row[5] or 0) * float(row[6] or 0)
        current_used = get_workstation_current_used_load(cursor, workstation_id, company_id)
        planned_batch_hours = float(row[7] or 0)
        preview_total = current_used + planned_batch_hours
        rows.append({
            "workstation_id": workstation_id,
            "name": row[1],
            "color": row[2] or "#3b82f6",
            "cost_per_hour": float(row[3] or 0),
            "monthly_capacity": monthly_capacity,
            "current_used": current_used,
            "planned_batch_hours": planned_batch_hours,
            "preview_total": preview_total,
            "preview_percent": (preview_total / monthly_capacity * 100) if monthly_capacity > 0 else 0,
            "batch_order_count": int(row[8] or 0),
            "job_template_count": int(row[9] or 0),
        })
    return rows


def fetch_product_groups(cursor, company_id):
    cursor.execute("""
        SELECT id, name, COALESCE(description, '')
        FROM product_groups
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    return [
        {"id": row[0], "name": row[1], "description": row[2]}
        for row in cursor.fetchall()
    ]


from functools import wraps

ROLE_DEFAULT_PERMISSIONS = {
    "admin": {
        "view_dashboard",
        "view_orders", "manage_orders",
        "view_jobs", "update_job_progress", "manage_jobs",
        "view_inventory", "manage_inventory",
        "view_products", "manage_products",
        "view_items", "manage_items",
        "view_workstations", "manage_workstations",
        "view_reports", "export_data",
        "manage_users",
        "manage_procurement",
        "view_suppliers",
        "manage_suppliers",
        "view_procurement",
        "manage_procurement",
    },
    "manager": {
        "view_dashboard",
        "view_orders", "manage_orders",
        "view_jobs", "update_job_progress", "manage_jobs",
        "view_inventory", "manage_inventory",
        "view_products", "manage_products",
        "view_items", "manage_items",
        "view_workstations", "manage_workstations",
        "view_reports", "export_data",
        "manage_procurement",
        "view_suppliers",
        "manage_suppliers",
        "view_procurement",
    },
    "worker": {
        "view_dashboard",
        "view_jobs",
        "update_job_progress",
        "view_procurement",
    },
}

ALL_PERMISSION_KEYS = [
    "view_dashboard",
    "view_orders",
    "manage_orders",
    "view_jobs",
    "update_job_progress",
    "manage_jobs",
    "view_inventory",
    "manage_inventory",
    "view_products",
    "manage_products",
    "view_items",
    "manage_items",
    "view_workstations",
    "manage_workstations",
    "view_reports",
    "export_data",
    "manage_users",
    "manage_procurement",
    "view_suppliers",
    "manage_suppliers",
    "view_procurement",
    "manage_procurement",
]


def get_current_user_role():
    return (session.get("user_role") or "worker").strip().lower()


def get_role_default_permissions(role):
    return set(ROLE_DEFAULT_PERMISSIONS.get((role or "worker").lower(), set()))


def get_user_permission_overrides(user_id):
    if not user_id:
        return {}

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT permission_key, allowed
            FROM user_permissions
            WHERE user_id = ?
              AND (company_id = ? OR company_id IS NULL)
        """, (user_id, company_id))

        rows = cursor.fetchall()
        overrides = {}

        for row in rows:
            permission_key = row[0]
            allowed = row[1]
            overrides[permission_key] = bool(allowed)

        return overrides

    except Exception:
        # jei lentelės dar nėra arba query nulūžta, negriaunam viso home puslapio
        return {}
    finally:
        conn.close()


def get_effective_permissions(user_id=None, role=None):
    if user_id is None:
        user_id = session.get("user_id")
    if role is None:
        role = get_current_user_role()

    permissions = set(get_role_default_permissions(role))
    overrides = get_user_permission_overrides(user_id)

    for permission_key, allowed in overrides.items():
        if allowed:
            permissions.add(permission_key)
        else:
            permissions.discard(permission_key)

    return permissions


def has_permission(permission_key, user_id=None, role=None):
    return permission_key in get_effective_permissions(user_id=user_id, role=role)


def get_dashboard_layout(cursor, user_id, company_id, page_key="dashboard"):
    cursor.execute("""
        SELECT layout_json
        FROM dashboard_layouts
        WHERE user_id = ? AND company_id = ? AND page_key = ?
        LIMIT 1
    """, (user_id, company_id, page_key))

    row = cursor.fetchone()

    if not row:
        return []

    try:
        return json.loads(row[0])
    except (TypeError, ValueError, json.JSONDecodeError):
        return []


def save_dashboard_layout_record(cursor, user_id, company_id, layout_state, page_key="dashboard"):
    cursor.execute("""
        INSERT INTO dashboard_layouts (user_id, company_id, page_key, layout_json, updated_at)
        VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(user_id, company_id, page_key)
        DO UPDATE SET
            layout_json = excluded.layout_json,
            updated_at = CURRENT_TIMESTAMP
    """, (user_id, company_id, page_key, json.dumps(layout_state)))

def permission_required(permission_key):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(*args, **kwargs):
            if not is_logged_in():
                return redirect(url_for("login"))

            if not has_permission(permission_key):
                flash("You do not have permission to access this page.", "error")
                return redirect(url_for("dashboard"))

            return view_func(*args, **kwargs)
        return wrapped_view
    return decorator





@app.context_processor
def inject_permissions():
    if not session.get("user_id"):
        return {
            "current_user_role": None,
            "effective_permissions": set(),
            "has_permission_ui": lambda permission_key: False,
            "all_permission_keys": ALL_PERMISSION_KEYS,
        }

    effective_permissions = get_effective_permissions()

    return {
        "current_user_role": get_current_user_role(),
        "effective_permissions": effective_permissions,
        "has_permission_ui": lambda permission_key: permission_key in effective_permissions,
        "all_permission_keys": ALL_PERMISSION_KEYS,
    }


@app.route("/health", methods=["GET", "HEAD"])
def health():
    return "ok", 200

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/landing")
def landing():
    return render_template("index.html")



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if not email or not password:
            flash("Email and password are required.", "error")
            return redirect(url_for("login"))

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT id, full_name, email, password, company_id, role
            FROM users
            WHERE email = %s
            """,
            (email,),
        )
        user = cursor.fetchone()
        conn.close()

        if not user:
            flash("Invalid email or password.", "error")
            return redirect(url_for("login"))

        user_id, full_name, user_email, hashed_password, company_id, role = user

        if not check_password_hash(hashed_password, password):
            flash("Invalid email or password.", "error")
            return redirect(url_for("login"))

        # session
        session["user_id"] = user_id
        session["user_name"] = full_name
        session["user_email"] = user_email
        session["company_id"] = company_id
        session["role"] = role

        return redirect(url_for("dashboard"))

    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        company_name = request.form.get("company_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not full_name or not company_name or not email or not password:
            flash("All fields are required.", "error")
            return redirect(url_for("register"))

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return redirect(url_for("register"))

        hashed_password = generate_password_hash(password)

        conn = get_connection()
        cursor = conn.cursor()

        try:
            # 1. sukuriam company
            cursor.execute(
                """
                INSERT INTO companies (company_name)
                VALUES (%s)
                RETURNING id
                """,
                (company_name,),
            )
            company_id = cursor.fetchone()[0]

            # 2. sukuriam user (BE company_name čia!)
            cursor.execute(
                """
                INSERT INTO users (full_name, email, password, company_id, role)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (full_name, email, hashed_password, company_id, "admin"),
            )

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Error: {str(e)}", "error")
            return redirect(url_for("register"))

        finally:
            conn.close()

        flash("Account created successfully. Please login.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))

@app.route("/dashboard")
def dashboard():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE company_id = ?
    """, (company_id,))
    total_orders = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status = 'Waiting' AND company_id = ?
    """, (company_id,))
    waiting_count = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status = 'Completed' AND company_id = ?
    """, (company_id,))
    completed_count = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM orders
        WHERE status = 'In Progress' AND company_id = ?
    """, (company_id,))
    in_progress_count = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM order_jobs
        WHERE status = 'Delayed' AND company_id = ?
    """, (company_id,))
    delayed_count = cursor.fetchone()[0]

    cursor.execute("""
        SELECT
            o.id,
            o.order_number,
            COALESCE(p.product_name, '-') AS product_name,
            o.quantity,
            o.status,
            o.due_date
        FROM orders o
        LEFT JOIN products p
          ON o.product_id = p.id
         AND p.company_id = o.company_id
        WHERE o.company_id = ?
        ORDER BY o.id DESC
        LIMIT 10
    """, (company_id,))
    order_rows = cursor.fetchall()

    recent_orders = []
    for row in order_rows:
        recent_orders.append({
            "id": row[0],
            "order_number": row[1],
            "product_name": row[2],
            "quantity": row[3],
            "status": row[4],
            "due_date": row[5]
        })

    cursor.execute("""
        SELECT
            w.id,
            w.name,
            (w.hours_per_shift * w.shifts_per_day * w.working_days_per_month) AS monthly_capacity,
            COALESCE((
                SELECT SUM(
                    oj.estimated_hours *
                    CASE
                        WHEN (oj.planned_quantity - oj.completed_quantity) < 0 THEN 0
                        ELSE (oj.planned_quantity - oj.completed_quantity)
                    END
                )
                FROM order_jobs oj
                WHERE oj.workstation_id = w.id
                AND oj.company_id = w.company_id
                AND oj.status != 'Done'
                AND (
                        oj.status = 'Waiting'
                        OR oj.status = 'Ongoing'
                        OR oj.status = 'Paused'
                        OR oj.status = 'Delayed'
                )
                AND (
                        oj.is_split_child = 1
                        OR NOT EXISTS (
                            SELECT 1
                            FROM order_jobs child
                            WHERE child.parent_job_id = oj.id
                            AND child.is_split_child = 1
                            AND child.company_id = oj.company_id
                        )
                )
            ), 0) AS used_load
        FROM workstations w
        WHERE w.company_id = ?
        ORDER BY w.name ASC
    """, (company_id,))
    workstation_rows = cursor.fetchall()

    workstation_load = []
    for row in workstation_rows:
        monthly_capacity = float(row[2] or 0)
        used_load = float(row[3] or 0)
        load_percent = round((used_load / monthly_capacity) * 100) if monthly_capacity > 0 else 0

        workstation_load.append({
            "id": row[0],
            "name": row[1],
            "monthly_capacity": round(monthly_capacity, 2),
            "used_load": round(used_load, 2),
            "load_percent": load_percent
        })

    saved_dashboard_layout = get_dashboard_layout(cursor, user_id, company_id, "dashboard")

    conn.close()

    return render_template(
        "dashboard.html",
        active_page="dashboard",
        total_orders=total_orders,
        waiting_count=waiting_count,
        completed_count=completed_count,
        in_progress_count=in_progress_count,
        delayed_count=delayed_count,
        recent_orders=recent_orders,
        workstation_load=workstation_load,
        saved_dashboard_layout=saved_dashboard_layout
    )


@app.route("/orders")
@permission_required("view_orders")
def orders():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    order_number = request.args.get("order_number", "").strip()
    product_name = request.args.get("product_name", "").strip()
    statuses = request.args.getlist("status")
    priority = request.args.get("priority", "").strip()
    due_date_from = request.args.get("due_date_from", "").strip()
    due_date_to = request.args.get("due_date_to", "").strip()
    group_id = request.args.get("group_id", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            o.id,
            o.order_number,
            p.product_name,
            o.quantity,
            o.status,
            o.due_date,
            o.priority,
            o.product_id,
            COALESCE(pg.name, '')
        FROM orders o
        LEFT JOIN products p
          ON o.product_id = p.id
         AND p.company_id = o.company_id
        LEFT JOIN product_groups pg
          ON pg.id = p.group_id
         AND pg.company_id = p.company_id
        WHERE o.company_id = ?
    """

    params = [company_id]

    if order_number:
        query += " AND o.order_number LIKE ?"
        params.append(f"%{order_number}%")

    if product_name:
        query += " AND p.product_name LIKE ?"
        params.append(f"%{product_name}%")

    if statuses and "All" not in statuses:
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND o.status IN ({placeholders})"
        params.extend(statuses)

    if priority:
        query += " AND o.priority = ?"
        params.append(priority)

    if due_date_from:
        query += " AND o.due_date >= ?"
        params.append(due_date_from)

    if due_date_to:
        query += " AND o.due_date <= ?"
        params.append(due_date_to)

    if group_id:
        query += " AND p.group_id = ?"
        params.append(group_id)

    query += " ORDER BY o.id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    product_groups = fetch_product_groups(cursor, company_id)

    orders = []
    for row in rows:
        cost_snapshot = get_order_cost_snapshot(cursor, row[7], row[3], company_id=company_id)
        orders.append({
            "id": row[0],
            "order_number": row[1],
            "product_name": row[2] if row[2] else "-",
            "quantity": float(row[3] or 0),
            "status": row[4],
            "due_date": row[5],
            "priority": row[6],
            "group_name": row[8] or "-",
            "materials_cost": cost_snapshot["materials_cost"],
            "jobs_cost": cost_snapshot["jobs_cost"],
            "total_cost": cost_snapshot["total_cost"],
        })

    conn.close()

    return render_template(
        "orders.html",
        orders=orders,
        active_page="orders",
        product_groups=product_groups,
        filters={
            "order_number": order_number,
            "product_name": product_name,
            "status": statuses,
            "priority": priority,
            "due_date_from": due_date_from,
            "due_date_to": due_date_to,
            "group_id": group_id,
        }
    )


@app.route("/orders/new", methods=["GET", "POST"])
@permission_required("manage_orders")
def new_order():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    selected_group_id = request.args.get("group_id", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        order_number = request.form["order_number"].strip()
        product_id_raw = request.form.get("product_id", "").strip()
        quantity_raw = request.form.get("quantity", "").strip()
        status = request.form["status"].strip()
        due_date = request.form["due_date"].strip()
        priority = request.form["priority"].strip()

        if not product_id_raw:
            conn.close()
            flash("No product selected. Create a product first or choose an existing one.", "error")
            return redirect(url_for("new_order"))

        try:
            product_id = int(product_id_raw)
            quantity = float(quantity_raw)
        except ValueError:
            conn.close()
            flash("Invalid product or quantity.", "error")
            return redirect(url_for("new_order"))

        if quantity <= 0:
            conn.close()
            flash("Quantity must be greater than 0.", "error")
            return redirect(url_for("new_order"))

        cursor.execute("""
            SELECT id
            FROM products
            WHERE id = ? AND company_id = ?
        """, (product_id, company_id))
        product_row = cursor.fetchone()

        if product_row is None:
            conn.close()
            flash("Selected product was not found in your current company.", "error")
            return redirect(url_for("new_order"))

        cursor.execute("""
            INSERT INTO orders (
                order_number,
                product_id,
                quantity,
                status,
                due_date,
                priority,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            order_number,
            product_id,
            quantity,
            status,
            due_date,
            priority,
            company_id
        ))

        order_id = cursor.lastrowid

        create_order_shortages_and_reservations(
            cursor,
            order_id,
            product_id,
            quantity,
            company_id=company_id
        )

        generate_order_jobs_recursive(
            cursor,
            order_id,
            product_id,
            quantity,
            planned_date=None,
            company_id=company_id
        )

        conn.commit()
        conn.close()

        flash("Order created successfully.", "success")
        return redirect(url_for("orders"))

    cursor.execute("""
        SELECT id, product_code, product_name, COALESCE(group_id, 0)
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC, product_code ASC
    """, (company_id,))
    products = [
        {
            "id": row[0],
            "product_code": row[1],
            "product_name": row[2],
            "group_id": row[3] if row[3] != 0 else None,
        }
        for row in cursor.fetchall()
    ]
    product_groups = fetch_product_groups(cursor, company_id)

    conn.close()

    return render_template(
        "new_order.html",
        products=products,
        product_groups=product_groups,
        selected_group_id=selected_group_id,
        active_page="orders"
    )


@app.route("/order-batches")
@permission_required("view_orders")
def order_batches():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            b.id,
            b.batch_number,
            b.name,
            b.status,
            b.notes,
            b.created_at,
            b.launched_at,
            COALESCE(COUNT(bo.id), 0) AS order_count,
            COALESCE(SUM(bo.quantity), 0) AS total_quantity
        FROM order_batches b
        LEFT JOIN batch_orders bo
          ON bo.batch_id = b.id
         AND bo.company_id = b.company_id
        WHERE b.company_id = ?
          AND COALESCE(b.is_deleted, 0) = 0
        GROUP BY b.id, b.batch_number, b.name, b.status, b.notes, b.created_at, b.launched_at
        ORDER BY b.id DESC
    """, (company_id,))
    rows = cursor.fetchall()

    batches = []
    for row in rows:
        cost_snapshot = get_batch_cost_snapshot(cursor, row[0], company_id=company_id)
        load_preview = get_batch_planned_workstation_group_load(cursor, row[0], company_id=company_id)
        batches.append({
            "id": row[0],
            "batch_number": row[1] or f"BAT-{row[0]:05d}",
            "name": row[2],
            "status": row[3],
            "notes": row[4] or "",
            "created_at": row[5],
            "launched_at": row[6],
            "order_count": int(row[7] or 0),
            "total_quantity": float(row[8] or 0),
            "materials_cost": cost_snapshot["materials_cost"],
            "jobs_cost": cost_snapshot["jobs_cost"],
            "total_cost": cost_snapshot["total_cost"],
            "planned_groups": len(load_preview),
            "planned_hours": sum(item["planned_batch_hours"] for item in load_preview),
        })

    conn.close()

    return render_template(
        "order_batches.html",
        batches=batches,
        active_page="orders"
    )


@app.route("/order-batches/new", methods=["GET", "POST"])
@permission_required("manage_orders")
def new_order_batch():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        notes = request.form.get("notes", "").strip()

        if not name:
            conn.close()
            flash("Batch name is required.", "error")
            return redirect(url_for("new_order_batch"))

        batch_number = generate_batch_number(cursor, company_id)

        cursor.execute("""
            INSERT INTO order_batches (
                company_id,
                batch_number,
                name,
                status,
                notes,
                created_by
            )
            VALUES (?, ?, ?, 'draft', ?, ?)
        """, (
            company_id,
            batch_number,
            name,
            notes or None,
            user_id
        ))

        batch_id = cursor.lastrowid
        conn.commit()
        conn.close()

        flash("Order batch created successfully.", "success")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    conn.close()
    return render_template(
        "new_order_batch.html",
        active_page="orders"
    )


@app.route("/order-batches/<int:batch_id>/delete", methods=["POST"])
@permission_required("manage_orders")
def delete_order_batch(batch_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        conn.close()
        flash("Batch not found.", "error")
        return redirect(url_for("order_batches"))

    if batch["status"] == "draft":
        cursor.execute("DELETE FROM batch_orders WHERE batch_id = ? AND company_id = ?", (batch_id, company_id))
        cursor.execute("DELETE FROM order_batches WHERE id = ? AND company_id = ?", (batch_id, company_id))
        conn.commit()
        conn.close()

        flash("Draft batch deleted.", "success")
        return redirect(url_for("order_batches"))

    cursor.execute("""
        UPDATE order_batches
        SET is_deleted = 1
        WHERE id = ?
          AND company_id = ?
    """, (batch_id, company_id))

    conn.commit()
    conn.close()

    flash("Launched batch removed from active planning list. Real launched orders/jobs were kept.", "success")
    return redirect(url_for("order_batches"))


@app.route("/order-batches/<int:batch_id>")
@permission_required("view_orders")
def view_order_batch(batch_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    selected_group_keys = request.args.getlist("preview_group")

    conn = get_connection()
    cursor = conn.cursor()

    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        conn.close()
        flash("Batch not found.", "error")
        return redirect(url_for("order_batches"))

    cursor.execute("""
        SELECT
            bo.id,
            bo.order_number,
            bo.product_id,
            COALESCE(p.product_name, '-'),
            COALESCE(p.product_code, '-'),
            bo.quantity,
            bo.due_date,
            bo.priority,
            bo.notes,
            bo.launched_order_id,
            COALESCE(pg.name, '')
        FROM batch_orders bo
        LEFT JOIN products p
          ON p.id = bo.product_id
         AND p.company_id = bo.company_id
        LEFT JOIN product_groups pg
          ON pg.id = p.group_id
         AND pg.company_id = p.company_id
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        ORDER BY bo.id DESC
    """, (batch_id, company_id))
    batch_order_rows = cursor.fetchall()

    batch_orders_list = []
    for row in batch_order_rows:
        cost_snapshot = get_order_cost_snapshot(cursor, row[2], row[5], company_id=company_id)
        batch_orders_list.append({
            "id": row[0],
            "order_number": row[1],
            "product_id": row[2],
            "product_name": row[3],
            "product_code": row[4],
            "quantity": float(row[5] or 0),
            "due_date": row[6],
            "priority": row[7] or "Medium",
            "notes": row[8] or "",
            "launched_order_id": row[9],
            "group_name": row[10] or "-",
            "materials_cost": cost_snapshot["materials_cost"],
            "jobs_cost": cost_snapshot["jobs_cost"],
            "total_cost": cost_snapshot["total_cost"],
        })

    materials = get_batch_material_requirements(cursor, batch_id, company_id)
    batch_cost = get_batch_cost_snapshot(cursor, batch_id, company_id=company_id)
    planned_groups = get_batch_planned_workstation_group_load(
        cursor,
        batch_id,
        company_id=company_id,
        selected_group_keys=selected_group_keys,
    )

    cursor.execute("""
        SELECT id, product_code, product_name, COALESCE(group_id, 0)
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC, product_code ASC
    """, (company_id,))
    product_rows = cursor.fetchall()
    products = [
        {
            "id": row[0],
            "product_code": row[1],
            "product_name": row[2],
            "group_id": row[3] if row[3] != 0 else None,
        }
        for row in product_rows
    ]
    product_groups = fetch_product_groups(cursor, company_id)

    cursor.execute("""
        SELECT DISTINCT
            CASE
                WHEN wg.id IS NOT NULL THEN 'g:' || wg.id
                ELSE 'w:' || w.id
            END AS group_key,
            CASE
                WHEN wg.id IS NOT NULL THEN wg.name
                ELSE 'Ungrouped — ' || w.name
            END AS group_name
        FROM workstations w
        JOIN product_job_templates pjt
          ON pjt.workstation_id = w.id
         AND pjt.company_id = w.company_id
        JOIN batch_orders bo
          ON bo.product_id = pjt.product_id
         AND bo.company_id = pjt.company_id
        LEFT JOIN workstation_groups wg
          ON wg.id = w.group_id
         AND wg.company_id = w.company_id
        WHERE bo.batch_id = ?
          AND bo.company_id = ?
        ORDER BY group_name ASC
    """, (batch_id, company_id))
    involved_groups = [
        {"group_key": row[0], "name": row[1]}
        for row in cursor.fetchall()
    ]

    conn.close()

    return render_template(
        "view_order_batch.html",
        batch=batch,
        batch_orders=batch_orders_list,
        materials=materials,
        products=products,
        product_groups=product_groups,
        selected_group_id=request.args.get("group_id", "").strip(),
        batch_cost=batch_cost,
        planned_groups=planned_groups,
        involved_groups=involved_groups,
        selected_group_keys=[str(value) for value in selected_group_keys],
        active_page="orders"
    )


@app.route("/order-batches/<int:batch_id>/add-order", methods=["POST"])
@permission_required("manage_orders")
def add_order_to_batch(batch_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        conn.close()
        flash("Batch not found.", "error")
        return redirect(url_for("order_batches"))

    if batch["status"] != "draft":
        conn.close()
        flash("Only draft batches can be changed.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    order_number = request.form.get("order_number", "").strip()
    product_id_raw = request.form.get("product_id", "").strip()
    quantity_raw = request.form.get("quantity", "").strip()
    due_date = request.form.get("due_date", "").strip()
    priority = request.form.get("priority", "Medium").strip()
    notes = request.form.get("notes", "").strip()

    if not order_number:
        conn.close()
        flash("Order number is required.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    try:
        product_id = int(product_id_raw)
        quantity = float(quantity_raw)
    except ValueError:
        conn.close()
        flash("Invalid product or quantity.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    if quantity <= 0:
        conn.close()
        flash("Quantity must be greater than 0.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    cursor.execute("""
        SELECT id
        FROM products
        WHERE id = ?
          AND company_id = ?
    """, (product_id, company_id))
    product_exists = cursor.fetchone()

    if product_exists is None:
        conn.close()
        flash("Selected product was not found.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    cursor.execute("""
        INSERT INTO batch_orders (
            batch_id,
            company_id,
            order_number,
            product_id,
            quantity,
            due_date,
            priority,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        batch_id,
        company_id,
        order_number,
        product_id,
        quantity,
        due_date or None,
        priority or "Medium",
        notes or None
    ))

    conn.commit()
    conn.close()

    flash("Order added to batch.", "success")
    return redirect(url_for("view_order_batch", batch_id=batch_id))


@app.route("/order-batches/<int:batch_id>/edit-order/<int:batch_order_id>", methods=["GET", "POST"])
@permission_required("manage_orders")
def edit_batch_order(batch_id, batch_order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        conn.close()
        flash("Batch not found.", "error")
        return redirect(url_for("order_batches"))

    if batch["status"] != "draft":
        conn.close()
        flash("Only draft batches can be changed.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    cursor.execute("""
        SELECT
            bo.id,
            bo.order_number,
            bo.product_id,
            bo.quantity,
            bo.due_date,
            bo.priority,
            bo.notes
        FROM batch_orders bo
        WHERE bo.id = ?
          AND bo.batch_id = ?
          AND bo.company_id = ?
    """, (batch_order_id, batch_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Batch order not found.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    if request.method == "POST":
        order_number = request.form.get("order_number", "").strip()
        product_id_raw = request.form.get("product_id", "").strip()
        quantity_raw = request.form.get("quantity", "").strip()
        due_date = request.form.get("due_date", "").strip()
        priority = request.form.get("priority", "Medium").strip()
        notes = request.form.get("notes", "").strip()

        if not order_number:
            conn.close()
            flash("Order number is required.", "error")
            return redirect(url_for("edit_batch_order", batch_id=batch_id, batch_order_id=batch_order_id))

        try:
            product_id = int(product_id_raw)
            quantity = float(quantity_raw)
        except ValueError:
            conn.close()
            flash("Invalid product or quantity.", "error")
            return redirect(url_for("edit_batch_order", batch_id=batch_id, batch_order_id=batch_order_id))

        if quantity <= 0:
            conn.close()
            flash("Quantity must be greater than 0.", "error")
            return redirect(url_for("edit_batch_order", batch_id=batch_id, batch_order_id=batch_order_id))

        cursor.execute("""
            SELECT id
            FROM products
            WHERE id = ?
              AND company_id = ?
        """, (product_id, company_id))
        product_exists = cursor.fetchone()

        if product_exists is None:
            conn.close()
            flash("Selected product was not found.", "error")
            return redirect(url_for("edit_batch_order", batch_id=batch_id, batch_order_id=batch_order_id))

        cursor.execute("""
            UPDATE batch_orders
            SET order_number = ?,
                product_id = ?,
                quantity = ?,
                due_date = ?,
                priority = ?,
                notes = ?
            WHERE id = ?
              AND batch_id = ?
              AND company_id = ?
        """, (
            order_number,
            product_id,
            quantity,
            due_date or None,
            priority or "Medium",
            notes or None,
            batch_order_id,
            batch_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Batch order updated.", "success")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    batch_order = {
        "id": row[0],
        "order_number": row[1],
        "product_id": row[2],
        "quantity": float(row[3] or 0),
        "due_date": row[4] or "",
        "priority": row[5] or "Medium",
        "notes": row[6] or ""
    }

    cursor.execute("""
        SELECT id, product_code, product_name, COALESCE(group_id, 0)
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC, product_code ASC
    """, (company_id,))
    product_rows = cursor.fetchall()
    products = [
        {
            "id": product_row[0],
            "product_code": product_row[1],
            "product_name": product_row[2],
            "group_id": product_row[3] if product_row[3] != 0 else None,
        }
        for product_row in product_rows
    ]
    product_groups = fetch_product_groups(cursor, company_id)

    conn.close()

    return render_template(
        "edit_batch_order.html",
        batch=batch,
        batch_order=batch_order,
        products=products,
        product_groups=product_groups,
        active_page="orders"
    )


@app.route("/order-batches/<int:batch_id>/delete-order/<int:batch_order_id>", methods=["POST"])
@permission_required("manage_orders")
def delete_batch_order(batch_id, batch_order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    batch = fetch_batch_with_stats(cursor, batch_id, company_id)
    if batch is None:
        conn.close()
        flash("Batch not found.", "error")
        return redirect(url_for("order_batches"))

    if batch["status"] != "draft":
        conn.close()
        flash("Only draft batches can be changed.", "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))

    cursor.execute("""
        DELETE FROM batch_orders
        WHERE id = ?
          AND batch_id = ?
          AND company_id = ?
    """, (batch_order_id, batch_id, company_id))

    conn.commit()
    conn.close()

    flash("Batch order removed.", "success")
    return redirect(url_for("view_order_batch", batch_id=batch_id))


@app.route("/order-batches/<int:batch_id>/buy-materials", methods=["POST"])
@permission_required("manage_procurement")
def buy_order_batch_materials(batch_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    try:
        created_count = create_batch_purchase_requests(cursor, batch_id, company_id, user_id)
        conn.commit()
        conn.close()

        if created_count == 0:
            flash("No missing materials found or purchase requests already exist for this batch.", "info")
        else:
            flash(f"Created {created_count} purchase request(s) for batch materials.", "success")
    except ValueError as e:
        conn.rollback()
        conn.close()
        flash(str(e), "error")

    return redirect(url_for("view_order_batch", batch_id=batch_id))


@app.route("/order-batches/<int:batch_id>/launch", methods=["POST"])
@permission_required("manage_orders")
def launch_order_batch(batch_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    try:
        launched_count = launch_batch_orders(cursor, batch_id, company_id)
        conn.commit()
        conn.close()

        flash(f"Batch launched successfully. Activated {launched_count} order(s).", "success")
        return redirect(url_for("orders"))
    except ValueError as e:
        conn.rollback()
        conn.close()
        flash(str(e), "error")
        return redirect(url_for("view_order_batch", batch_id=batch_id))



@app.route("/orders/edit/<int:order_id>", methods=["GET", "POST"])
def edit_order(order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, order_number, product_id, quantity, status, due_date, priority
        FROM orders
        WHERE id = ? AND company_id = ?
    """, (order_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        return "Order not found", 404

    if request.method == "POST":
        order_number = request.form["order_number"].strip()
        product_id = int(request.form["product_id"])
        quantity = float(request.form["quantity"])
        status = request.form["status"].strip()
        due_date = request.form["due_date"].strip()
        priority = request.form["priority"].strip()

        try:
            require_company_record(cursor, "products", product_id, company_id, "Product not found.")

            cursor.execute("""
                UPDATE orders
                SET order_number = ?, product_id = ?, quantity = ?, status = ?, due_date = ?, priority = ?,
                    materials_consumed = 0,
                    finished_stock_added = 0
                WHERE id = ? AND company_id = ?
            """, (order_number, product_id, quantity, status, due_date, priority, order_id, company_id))

            rebuild_order_jobs(
                cursor,
                order_id,
                int(product_id),
                quantity,
                due_date,
                company_id=company_id
            )

            conn.commit()
            conn.close()

            flash("Order updated successfully.", "success")
            return redirect(url_for("orders"))
        except ValueError as e:
            conn.rollback()
            conn.close()
            flash(str(e), "error")
            return redirect(url_for("edit_order", order_id=order_id))

    order = {
        "id": row[0],
        "order_number": row[1],
        "product_id": row[2],
        "quantity": row[3] if row[3] is not None else 1,
        "status": row[4],
        "due_date": row[5],
        "priority": row[6]
    }

    cursor.execute("""
        SELECT id, product_code, product_name
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC
    """, (company_id,))
    products = cursor.fetchall()

    conn.close()

    return render_template(
        "edit_order.html",
        order=order,
        products=products,
        active_page="orders"
    )


@app.route("/orders/<int:order_id>/materials")
@permission_required("view_orders")
def order_materials(order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            o.id,
            o.order_number,
            o.quantity,
            p.id,
            p.product_name,
            COALESCE(pg.name, '')
        FROM orders o
        JOIN products p
          ON o.product_id = p.id
         AND p.company_id = o.company_id
        LEFT JOIN product_groups pg
          ON pg.id = p.group_id
         AND pg.company_id = p.company_id
        WHERE o.id = ? AND o.company_id = ?
    """, (order_id, company_id))

    order = cursor.fetchone()

    if not order:
        conn.close()
        return redirect(url_for("orders"))

    product_id = order[3]
    order_quantity = float(order[2] or 0)

    per_unit_exploded = explode_bom_items_recursive(cursor, product_id, 1, company_id=company_id)
    total_exploded = explode_bom_items_recursive(cursor, product_id, order_quantity, company_id=company_id)

    materials = []
    total_material_cost = 0

    item_ids = list(total_exploded.keys())

    if item_ids:
        placeholders = ",".join(["?"] * len(item_ids))
        cursor.execute(f"""
            SELECT id, item_name, item_code, measurement_unit, unit_price
            FROM items
            WHERE id IN ({placeholders}) AND company_id = ?
            ORDER BY item_name ASC
        """, item_ids + [company_id])

        for item_id, item_name, item_code, measurement_unit, unit_price in cursor.fetchall():
            bom_quantity = float(per_unit_exploded.get(item_id, 0) or 0)
            total_quantity = float(total_exploded.get(item_id, 0) or 0)
            unit_price = float(unit_price or 0)
            total_cost = total_quantity * unit_price
            total_material_cost += total_cost

            materials.append({
                "item_name": item_name,
                "item_code": item_code,
                "unit": measurement_unit,
                "bom_quantity": bom_quantity,
                "total_quantity": total_quantity,
                "unit_price": unit_price,
                "total_cost": total_cost
            })

    order_cost = get_order_cost_snapshot(cursor, product_id, order_quantity, company_id=company_id)
    conn.close()

    return render_template(
        "order_materials.html",
        order={
            "order_number": order[1],
            "quantity": order_quantity,
            "product_name": order[4],
            "group_name": order[5] or "Ungrouped",
        },
        materials=materials,
        total_material_cost=total_material_cost,
        total_jobs_cost=order_cost["jobs_cost"],
        total_cost=order_cost["total_cost"],
        active_page="orders"
    )


@app.route("/orders/delete/<int:order_id>", methods=["POST"])
def delete_order(order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    require_company_record(cursor, "orders", order_id, company_id)

    release_order_reservations(cursor, order_id, company_id=company_id)
    clear_order_shortages(cursor, order_id, company_id=company_id)

    cursor.execute("""
        DELETE FROM order_jobs
        WHERE order_id = ?
          AND company_id = ?
    """, (order_id, company_id))

    cursor.execute("""
        DELETE FROM orders
        WHERE id = ?
          AND company_id = ?
    """, (order_id, company_id))

    rebuild_company_reserved_quantities(cursor, company_id=company_id)

    conn.commit()
    conn.close()

    return redirect(url_for("orders"))

@app.route("/items")
def items():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    item_code = request.args.get("item_code", "").strip()
    item_name = request.args.get("item_name", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            id,
            item_code,
            item_name,
            description,
            measurement_unit,
            unit_price,
            stock_quantity,
            min_stock
        FROM items
        WHERE company_id = ?
    """
    params = [company_id]

    if item_code:
        query += " AND item_code LIKE ?"
        params.append(f"%{item_code}%")

    if item_name:
        query += " AND item_name LIKE ?"
        params.append(f"%{item_name}%")

    query += " ORDER BY id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    items = []
    for row in rows:
        items.append({
            "id": row[0],
            "item_code": row[1],
            "item_name": row[2],
            "description": row[3],
            "measurement_unit": row[4],
            "unit_price": row[5] if row[5] is not None else 0,
            "stock_quantity": row[6] if row[6] is not None else 0,
            "min_stock": row[7] if row[7] is not None else 0
        })

    filters = {
        "item_code": item_code,
        "item_name": item_name
    }

    return render_template(
        "items.html",
        items=items,
        filters=filters,
        active_page="items"
    )


@app.route("/items/new", methods=["GET", "POST"])
def new_item():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        item_code = request.form["item_code"].strip()
        item_name = request.form["item_name"].strip()
        description = request.form["description"].strip()
        measurement_unit = request.form["measurement_unit"].strip()
        unit_price = float(request.form["unit_price"] or 0)
        stock_quantity = float(request.form["stock_quantity"] or 0)
        min_stock = float(request.form["min_stock"] or 0)
        supplier_id_raw = request.form.get("supplier_id", "").strip()

        supplier_id = None
        if supplier_id_raw:
            try:
                supplier_id = int(supplier_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid supplier selected.", "error")
                return redirect(url_for("new_item"))

            cursor.execute("""
                SELECT id
                FROM suppliers
                WHERE id = ? AND company_id = ?
            """, (supplier_id, company_id))
            supplier_row = cursor.fetchone()

            if supplier_row is None:
                conn.close()
                flash("Selected supplier not found.", "error")
                return redirect(url_for("new_item"))

        cursor.execute("""
            INSERT INTO items (
                item_code,
                item_name,
                description,
                measurement_unit,
                unit_price,
                stock_quantity,
                min_stock,
                supplier_id,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            item_code,
            item_name,
            description,
            measurement_unit,
            unit_price,
            stock_quantity,
            min_stock,
            supplier_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Item created successfully.", "success")
        return redirect(url_for("items"))

    cursor.execute("""
        SELECT id, name
        FROM suppliers
        WHERE company_id = ? AND is_active = 1
        ORDER BY name ASC
    """, (company_id,))
    suppliers = cursor.fetchall()

    conn.close()

    return render_template(
        "new_item.html",
        suppliers=suppliers,
        active_page="items"
    )


@app.route("/items/edit/<int:item_id>", methods=["GET", "POST"])
def edit_item(item_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "items", item_id, company_id, "Item not found.")

        if request.method == "POST":
            item_code = request.form["item_code"].strip()
            item_name = request.form["item_name"].strip()
            description = request.form["description"].strip()
            measurement_unit = request.form["measurement_unit"].strip()
            unit_price = float(request.form["unit_price"] or 0)
            stock_quantity = float(request.form["stock_quantity"] or 0)
            min_stock = float(request.form["min_stock"] or 0)
            supplier_id_raw = request.form.get("supplier_id", "").strip()

            supplier_id = None
            if supplier_id_raw:
                try:
                    supplier_id = int(supplier_id_raw)
                except ValueError:
                    conn.close()
                    flash("Invalid supplier selected.", "error")
                    return redirect(url_for("edit_item", item_id=item_id))

                cursor.execute("""
                    SELECT id
                    FROM suppliers
                    WHERE id = ? AND company_id = ?
                """, (supplier_id, company_id))
                supplier_row = cursor.fetchone()

                if supplier_row is None:
                    conn.close()
                    flash("Selected supplier not found.", "error")
                    return redirect(url_for("edit_item", item_id=item_id))

            cursor.execute("""
                UPDATE items
                SET item_code = ?, item_name = ?, description = ?, measurement_unit = ?,
                    unit_price = ?, stock_quantity = ?, min_stock = ?, supplier_id = ?
                WHERE id = ? AND company_id = ?
            """, (
                item_code,
                item_name,
                description,
                measurement_unit,
                unit_price,
                stock_quantity,
                min_stock,
                supplier_id,
                item_id,
                company_id
            ))

            conn.commit()
            conn.close()

            flash("Item updated successfully.", "success")
            return redirect(url_for("items"))

        cursor.execute("""
            SELECT
                id,
                item_code,
                item_name,
                description,
                measurement_unit,
                unit_price,
                stock_quantity,
                min_stock,
                supplier_id
            FROM items
            WHERE id = ? AND company_id = ?
        """, (item_id, company_id))
        row = cursor.fetchone()

        cursor.execute("""
            SELECT id, name
            FROM suppliers
            WHERE company_id = ? AND is_active = 1
            ORDER BY name ASC
        """, (company_id,))
        supplier_rows = cursor.fetchall()

        conn.close()

        item = {
            "id": row[0],
            "item_code": row[1],
            "item_name": row[2],
            "description": row[3],
            "measurement_unit": row[4],
            "unit_price": row[5] if row[5] is not None else 0,
            "stock_quantity": row[6] if row[6] is not None else 0,
            "min_stock": row[7] if row[7] is not None else 0,
            "supplier_id": row[8]
        }

        return render_template(
            "edit_item.html",
            item=item,
            suppliers=supplier_rows,
            active_page="items"
        )
    except ValueError:
        conn.close()
        return "Item not found", 404


@app.route("/items/delete/<int:item_id>", methods=["POST"])
def delete_item(item_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "items", item_id, company_id, "Item not found.")

        cursor.execute("""
            DELETE FROM bom
            WHERE item_id = ? AND company_id = ?
        """, (item_id, company_id))

        cursor.execute("""
            DELETE FROM items
            WHERE id = ? AND company_id = ?
        """, (item_id, company_id))

        conn.commit()
        conn.close()

        flash("Item deleted successfully.", "info")
        return redirect(url_for("items"))
    except ValueError:
        conn.close()
        flash("Item not found.", "error")
        return redirect(url_for("items"))


@app.route("/products")
@permission_required("view_products")
def products():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    product_code = request.args.get("product_code", "").strip()
    product_name = request.args.get("product_name", "").strip()
    group_id = request.args.get("group_id", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            p.id,
            p.product_code,
            p.product_name,
            p.description,
            p.measurement_unit,
            p.time_per_unit,
            p.stock_quantity,
            COALESCE(pg.name, '') AS group_name,
            p.group_id
        FROM products p
        LEFT JOIN product_groups pg
          ON pg.id = p.group_id
         AND pg.company_id = p.company_id
        WHERE p.company_id = ?
    """
    params = [company_id]

    if product_code:
        query += " AND p.product_code LIKE ?"
        params.append(f"%{product_code}%")

    if product_name:
        query += " AND p.product_name LIKE ?"
        params.append(f"%{product_name}%")

    if group_id:
        query += " AND p.group_id = ?"
        params.append(group_id)

    query += " ORDER BY p.product_name ASC, p.product_code ASC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    product_groups = fetch_product_groups(cursor, company_id)

    products = []
    for row in rows:
        cost_snapshot = get_product_cost_snapshot(cursor, row[0], company_id=company_id)
        products.append({
            "id": row[0],
            "product_code": row[1],
            "product_name": row[2],
            "description": row[3],
            "measurement_unit": row[4],
            "time_per_unit": row[5] if row[5] is not None else 0,
            "stock_quantity": row[6] if row[6] is not None else 0,
            "group_name": row[7] or "Ungrouped",
            "group_id": row[8],
            "materials_cost": cost_snapshot["materials_cost"],
            "jobs_cost": cost_snapshot["jobs_cost"],
            "total_cost": cost_snapshot["total_cost"],
        })

    conn.close()

    filters = {
        "product_code": product_code,
        "product_name": product_name,
        "group_id": group_id,
    }

    return render_template(
        "products.html",
        products=products,
        product_groups=product_groups,
        filters=filters,
        active_page="products"
    )


@app.route("/products/new", methods=["GET", "POST"])
@permission_required("manage_products")
def new_product():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        product_code = request.form["product_code"].strip()
        product_name = request.form["product_name"].strip()
        description = request.form["description"].strip()
        measurement_unit = request.form["measurement_unit"].strip()
        time_per_unit = float(request.form["time_per_unit"] or 0)
        stock_quantity = float(request.form["stock_quantity"] or 0)
        group_id_raw = request.form.get("group_id", "").strip()
        group_id = int(group_id_raw) if group_id_raw else None

        cursor.execute("""
            INSERT INTO products (
                product_code,
                product_name,
                description,
                measurement_unit,
                time_per_unit,
                stock_quantity,
                group_id,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            product_code,
            product_name,
            description,
            measurement_unit,
            time_per_unit,
            stock_quantity,
            group_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Product created successfully.", "success")
        return redirect(url_for("products"))

    product_groups = fetch_product_groups(cursor, company_id)
    conn.close()
    return render_template(
        "new_product.html",
        product_groups=product_groups,
        active_page="products"
    )


@app.route("/products/edit/<int:product_id>", methods=["GET", "POST"])
@permission_required("manage_products")
def edit_product(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "products", product_id, company_id, "Product not found.")

        if request.method == "POST":
            product_code = request.form["product_code"].strip()
            product_name = request.form["product_name"].strip()
            description = request.form["description"].strip()
            measurement_unit = request.form["measurement_unit"].strip()
            time_per_unit = float(request.form["time_per_unit"] or 0)
            stock_quantity = float(request.form["stock_quantity"] or 0)
            group_id_raw = request.form.get("group_id", "").strip()
            group_id = int(group_id_raw) if group_id_raw else None

            cursor.execute("""
                UPDATE products
                SET product_code = ?, product_name = ?, description = ?, measurement_unit = ?,
                    time_per_unit = ?, stock_quantity = ?, group_id = ?
                WHERE id = ? AND company_id = ?
            """, (
                product_code,
                product_name,
                description,
                measurement_unit,
                time_per_unit,
                stock_quantity,
                group_id,
                product_id,
                company_id
            ))

            conn.commit()
            conn.close()

            flash("Product updated successfully.", "success")
            return redirect(url_for("products"))

        cursor.execute("""
            SELECT
                id,
                product_code,
                product_name,
                description,
                measurement_unit,
                time_per_unit,
                stock_quantity,
                group_id
            FROM products
            WHERE id = ? AND company_id = ?
        """, (product_id, company_id))
        row = cursor.fetchone()
        product_groups = fetch_product_groups(cursor, company_id)
        conn.close()

        product = {
            "id": row[0],
            "product_code": row[1],
            "product_name": row[2],
            "description": row[3],
            "measurement_unit": row[4],
            "time_per_unit": row[5] if row[5] is not None else 0,
            "stock_quantity": row[6] if row[6] is not None else 0,
            "group_id": row[7],
        }

        return render_template(
            "edit_product.html",
            product=product,
            product_groups=product_groups,
            active_page="products"
        )
    except ValueError:
        conn.close()
        return "Product not found", 404


@app.route("/product-groups", methods=["GET", "POST"])
@permission_required("manage_products")
def product_groups():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Group name is required.", "error")
            conn.close()
            return redirect(url_for("product_groups"))
        try:
            cursor.execute(
                "INSERT INTO product_groups (company_id, name, description) VALUES (?, ?, ?)",
                (company_id, name, description or None),
            )
            conn.commit()
            flash("Product group created.", "success")
        except sqlite3.IntegrityError:
            flash("A product group with this name already exists.", "error")
        conn.close()
        return redirect(url_for("product_groups"))

    groups = fetch_product_groups(cursor, company_id)
    for group in groups:
        cursor.execute(
            "SELECT COUNT(*) FROM products WHERE company_id = ? AND group_id = ?",
            (company_id, group["id"]),
        )
        group["product_count"] = int(cursor.fetchone()[0] or 0)
    conn.close()
    return render_template("product_groups.html", groups=groups, active_page="products")


@app.route("/product-groups/<int:group_id>/edit", methods=["GET", "POST"])
@permission_required("manage_products")
def edit_product_group(group_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, name, COALESCE(description, '') FROM product_groups WHERE id = ? AND company_id = ?",
        (group_id, company_id),
    )
    row = cursor.fetchone()
    if row is None:
        conn.close()
        flash("Product group not found.", "error")
        return redirect(url_for("product_groups"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            conn.close()
            flash("Group name is required.", "error")
            return redirect(url_for("edit_product_group", group_id=group_id))
        try:
            cursor.execute(
                "UPDATE product_groups SET name = ?, description = ? WHERE id = ? AND company_id = ?",
                (name, description or None, group_id, company_id),
            )
            conn.commit()
            flash("Product group updated.", "success")
            conn.close()
            return redirect(url_for("product_groups"))
        except sqlite3.IntegrityError:
            conn.close()
            flash("A product group with this name already exists.", "error")
            return redirect(url_for("edit_product_group", group_id=group_id))

    group = {"id": row[0], "name": row[1], "description": row[2]}
    conn.close()
    return render_template("edit_product_group.html", group=group, active_page="products")


@app.route("/product-groups/<int:group_id>/delete", methods=["POST"])
@permission_required("manage_products")
def delete_product_group(group_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET group_id = NULL WHERE group_id = ? AND company_id = ?", (group_id, company_id))
    cursor.execute("DELETE FROM product_groups WHERE id = ? AND company_id = ?", (group_id, company_id))
    conn.commit()
    conn.close()
    flash("Product group deleted.", "success")
    return redirect(url_for("product_groups"))


@app.route("/products/delete/<int:product_id>", methods=["POST"])
def delete_product(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "products", product_id, company_id, "Product not found.")

        cursor.execute("""
            DELETE FROM bom
            WHERE product_id = ? AND company_id = ?
        """, (product_id, company_id))

        cursor.execute("""
            DELETE FROM bom
            WHERE child_product_id = ? AND company_id = ?
        """, (product_id, company_id))

        cursor.execute("""
            DELETE FROM product_job_templates
            WHERE product_id = ? AND company_id = ?
        """, (product_id, company_id))

        cursor.execute("""
            DELETE FROM products
            WHERE id = ? AND company_id = ?
        """, (product_id, company_id))

        conn.commit()
        conn.close()

        flash("Product deleted successfully.", "info")
        return redirect(url_for("products"))
    except ValueError:
        conn.close()
        flash("Product not found.", "error")
        return redirect(url_for("products"))

@app.route("/products/<int:product_id>/jobs")
def product_jobs(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, product_code, product_name, description, measurement_unit, time_per_unit
        FROM products
        WHERE id = ? AND company_id = ?
    """, (product_id, company_id))
    product_row = cursor.fetchone()

    if product_row is None:
        conn.close()
        return "Product not found", 404

    product = {
        "id": product_row[0],
        "product_code": product_row[1],
        "product_name": product_row[2],
        "description": product_row[3],
        "measurement_unit": product_row[4],
        "time_per_unit": product_row[5]
    }

    cursor.execute("""
        SELECT
            pjt.id,
            pjt.job_name,
            pjt.sequence,
            pjt.estimated_hours,
            pjt.workstation_id,
            w.name
        FROM product_job_templates pjt
        JOIN workstations w
          ON pjt.workstation_id = w.id
         AND w.company_id = pjt.company_id
        WHERE pjt.product_id = ?
          AND pjt.company_id = ?
        ORDER BY pjt.sequence ASC, pjt.id ASC
    """, (product_id, company_id))
    rows = cursor.fetchall()

    job_templates = []
    for row in rows:
        job_templates.append({
            "id": row[0],
            "job_name": row[1],
            "sequence": row[2],
            "estimated_hours": row[3],
            "workstation_id": row[4],
            "workstation_name": row[5]
        })

    cursor.execute("""
        SELECT id, name
        FROM workstations
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    workstation_rows = cursor.fetchall()

    workstations = [{"id": row[0], "name": row[1]} for row in workstation_rows]

    conn.close()

    return render_template(
        "product_jobs.html",
        product=product,
        job_templates=job_templates,
        workstations=workstations,
        active_page="products"
    )


@app.route("/products/<int:product_id>/jobs/add", methods=["POST"])
def add_product_job(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    job_name = request.form["job_name"].strip()
    workstation_id = int(request.form["workstation_id"])
    sequence = int(request.form["sequence"])
    estimated_hours = float(request.form["estimated_hours"] or 0)

    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "products", product_id, company_id, "Product not found.")
        require_company_record(cursor, "workstations", workstation_id, company_id, "Workstation not found.")

        cursor.execute("""
            INSERT INTO product_job_templates (
                product_id,
                workstation_id,
                job_name,
                sequence,
                estimated_hours,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (product_id, workstation_id, job_name, sequence, estimated_hours, company_id))

        conn.commit()
        conn.close()

        flash("Product job template added successfully.", "success")
        return redirect(url_for("product_jobs", product_id=product_id))
    except ValueError as e:
        conn.close()
        flash(str(e), "error")
        return redirect(url_for("product_jobs", product_id=product_id))


@app.route("/products/jobs/edit/<int:job_id>", methods=["GET", "POST"])
def edit_product_job(job_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, product_id, workstation_id, job_name, sequence, estimated_hours
        FROM product_job_templates
        WHERE id = ? AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        return "Product job template not found", 404

    if request.method == "POST":
        job_name = request.form["job_name"].strip()
        workstation_id = int(request.form["workstation_id"])
        sequence = int(request.form["sequence"])
        estimated_hours = float(request.form["estimated_hours"] or 0)

        try:
            require_company_record(cursor, "workstations", workstation_id, company_id, "Workstation not found.")

            cursor.execute("""
                UPDATE product_job_templates
                SET job_name = ?, workstation_id = ?, sequence = ?, estimated_hours = ?
                WHERE id = ? AND company_id = ?
            """, (job_name, workstation_id, sequence, estimated_hours, job_id, company_id))

            conn.commit()
            product_id = row[1]
            conn.close()

            flash("Product job template updated successfully.", "success")
            return redirect(url_for("product_jobs", product_id=product_id))
        except ValueError as e:
            conn.close()
            flash(str(e), "error")
            return redirect(url_for("product_jobs", product_id=row[1]))

    job_template = {
        "id": row[0],
        "product_id": row[1],
        "workstation_id": row[2],
        "job_name": row[3],
        "sequence": row[4],
        "estimated_hours": row[5]
    }

    cursor.execute("""
        SELECT id, name
        FROM workstations
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    workstation_rows = cursor.fetchall()
    workstations = [{"id": r[0], "name": r[1]} for r in workstation_rows]

    cursor.execute("""
        SELECT id, product_code, product_name
        FROM products
        WHERE id = ? AND company_id = ?
    """, (job_template["product_id"], company_id))
    product_row = cursor.fetchone()
    conn.close()

    product = {
        "id": product_row[0],
        "product_code": product_row[1],
        "product_name": product_row[2]
    }

    return render_template(
        "edit_product_job.html",
        job_template=job_template,
        workstations=workstations,
        product=product,
        active_page="products"
    )


@app.route("/products/jobs/delete/<int:job_id>", methods=["POST"])
def delete_product_job(job_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT product_id
        FROM product_job_templates
        WHERE id = ? AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Product job template not found.", "error")
        return redirect(url_for("products"))

    product_id = row[0]

    cursor.execute("""
        DELETE FROM product_job_templates
        WHERE id = ? AND company_id = ?
    """, (job_id, company_id))

    conn.commit()
    conn.close()

    flash("Product job template deleted successfully.", "info")
    return redirect(url_for("product_jobs", product_id=product_id))


@app.route("/bom/<int:product_id>")
def product_bom(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, product_code, product_name, description, measurement_unit, time_per_unit
        FROM products
        WHERE id = ? AND company_id = ?
    """, (product_id, company_id))
    product_row = cursor.fetchone()

    if product_row is None:
        conn.close()
        return "Product not found", 404

    product = {
        "id": product_row[0],
        "product_code": product_row[1],
        "product_name": product_row[2],
        "description": product_row[3],
        "measurement_unit": product_row[4],
        "time_per_unit": product_row[5]
    }

    cursor.execute("""
        SELECT
            bom.id,
            bom.component_type,
            bom.quantity,
            items.item_code,
            items.item_name,
            items.measurement_unit,
            products.product_code,
            products.product_name,
            products.measurement_unit
        FROM bom
        LEFT JOIN items
          ON bom.item_id = items.id
         AND bom.component_type = 'item'
         AND items.company_id = bom.company_id
        LEFT JOIN products
          ON bom.child_product_id = products.id
         AND bom.component_type = 'product'
         AND products.company_id = bom.company_id
        WHERE bom.product_id = ?
          AND bom.company_id = ?
        ORDER BY bom.id DESC
    """, (product_id, company_id))
    bom_rows = cursor.fetchall()

    bom_items = []
    for row in bom_rows:
        if row[1] == "product":
            bom_items.append({
                "id": row[0],
                "component_type": "product",
                "component_code": row[6],
                "component_name": row[7],
                "measurement_unit": row[8],
                "quantity": row[2]
            })
        else:
            bom_items.append({
                "id": row[0],
                "component_type": "item",
                "component_code": row[3],
                "component_name": row[4],
                "measurement_unit": row[5],
                "quantity": row[2]
            })

    cursor.execute("""
        SELECT id, item_code, item_name, measurement_unit
        FROM items
        WHERE company_id = ?
        ORDER BY item_name ASC
    """, (company_id,))
    item_rows = cursor.fetchall()
    items = [{"id": r[0], "item_code": r[1], "item_name": r[2], "measurement_unit": r[3]} for r in item_rows]

    cursor.execute("""
        SELECT id, product_code, product_name, measurement_unit
        FROM products
        WHERE id != ?
          AND company_id = ?
        ORDER BY product_name ASC
    """, (product_id, company_id))
    product_rows = cursor.fetchall()
    child_products = [{"id": r[0], "product_code": r[1], "product_name": r[2], "measurement_unit": r[3]} for r in product_rows]

    conn.close()

    return render_template(
        "bom.html",
        product=product,
        bom_items=bom_items,
        items=items,
        child_products=child_products,
        active_page="products"
    )


@app.route("/products/<int:product_id>/cost")
@permission_required("view_products")
def product_cost(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, product_code, product_name, measurement_unit, COALESCE(pg.name, '')
        FROM products p
        LEFT JOIN product_groups pg
          ON pg.id = p.group_id
         AND pg.company_id = p.company_id
        WHERE p.id = ? AND p.company_id = ?
    """, (product_id, company_id))
    product_row = cursor.fetchone()

    if product_row is None:
        conn.close()
        return "Product not found", 404

    product = {
        "id": product_row[0],
        "product_code": product_row[1],
        "product_name": product_row[2],
        "measurement_unit": product_row[3],
        "group_name": product_row[4] or "Ungrouped",
    }

    cost_snapshot = get_product_cost_snapshot(cursor, product_id, company_id=company_id)

    conn.close()

    return render_template(
        "product_cost.html",
        product=product,
        materials=cost_snapshot["materials"],
        jobs=cost_snapshot["jobs"],
        total_material_cost=cost_snapshot["materials_cost"],
        total_jobs_cost=cost_snapshot["jobs_cost"],
        total_cost=cost_snapshot["total_cost"],
        active_page="products"
    )


@app.route("/bom/<int:product_id>/add", methods=["POST"])
def add_bom_item(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    component_type = request.form.get("component_type", "item").strip()
    quantity = float(request.form.get("quantity", 0) or 0)

    item_id = request.form.get("item_id")
    child_product_id = request.form.get("child_product_id")

    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "products", product_id, company_id, "Product not found.")

        if component_type == "product":
            if not child_product_id:
                raise ValueError("Child product is required.")
            child_product_id = int(child_product_id)
            require_company_record(cursor, "products", child_product_id, company_id, "Child product not found.")

            cursor.execute("""
                INSERT INTO bom (
                    product_id,
                    item_id,
                    quantity,
                    component_type,
                    child_product_id,
                    company_id
                )
                VALUES (?, ?, ?, ?, ?, ?)
            """, (product_id, 0, quantity, "product", child_product_id, company_id))
        else:
            if not item_id:
                raise ValueError("Item is required.")
            item_id = int(item_id)
            require_company_record(cursor, "items", item_id, company_id, "Item not found.")

            cursor.execute("""
                INSERT INTO bom (
                    product_id,
                    item_id,
                    quantity,
                    component_type,
                    child_product_id,
                    company_id
                )
                VALUES (?, ?, ?, ?, ?, ?)
            """, (product_id, item_id, quantity, "item", None, company_id))

        conn.commit()
        conn.close()

        flash("BOM item added successfully.", "success")
        return redirect(url_for("product_bom", product_id=product_id))
    except ValueError as e:
        conn.close()
        flash(str(e), "error")
        return redirect(url_for("product_bom", product_id=product_id))


@app.route("/bom/delete/<int:bom_id>", methods=["POST"])
def delete_bom_item(bom_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT product_id
        FROM bom
        WHERE id = ? AND company_id = ?
    """, (bom_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("BOM row not found.", "error")
        return redirect(url_for("products"))

    product_id = row[0]

    cursor.execute("""
        DELETE FROM bom
        WHERE id = ? AND company_id = ?
    """, (bom_id, company_id))

    conn.commit()
    conn.close()

    flash("BOM item deleted successfully.", "info")
    return redirect(url_for("product_bom", product_id=product_id))


@app.route("/workstations/new", methods=["GET", "POST"])
@permission_required("manage_workstations")
def new_workstation():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form["name"].strip()
        description = request.form["description"].strip()
        hours_per_shift = float(request.form["hours_per_shift"] or 8)
        shifts_per_day = int(request.form["shifts_per_day"] or 1)
        working_days_per_month = int(request.form["working_days_per_month"] or 20)
        cost_per_hour = float(request.form.get("cost_per_hour", 0) or 0)
        color = request.form.get("color", "#3b82f6").strip() or "#3b82f6"
        group_id_raw = request.form.get("group_id", "").strip()

        group_id = None
        if group_id_raw:
            try:
                group_id = int(group_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid workstation group selected.", "error")
                return redirect(url_for("new_workstation"))

            cursor.execute("""
                SELECT id
                FROM workstation_groups
                WHERE id = ?
                  AND company_id = ?
            """, (group_id, company_id))
            group_exists = cursor.fetchone()

            if group_exists is None:
                conn.close()
                flash("Selected workstation group was not found.", "error")
                return redirect(url_for("new_workstation"))

        cursor.execute("""
            INSERT INTO workstations (
                name,
                description,
                hours_per_shift,
                shifts_per_day,
                working_days_per_month,
                color,
                cost_per_hour,
                group_id,
                company_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            name,
            description,
            hours_per_shift,
            shifts_per_day,
            working_days_per_month,
            color,
            cost_per_hour,
            group_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Workstation created successfully.", "success")
        return redirect(url_for("workstations"))

    workstation_groups = fetch_workstation_groups(cursor, company_id)
    conn.close()

    return render_template(
        "new_workstation.html",
        workstation_groups=workstation_groups,
        active_page="workstations"
    )


@app.route("/workstations/edit/<int:workstation_id>", methods=["GET", "POST"])
@permission_required("manage_workstations")
def edit_workstation(workstation_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            name,
            description,
            hours_per_shift,
            shifts_per_day,
            working_days_per_month,
            color,
            COALESCE(cost_per_hour, 0),
            group_id
        FROM workstations
        WHERE id = ? AND company_id = ?
    """, (workstation_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        return "Workstation not found", 404

    if request.method == "POST":
        name = request.form["name"].strip()
        description = request.form["description"].strip()
        hours_per_shift = float(request.form["hours_per_shift"] or 8)
        shifts_per_day = int(request.form["shifts_per_day"] or 1)
        working_days_per_month = int(request.form["working_days_per_month"] or 20)
        cost_per_hour = float(request.form.get("cost_per_hour", 0) or 0)
        color = request.form.get("color", "#3b82f6").strip() or "#3b82f6"
        group_id_raw = request.form.get("group_id", "").strip()

        group_id = None
        if group_id_raw:
            try:
                group_id = int(group_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid workstation group selected.", "error")
                return redirect(url_for("edit_workstation", workstation_id=workstation_id))

            cursor.execute("""
                SELECT id
                FROM workstation_groups
                WHERE id = ?
                  AND company_id = ?
            """, (group_id, company_id))
            group_exists = cursor.fetchone()

            if group_exists is None:
                conn.close()
                flash("Selected workstation group was not found.", "error")
                return redirect(url_for("edit_workstation", workstation_id=workstation_id))

        cursor.execute("""
            UPDATE workstations
            SET name = ?,
                description = ?,
                hours_per_shift = ?,
                shifts_per_day = ?,
                working_days_per_month = ?,
                color = ?,
                cost_per_hour = ?,
                group_id = ?
            WHERE id = ? AND company_id = ?
        """, (
            name,
            description,
            hours_per_shift,
            shifts_per_day,
            working_days_per_month,
            color,
            cost_per_hour,
            group_id,
            workstation_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Workstation updated successfully.", "success")
        return redirect(url_for("workstations"))

    workstation = {
        "id": row[0],
        "name": row[1],
        "description": row[2],
        "hours_per_shift": row[3],
        "shifts_per_day": row[4],
        "working_days_per_month": row[5],
        "color": row[6],
        "cost_per_hour": float(row[7] or 0),
        "group_id": row[8],
    }

    workstation_groups = fetch_workstation_groups(cursor, company_id)
    conn.close()

    return render_template(
        "edit_workstation.html",
        workstation=workstation,
        workstation_groups=workstation_groups,
        active_page="workstations"
    )


@app.route("/workstations/delete/<int:workstation_id>", methods=["POST"])
def delete_workstation(workstation_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*)
        FROM product_job_templates
        WHERE workstation_id = ? AND company_id = ?
    """, (workstation_id, company_id))
    template_count = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM order_jobs
        WHERE workstation_id = ? AND company_id = ?
    """, (workstation_id, company_id))
    job_count = cursor.fetchone()[0]

    if template_count > 0 or job_count > 0:
        conn.close()
        flash("Cannot delete workstation that is used in jobs or templates.", "error")
        return redirect(url_for("workstations"))

    cursor.execute("""
        DELETE FROM workstations
        WHERE id = ? AND company_id = ?
    """, (workstation_id, company_id))

    conn.commit()
    conn.close()

    flash("Workstation deleted successfully.", "info")
    return redirect(url_for("workstations"))



@app.route("/jobs")
@permission_required("view_jobs")
def jobs():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    order_number = request.args.get("order_number", "").strip()
    product_name = request.args.get("product_name", "").strip()
    job_name = request.args.get("job_name", "").strip()
    workstation = request.args.get("workstation", "").strip()
    workstation_text = request.args.get("workstation_text", "").strip()
    due_date_from = request.args.get("due_date_from", "").strip()
    due_date_to = request.args.get("due_date_to", "").strip()

    selected_statuses = request.args.getlist("status")
    selected_statuses = [s.strip() for s in selected_statuses if s.strip()]

    statuses = selected_statuses[:]
    if "All" in statuses:
        statuses = []

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            oj.id,
            o.order_number,
            jp.product_name,
            oj.job_name,
            w.name,
            oj.workstation_id,
            oj.sequence,
            oj.planned_quantity,
            oj.completed_quantity,
            oj.estimated_hours,
            oj.status,
            o.due_date,
            oj.planned_start,
            oj.planned_end,
            oj.parent_job_id,
            oj.is_split_child,
            (
                SELECT COUNT(*)
                FROM order_jobs child
                WHERE child.parent_job_id = oj.id
                  AND child.is_split_child = 1
                  AND child.company_id = oj.company_id
            ) AS child_count
        FROM order_jobs oj
        JOIN orders o
          ON oj.order_id = o.id
         AND o.company_id = oj.company_id
        LEFT JOIN products jp
          ON oj.job_product_id = jp.id
         AND jp.company_id = oj.company_id
        JOIN workstations w
          ON oj.workstation_id = w.id
         AND w.company_id = oj.company_id
        WHERE oj.company_id = ?
    """
    params = [company_id]

    if order_number:
        query += " AND o.order_number LIKE ?"
        params.append(f"%{order_number}%")

    if product_name:
        query += " AND jp.product_name LIKE ?"
        params.append(f"%{product_name}%")

    if job_name:
        query += " AND oj.job_name LIKE ?"
        params.append(f"%{job_name}%")

    if workstation:
        query += " AND oj.workstation_id = ?"
        params.append(workstation)

    if workstation_text:
        query += " AND w.name LIKE ?"
        params.append(f"%{workstation_text}%")

    if statuses:
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND oj.status IN ({placeholders})"
        params.extend(statuses)

    if due_date_from:
        query += " AND o.due_date >= ?"
        params.append(due_date_from)

    if due_date_to:
        query += " AND o.due_date <= ?"
        params.append(due_date_to)

    query += " ORDER BY o.due_date ASC, o.order_number ASC, oj.sequence ASC, oj.id ASC"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    jobs = []
    for row in rows:
        job_id = row[0]
        planned_quantity = float(row[7] or 0)
        completed_quantity = float(row[8] or 0)
        child_count = int(row[16] or 0)

        progress_percent = 0
        if planned_quantity > 0:
            progress_percent = min(100, max(0, (completed_quantity / planned_quantity) * 100))

        jobs.append({
            "id": row[0],
            "order_number": row[1],
            "product_name": row[2] if row[2] else "-",
            "job_name": row[3],
            "workstation_name": row[4],
            "workstation_id": row[5],
            "sequence": row[6],
            "planned_quantity": planned_quantity,
            "completed_quantity": completed_quantity,
            "estimated_hours": row[9],
            "status": row[10],
            "due_date": row[11],
            "planned_start": row[12],
            "planned_end": row[13],
            "progress_percent": progress_percent,
            "parent_job_id": row[14],
            "is_split_child": int(row[15] or 0),
            "child_count": child_count,
            "can_start": can_start_job(cursor, job_id, company_id=company_id) if child_count == 0 else False
        })

    cursor.execute("""
        SELECT id, name
        FROM workstations
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    ws_rows = cursor.fetchall()
    conn.close()

    workstations = [{"id": w[0], "name": w[1]} for w in ws_rows]

    filters = {
        "order_number": order_number,
        "product_name": product_name,
        "job_name": job_name,
        "workstation": workstation,
        "workstation_text": workstation_text,
        "status": selected_statuses,
        "due_date_from": due_date_from,
        "due_date_to": due_date_to
    }

    return render_template(
        "jobs.html",
        jobs=jobs,
        workstations=workstations,
        filters=filters,
        active_page="jobs"
    )


@app.route("/jobs/update_workstation/<int:job_id>", methods=["POST"])
def update_job_workstation(job_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    new_workstation_id = int(request.form["workstation_id"])

    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "order_jobs", job_id, company_id, "Job not found.")
        require_company_record(cursor, "workstations", new_workstation_id, company_id, "Workstation not found.")

        cursor.execute("""
            UPDATE order_jobs
            SET workstation_id = ?
            WHERE id = ? AND company_id = ?
        """, (new_workstation_id, job_id, company_id))

        recalculate_job_dates(cursor, job_id)

        conn.commit()
        conn.close()

        flash("Workstation updated.", "success")
        return redirect(request.referrer or url_for("jobs"))
    except ValueError as e:
        conn.close()
        flash(str(e), "error")
        return redirect(request.referrer or url_for("jobs"))



@app.route("/jobs/update_status/<int:job_id>/<new_status>", methods=["POST"])
def update_job_status(job_id, new_status):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    allowed_statuses = ["Waiting", "Ongoing", "Paused", "Done"]

    if new_status not in allowed_statuses:
        flash("Invalid status.", "error")
        return redirect_back("jobs")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            order_id,
            parent_job_id,
            job_product_id,
            planned_quantity,
            completed_quantity,
            status
        FROM order_jobs
        WHERE id = ? AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Job not found.", "error")
        return redirect_back("jobs")

    order_id = row[0]
    parent_job_id = row[1]
    job_product_id = row[2]
    planned_quantity = float(row[3] or 0)
    completed_quantity = float(row[4] or 0)

    if new_status == "Ongoing":
        if not can_start_job(cursor, job_id, company_id=company_id):
            conn.close()
            flash("Cannot start this job yet. Previous sequence jobs are not done.", "error")
            return redirect_back("jobs")

        cursor.execute("""
            UPDATE order_jobs
            SET status = ?
            WHERE id = ? AND company_id = ?
        """, (new_status, job_id, company_id))

    elif new_status == "Done":
        if completed_quantity < planned_quantity:
            completed_quantity = planned_quantity
            cursor.execute("""
                UPDATE order_jobs
                SET completed_quantity = ?, status = ?
                WHERE id = ? AND company_id = ?
            """, (completed_quantity, "Done", job_id, company_id))
        else:
            cursor.execute("""
                UPDATE order_jobs
                SET status = ?
                WHERE id = ? AND company_id = ?
            """, ("Done", job_id, company_id))

        if is_final_job(cursor, order_id, job_id, company_id=company_id):
            cursor.execute("""
                SELECT
                    product_id,
                    quantity,
                    finished_stock_added,
                    materials_consumed
                FROM orders
                WHERE id = ?
                  AND company_id = ?
            """, (order_id, company_id))
            order_row = cursor.fetchone()

            if order_row is not None:
                root_product_id = order_row[0]
                root_order_quantity = float(order_row[1] or 0)
                finished_stock_added = int(order_row[2] or 0)
                materials_consumed = int(order_row[3] or 0)

                if materials_consumed == 0:
                    consume_job_materials(
                        cursor,
                        root_product_id,
                        root_order_quantity,
                        company_id=company_id
                    )

                if finished_stock_added == 0:
                    add_finished_product_stock(
                        cursor,
                        root_product_id,
                        root_order_quantity,
                        company_id=company_id
                    )

            release_order_reservations(cursor, order_id, company_id=company_id)
            clear_order_shortages(cursor, order_id, company_id=company_id)
            rebuild_company_reserved_quantities(cursor, company_id=company_id)

            cursor.execute("""
                UPDATE orders
                SET finished_stock_added = 1,
                    materials_consumed = 1,
                    materials_reserved = 0
                WHERE id = ?
                  AND company_id = ?
            """, (order_id, company_id))

    else:
        cursor.execute("""
            UPDATE order_jobs
            SET status = ?
            WHERE id = ? AND company_id = ?
        """, (new_status, job_id, company_id))

    if parent_job_id:
        sync_parent_job_status(cursor, parent_job_id, company_id=company_id)

    sync_order_status(cursor, order_id, company_id=company_id)

    conn.commit()
    conn.close()

    flash(f"Job status updated to {new_status}.", "success")
    return redirect_back("jobs")


@app.route("/jobs/update_progress/<int:job_id>", methods=["POST"])
@permission_required("update_job_progress")
def update_job_progress(job_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    completed_quantity = float(request.form.get("completed_quantity", 0) or 0)

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT planned_quantity, parent_job_id, order_id
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Job not found.", "error")
        return redirect_back("jobs")

    planned_quantity = float(row[0] or 0)
    parent_job_id = row[1]
    order_id = row[2]

    if completed_quantity < 0:
        completed_quantity = 0

    if completed_quantity > planned_quantity:
        completed_quantity = planned_quantity

    if completed_quantity >= planned_quantity and planned_quantity > 0:
        new_status = "Done"
    elif completed_quantity > 0:
        new_status = "Ongoing"
    else:
        new_status = "Waiting"

    cursor.execute("""
        UPDATE order_jobs
        SET completed_quantity = ?, status = ?
        WHERE id = ?
          AND company_id = ?
    """, (completed_quantity, new_status, job_id, company_id))

    recalculate_job_dates(cursor, job_id)

    if parent_job_id:
        sync_parent_job_status(cursor, parent_job_id, company_id=company_id)

    sync_order_status(cursor, order_id, company_id=company_id)

    conn.commit()
    conn.close()

    flash("Job progress updated.", "success")
    return redirect_back("jobs")



@app.route("/transfers")
@permission_required("view_inventory")
def product_transfers():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    product_id = request.args.get("product_id", "").strip()
    destination_id = request.args.get("destination_id", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            product_code,
            product_name,
            measurement_unit,
            COALESCE(stock_quantity, 0) AS stock_quantity
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC, product_code ASC
    """, (company_id,))
    product_rows = cursor.fetchall()

    cursor.execute("""
        SELECT
            id,
            name,
            destination_code,
            notes
        FROM stock_destinations
        WHERE company_id = ?
          AND COALESCE(is_active, 1) = 1
        ORDER BY name ASC
    """, (company_id,))
    destination_rows = cursor.fetchall()

    where_parts = ["t.company_id = ?"]
    params = [company_id]

    if product_id:
        where_parts.append("t.product_id = ?")
        params.append(product_id)

    if destination_id:
        where_parts.append("t.destination_id = ?")
        params.append(destination_id)

    if date_from:
        where_parts.append("date(t.transfer_date) >= date(?)")
        params.append(date_from)

    if date_to:
        where_parts.append("date(t.transfer_date) <= date(?)")
        params.append(date_to)

    where_sql = " AND ".join(where_parts)

    cursor.execute(f"""
        SELECT
            t.id,
            t.transfer_date,
            t.quantity,
            t.notes,
            p.product_code,
            p.product_name,
            p.measurement_unit,
            d.name AS destination_name,
            d.destination_code,
            u.full_name AS created_by_name
        FROM product_transfers_out t
        JOIN products p
          ON p.id = t.product_id
         AND p.company_id = t.company_id
        JOIN stock_destinations d
          ON d.id = t.destination_id
         AND d.company_id = t.company_id
        LEFT JOIN users u
          ON u.id = t.created_by
        WHERE {where_sql}
        ORDER BY t.transfer_date DESC, t.id DESC
    """, tuple(params))
    transfer_rows = cursor.fetchall()

    cursor.execute(f"""
        SELECT
            COUNT(*) AS transfer_count,
            COALESCE(SUM(t.quantity), 0) AS total_quantity
        FROM product_transfers_out t
        WHERE {where_sql}
    """, tuple(params))
    stats_row = cursor.fetchone()

    destination_stock_where = ["t.company_id = ?"]
    destination_stock_params = [company_id]

    if destination_id:
        destination_stock_where.append("t.destination_id = ?")
        destination_stock_params.append(destination_id)

    if product_id:
        destination_stock_where.append("t.product_id = ?")
        destination_stock_params.append(product_id)

    destination_stock_where_sql = " AND ".join(destination_stock_where)

    cursor.execute(f"""
        SELECT
            d.id AS destination_id,
            d.name AS destination_name,
            d.destination_code,
            p.id AS product_id,
            p.product_code,
            p.product_name,
            COALESCE(p.measurement_unit, 'pcs') AS measurement_unit,
            ROUND(COALESCE(SUM(t.quantity), 0)::numeric, 2) AS qty_at_destination
        FROM product_transfers_out t
        JOIN stock_destinations d
          ON d.id = t.destination_id
         AND d.company_id = t.company_id
        JOIN products p
          ON p.id = t.product_id
         AND p.company_id = t.company_id
        WHERE {destination_stock_where_sql}
        GROUP BY
            d.id, d.name, d.destination_code,
            p.id, p.product_code, p.product_name, p.measurement_unit
        HAVING COALESCE(SUM(t.quantity), 0) > 0
        ORDER BY d.name ASC, p.product_name ASC, p.product_code ASC
    """, tuple(destination_stock_params))
    destination_stock_rows = cursor.fetchall()

    cursor.execute("""
        SELECT
            d.id AS destination_id,
            d.name AS destination_name,
            d.destination_code,
            ROUND(COALESCE(SUM(t.quantity), 0)::numeric, 2) AS total_qty
        FROM stock_destinations d
        LEFT JOIN product_transfers_out t
          ON t.destination_id = d.id
         AND t.company_id = d.company_id
        WHERE d.company_id = ?
          AND COALESCE(d.is_active, 1) = 1
        GROUP BY d.id, d.name, d.destination_code
        ORDER BY d.name ASC
    """, (company_id,))
    destination_total_rows = cursor.fetchall()

    conn.close()

    products = [dict(row) for row in product_rows]
    destinations = [dict(row) for row in destination_rows]
    transfers = [dict(row) for row in transfer_rows]
    destination_stock = [dict(row) for row in destination_stock_rows]
    destination_totals = [dict(row) for row in destination_total_rows]

    stats = {
        "destination_count": len(destinations),
        "transfer_count": stats_row["transfer_count"] if stats_row else 0,
        "total_quantity": stats_row["total_quantity"] if stats_row else 0,
    }

    filters = {
        "product_id": product_id,
        "destination_id": destination_id,
        "date_from": date_from,
        "date_to": date_to,
    }

    return render_template(
        "product_transfers.html",
        products=products,
        destinations=destinations,
        transfers=transfers,
        destination_stock=destination_stock,
        destination_totals=destination_totals,
        stats=stats,
        filters=filters,
        today=datetime.now().strftime("%Y-%m-%d"),
        active_page="transfers",
    )


@app.route("/transfers/destinations/new", methods=["POST"])
@permission_required("manage_inventory")
def create_stock_destination():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    name = request.form.get("name", "").strip()
    destination_code = request.form.get("destination_code", "").strip()
    notes = request.form.get("notes", "").strip()

    if not name:
        flash("Destination name is required.", "error")
        return redirect(url_for("product_transfers"))

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM stock_destinations
        WHERE company_id = ?
          AND LOWER(name) = LOWER(?)
          AND COALESCE(is_active, 1) = 1
        LIMIT 1
    """, (company_id, name))
    existing = cursor.fetchone()

    if existing:
        conn.close()
        flash("Destination with this name already exists.", "error")
        return redirect(url_for("product_transfers"))

    cursor.execute("""
        INSERT INTO stock_destinations (
            company_id,
            name,
            destination_code,
            notes,
            is_active
        )
        VALUES (?, ?, ?, ?, 1)
    """, (
        company_id,
        name,
        destination_code if destination_code else None,
        notes if notes else None,
    ))

    conn.commit()
    conn.close()

    flash("Destination created successfully.", "success")
    return redirect(url_for("product_transfers"))




@app.route("/transfers/<int:transfer_id>/undo", methods=["POST"])
@permission_required("manage_inventory")
def undo_product_transfer(transfer_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            t.id,
            t.product_id,
            t.destination_id,
            t.quantity,
            t.transfer_date,
            t.notes,
            p.product_name,
            p.measurement_unit,
            d.name AS destination_name
        FROM product_transfers_out t
        JOIN products p
          ON p.id = t.product_id
         AND p.company_id = t.company_id
        JOIN stock_destinations d
          ON d.id = t.destination_id
         AND d.company_id = t.company_id
        WHERE t.id = ?
          AND t.company_id = ?
        LIMIT 1
    """, (transfer_id, company_id))
    transfer = cursor.fetchone()

    if not transfer:
        conn.close()
        flash("Transfer not found.", "error")
        return redirect(url_for("product_transfers"))

    cursor.execute("""
        UPDATE products
        SET stock_quantity = COALESCE(stock_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (
        float(transfer["quantity"] or 0),
        transfer["product_id"],
        company_id,
    ))

    cursor.execute("""
        DELETE FROM product_transfers_out
        WHERE id = ?
          AND company_id = ?
    """, (transfer_id, company_id))

    conn.commit()
    conn.close()

    flash(
        f"Transfer undone. Restored {float(transfer['quantity'] or 0):g} "
        f"{transfer['measurement_unit'] or 'pcs'} of {transfer['product_name']} from {transfer['destination_name']}.",
        "success"
    )
    return redirect(url_for("product_transfers"))


@app.route("/transfers/new", methods=["POST"])
@permission_required("manage_inventory")
def create_product_transfer():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    product_id_raw = request.form.get("product_id", "").strip()
    destination_id_raw = request.form.get("destination_id", "").strip()
    quantity_raw = request.form.get("quantity", "").strip()
    transfer_date = request.form.get("transfer_date", "").strip()
    notes = request.form.get("notes", "").strip()

    if not product_id_raw or not destination_id_raw or not quantity_raw or not transfer_date:
        flash("Product, destination, quantity and date are required.", "error")
        return redirect(url_for("product_transfers"))

    try:
        product_id = int(product_id_raw)
        destination_id = int(destination_id_raw)
        quantity = float(quantity_raw)
    except ValueError:
        flash("Invalid transfer values.", "error")
        return redirect(url_for("product_transfers"))

    if quantity <= 0:
        flash("Transfer quantity must be greater than 0.", "error")
        return redirect(url_for("product_transfers"))

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            product_code,
            product_name,
            measurement_unit,
            COALESCE(stock_quantity, 0) AS stock_quantity
        FROM products
        WHERE id = ?
          AND company_id = ?
        LIMIT 1
    """, (product_id, company_id))
    product = cursor.fetchone()

    if not product:
        conn.close()
        flash("Product not found.", "error")
        return redirect(url_for("product_transfers"))

    cursor.execute("""
        SELECT
            id,
            name
        FROM stock_destinations
        WHERE id = ?
          AND company_id = ?
          AND COALESCE(is_active, 1) = 1
        LIMIT 1
    """, (destination_id, company_id))
    destination = cursor.fetchone()

    if not destination:
        conn.close()
        flash("Destination not found.", "error")
        return redirect(url_for("product_transfers"))

    current_stock = float(product["stock_quantity"] or 0)

    if quantity > current_stock:
        conn.close()
        flash(
            f"Not enough stock for {product['product_name']}. Available: {current_stock:g} {product['measurement_unit'] or 'pcs'}.",
            "error"
        )
        return redirect(url_for("product_transfers"))

    cursor.execute("""
        INSERT INTO product_transfers_out (
            company_id,
            product_id,
            destination_id,
            quantity,
            transfer_date,
            notes,
            created_by
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        company_id,
        product_id,
        destination_id,
        quantity,
        transfer_date,
        notes if notes else None,
        user_id,
    ))

    cursor.execute("""
        UPDATE products
        SET stock_quantity = COALESCE(stock_quantity, 0) - ?
        WHERE id = ?
          AND company_id = ?
    """, (quantity, product_id, company_id))

    conn.commit()
    conn.close()

    flash(
        f"Transferred {quantity:g} {product['measurement_unit'] or 'pcs'} of {product['product_name']} to {destination['name']}.",
        "success"
    )
    return redirect(url_for("product_transfers"))



@app.route("/inventory")
@permission_required("view_inventory")
def inventory():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            item_code,
            item_name,
            measurement_unit,
            COALESCE(unit_price, 0),
            COALESCE(stock_quantity, 0),
            COALESCE(min_stock, 0),
            COALESCE(reserved_quantity, 0),
            COALESCE(available_quantity, COALESCE(stock_quantity, 0) - COALESCE(reserved_quantity, 0))
        FROM items
        WHERE company_id = ?
        ORDER BY item_name ASC
    """, (company_id,))
    item_rows = cursor.fetchall()

    items_inventory = []
    total_items_value = 0

    for row in item_rows:
        stock_quantity = float(row[5] or 0)
        min_stock = float(row[6] or 0)
        reserved_quantity = float(row[7] or 0)
        available_quantity = float(row[8] or 0)
        unit_price = float(row[4] or 0)
        stock_value = stock_quantity * unit_price
        total_items_value += stock_value

        if stock_quantity <= 0:
            stock_status = "Out"
        elif available_quantity < min_stock:
            stock_status = "Low"
        else:
            stock_status = "OK"

        items_inventory.append({
            "id": row[0],
            "item_code": row[1],
            "item_name": row[2],
            "measurement_unit": row[3],
            "unit_price": unit_price,
            "stock_quantity": stock_quantity,
            "min_stock": min_stock,
            "reserved_quantity": reserved_quantity,
            "available_quantity": available_quantity,
            "stock_value": stock_value,
            "stock_status": stock_status
        })

    cursor.execute("""
        SELECT
            id,
            product_code,
            product_name,
            measurement_unit,
            COALESCE(stock_quantity, 0)
        FROM products
        WHERE company_id = ?
        ORDER BY product_name ASC
    """, (company_id,))
    product_rows = cursor.fetchall()

    products_inventory = []
    total_products_value = 0

    for row in product_rows:
        product_id = row[0]
        stock_quantity = float(row[4] or 0)
        material_cost_per_unit = calculate_product_material_cost(cursor, product_id, company_id=company_id)
        stock_value = stock_quantity * material_cost_per_unit
        total_products_value += stock_value

        products_inventory.append({
            "id": product_id,
            "product_code": row[1],
            "product_name": row[2],
            "measurement_unit": row[3],
            "stock_quantity": stock_quantity,
            "material_cost_per_unit": material_cost_per_unit,
            "stock_value": stock_value
        })

    conn.close()

    return render_template(
        "inventory.html",
        items_inventory=items_inventory,
        products_inventory=products_inventory,
        total_items_value=total_items_value,
        total_products_value=total_products_value,
        active_page="inventory"
    )


@app.route("/materials-shortage")
@permission_required("manage_procurement")
def materials_shortage():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            i.id,
            i.item_code,
            i.item_name,
            i.measurement_unit,
            COALESCE(i.stock_quantity, 0),
            COALESCE(i.min_stock, 0),
            i.supplier_id,
            s.name
        FROM items i
        LEFT JOIN suppliers s
          ON i.supplier_id = s.id
         AND s.company_id = i.company_id
        WHERE i.company_id = ?
          AND COALESCE(i.stock_quantity, 0) < COALESCE(i.min_stock, 0)
        ORDER BY i.item_name ASC
    """, (company_id,))
    rows = cursor.fetchall()

    conn.close()

    shortage_items = []
    for row in rows:
        stock_quantity = float(row[4] or 0)
        min_stock = float(row[5] or 0)

        shortage_items.append({
            "id": row[0],
            "item_code": row[1],
            "item_name": row[2],
            "unit": row[3],
            "stock_quantity": stock_quantity,
            "min_stock": min_stock,
            "supplier_id": row[6],
            "supplier_name": row[7],
            "required_to_order": max(0, min_stock - stock_quantity)
        })

    return render_template(
        "materials_shortage.html",
        shortage_items=shortage_items,
        active_page="shortage"
    )







def assign_planner_lanes(job_rows):
    lanes = []

    for job in sorted(job_rows, key=lambda item: (item["start_day"], item["end_day"], item["id"])):
        placed = False
        for lane_index, lane_end in enumerate(lanes):
            if job["start_day"] > lane_end:
                job["lane"] = lane_index
                lanes[lane_index] = job["end_day"]
                placed = True
                break

        if not placed:
            job["lane"] = len(lanes)
            lanes.append(job["end_day"])

    return max(1, len(lanes))

def assign_planner_lanes(jobs):
    if not jobs:
        return []

    sorted_jobs = sorted(
        jobs,
        key=lambda job: (
            job["start_day"],
            job["end_day"],
            job["id"]
        )
    )

    lane_end_days = []

    for job in sorted_jobs:
        assigned_lane = None

        for idx, lane_end_day in enumerate(lane_end_days):
            if job["start_day"] > lane_end_day:
                assigned_lane = idx
                lane_end_days[idx] = job["end_day"]
                break

        if assigned_lane is None:
            assigned_lane = len(lane_end_days)
            lane_end_days.append(job["end_day"])

        job["lane_index"] = assigned_lane

    lane_count = max(1, len(lane_end_days))

    for job in sorted_jobs:
        job["lane_count"] = lane_count

    return sorted_jobs

def build_jobs_filter_url(workstation_id, order_number, job_name):
    query = {
        "workstation": workstation_id,
        "order_number": order_number or "",
        "job_name": job_name or ""
    }
    return url_for("jobs", **query)


def parse_split_rows_from_request(form):
    workstation_ids = form.getlist("split_workstation_id")
    quantities = form.getlist("split_quantity")

    split_rows = []

    for workstation_id, quantity in zip(workstation_ids, quantities):
        workstation_id = (workstation_id or "").strip()
        quantity = (quantity or "").strip()

        if not workstation_id or not quantity:
            continue

        try:
            ws_id = int(workstation_id)
            qty_value = float(quantity)
        except ValueError:
            raise ValueError("Invalid split values.")

        if qty_value <= 0:
            continue

        split_rows.append({
            "workstation_id": ws_id,
            "quantity": qty_value
        })

    if len(split_rows) < 2:
        raise ValueError("Split requires at least 2 valid rows.")

    return split_rows


def assign_planner_lanes(jobs):
    """
    Kiekvienam workstation job list priskiria lane_index ir lane_count,
    kad persidengiantys darbai matytųsi vienas po kitu.
    """
    if not jobs:
        return []

    sorted_jobs = sorted(
        jobs,
        key=lambda j: (
            j["start_day"],
            j["end_day"],
            j["id"]
        )
    )

    lane_end_days = []

    for job in sorted_jobs:
        assigned_lane = None

        for idx, lane_end_day in enumerate(lane_end_days):
            if job["start_day"] > lane_end_day:
                assigned_lane = idx
                lane_end_days[idx] = job["end_day"]
                break

        if assigned_lane is None:
            assigned_lane = len(lane_end_days)
            lane_end_days.append(job["end_day"])

        job["lane_index"] = assigned_lane

    lane_count = max(1, len(lane_end_days))

    for job in sorted_jobs:
        job["lane_count"] = lane_count

    return sorted_jobs


@app.route("/planner")
def planner():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    today = datetime.today()

    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int) or today.month

    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    month_days = build_month_days(year, month)
    month_start = f"{year:04d}-{month:02d}-01"
    month_end = f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"

    prev_month = 12 if month == 1 else month - 1
    prev_year = year - 1 if month == 1 else year
    next_month = 1 if month == 12 else month + 1
    next_year = year + 1 if month == 12 else year

    conn = get_connection()
    cursor = conn.cursor()

    # svarbu: perstatom visas planner datas pagal dabartinę formulę,
    # kad seni jobai nebeliktų su senais per ilgais planned_end
    recalculate_company_planner_dates(cursor, company_id=company_id)
    conn.commit()

    cursor.execute("""
        SELECT
            id,
            name,
            color,
            hours_per_shift,
            shifts_per_day
        FROM workstations
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    workstation_rows = cursor.fetchall()

    workstations = []
    workstation_map = {}

    for row in workstation_rows:
        workstation = {
            "id": row[0],
            "name": row[1],
            "color": row[2] or "#3b82f6",
            "hours_per_shift": float(row[3] or 0),
            "shifts_per_day": float(row[4] or 0),
            "row_lane_count": 1
        }
        workstations.append(workstation)
        workstation_map[row[0]] = workstation

    jobs_by_workstation = {w["id"]: [] for w in workstations}
    unscheduled_jobs = []

    cursor.execute("""
        SELECT
            oj.id,
            oj.workstation_id,
            oj.job_name,
            oj.status,
            oj.planned_start,
            oj.planned_end,
            oj.completed_quantity,
            oj.planned_quantity,
            oj.estimated_hours,
            o.order_number,
            p.product_name,
            oj.parent_job_id,
            oj.is_split_child,
            oj.sequence
        FROM order_jobs oj
        JOIN orders o
          ON oj.order_id = o.id
         AND o.company_id = oj.company_id
        LEFT JOIN products p
          ON oj.job_product_id = p.id
         AND p.company_id = oj.company_id
        WHERE oj.company_id = ?
        ORDER BY o.order_number ASC, oj.sequence ASC, oj.id ASC
    """, (company_id,))
    rows = cursor.fetchall()
    conn.close()

    for row in rows:
        workstation_id = row[1]
        planned_start = row[4]
        planned_end = row[5]
        completed_quantity = float(row[6] or 0)
        planned_quantity = float(row[7] or 0)

        progress_percent = 0
        if planned_quantity > 0:
            progress_percent = min(100, max(0, (completed_quantity / planned_quantity) * 100))

        job = {
            "id": row[0],
            "workstation_id": workstation_id,
            "job_name": row[2],
            "status": row[3],
            "planned_start": planned_start,
            "planned_end": planned_end,
            "completed_quantity": completed_quantity,
            "planned_quantity": planned_quantity,
            "estimated_hours": float(row[8] or 0),
            "order_number": row[9],
            "product_name": row[10] if row[10] else "-",
            "parent_job_id": row[11],
            "is_split_child": int(row[12] or 0),
            "sequence": row[13],
            "progress_percent": progress_percent,
            "jobs_url": url_for(
                "jobs",
                workstation=workstation_id,
                order_number=row[9],
                job_name=row[2]
            )
        }

        if not planned_start or not planned_end or workstation_id not in jobs_by_workstation:
            unscheduled_jobs.append(job)
            continue

        if planned_end < month_start or planned_start > month_end:
            continue

        start_date = max(planned_start, month_start)
        end_date = min(planned_end, month_end)

        start_day = int(start_date[-2:])
        end_day = int(end_date[-2:])
        span_days = max(1, (end_day - start_day) + 1)

        ws_color = workstation_map[workstation_id]["color"] if workstation_id in workstation_map else "#3b82f6"

        jobs_by_workstation[workstation_id].append({
            "id": job["id"],
            "order_number": job["order_number"],
            "product_name": job["product_name"],
            "job_name": job["job_name"],
            "status": job["status"],
            "planned_start": job["planned_start"],
            "planned_end": job["planned_end"],
            "completed_quantity": job["completed_quantity"],
            "planned_quantity": job["planned_quantity"],
            "sequence": job["sequence"],
            "start_day": start_day,
            "end_day": end_day,
            "span_days": span_days,
            "color": ws_color,
            "progress_percent": job["progress_percent"],
            "parent_job_id": job["parent_job_id"],
            "is_split_child": job["is_split_child"],
            "jobs_url": job["jobs_url"]
        })

    for workstation in workstations:
        ws_id = workstation["id"]
        assigned_jobs = assign_planner_lanes(jobs_by_workstation.get(ws_id, []))
        jobs_by_workstation[ws_id] = assigned_jobs
        workstation["row_lane_count"] = max([job.get("lane_count", 1) for job in assigned_jobs] or [1])

    return render_template(
        "planner.html",
        workstations=workstations,
        month_days=month_days,
        jobs_by_workstation=jobs_by_workstation,
        unscheduled_jobs=unscheduled_jobs,
        planner_year=year,
        planner_month=month,
        planner_month_name=calendar.month_name[month],
        prev_month=prev_month,
        prev_year=prev_year,
        next_month=next_month,
        next_year=next_year,
        active_page="planner"
    )



@app.route("/planner/update-job-date/<int:job_id>", methods=["POST"])
def update_planner_job_date(job_id):
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    company_id = get_company_id()
    planned_start = request.form.get("planned_start", "").strip()
    workstation_id = request.form.get("workstation_id", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM order_jobs
        WHERE id = ?
          AND company_id = ?
    """, (job_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        return jsonify({"ok": False, "error": "Job not found"}), 404

    if workstation_id:
        try:
            workstation_id = int(workstation_id)
        except ValueError:
            conn.close()
            return jsonify({"ok": False, "error": "Invalid workstation"}), 400

        cursor.execute("""
            SELECT id
            FROM workstations
            WHERE id = ?
              AND company_id = ?
        """, (workstation_id, company_id))
        ws_row = cursor.fetchone()

        if ws_row is None:
            conn.close()
            return jsonify({"ok": False, "error": "Workstation not found"}), 404

        cursor.execute("""
            UPDATE order_jobs
            SET workstation_id = ?
            WHERE id = ?
              AND company_id = ?
        """, (workstation_id, job_id, company_id))

    recalculate_job_dates(cursor, job_id, planned_start or None)

    cursor.execute("""
        SELECT
            oj.id,
            oj.workstation_id,
            oj.planned_start,
            oj.planned_end
        FROM order_jobs oj
        WHERE oj.id = ?
          AND oj.company_id = ?
    """, (job_id, company_id))
    updated = cursor.fetchone()

    conn.commit()
    conn.close()

    if updated is None:
        return jsonify({"ok": False, "error": "Updated job not found"}), 404

    return jsonify({
        "ok": True,
        "job": {
            "id": updated[0],
            "workstation_id": updated[1],
            "planned_start": updated[2],
            "planned_end": updated[3]
        }
    })



@app.route("/jobs/split/<int:job_id>", methods=["POST"])
def split_job(job_id):
    if not is_logged_in():
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "Unauthorized"}), 401
        return redirect(url_for("login"))

    company_id = get_company_id()
    workstation_ids = request.form.getlist("split_workstation_id")
    quantities = request.form.getlist("split_quantity")

    split_rows = []

    for workstation_id, quantity in zip(workstation_ids, quantities):
        workstation_id = (workstation_id or "").strip()
        quantity = (quantity or "").strip()

        if not workstation_id or not quantity:
            continue

        try:
            ws_id = int(workstation_id)
            qty_value = float(quantity)
        except ValueError:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"ok": False, "error": "Invalid split values."}), 400
            flash("Invalid split values.", "error")
            return redirect_back("jobs")

        if qty_value <= 0:
            continue

        split_rows.append({
            "workstation_id": ws_id,
            "quantity": qty_value
        })

    if len(split_rows) < 2:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "Split requires at least 2 valid rows."}), 400
        flash("Split requires at least 2 valid rows.", "error")
        return redirect_back("jobs")

    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "order_jobs", job_id, company_id, "Job not found.")

        for row in split_rows:
            require_company_record(cursor, "workstations", row["workstation_id"], company_id, "Workstation not found.")

        create_split_children(cursor, job_id, split_rows, company_id=company_id)

        conn.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": True})

        flash("Job split successfully.", "success")
        return redirect_back("jobs")

    except ValueError as e:
        conn.rollback()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": str(e)}), 400

        flash(str(e), "error")
        return redirect_back("jobs")

    finally:
        conn.close()






@app.route("/planner/split-job/<int:job_id>", methods=["POST"])
def planner_split_job(job_id):
    if not is_logged_in():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    company_id = get_company_id()

    try:
        split_rows = parse_split_rows_from_request(request.form)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    conn = get_connection()
    cursor = conn.cursor()

    try:
        require_company_record(cursor, "order_jobs", job_id, company_id, "Job not found.")

        cursor.execute("""
            SELECT parent_job_id, is_split_child
            FROM order_jobs
            WHERE id = ? AND company_id = ?
        """, (job_id, company_id))
        row = cursor.fetchone()

        if not row:
            raise ValueError("Job not found.")

        if int(row[1] or 0) == 1 or row[0] is not None:
            raise ValueError("Only parent jobs can be split.")

        for split_row in split_rows:
            require_company_record(cursor, "workstations", split_row["workstation_id"], company_id, "Workstation not found.")

        create_split_children(cursor, job_id, split_rows, company_id=company_id)

        cursor.execute("SELECT order_id FROM order_jobs WHERE id = ? AND company_id = ?", (job_id, company_id))
        order_row = cursor.fetchone()
        if order_row:
            sync_order_status(cursor, order_row[0], company_id=company_id)

        conn.commit()
        return jsonify({"ok": True})
    except ValueError as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception:
        conn.rollback()
        return jsonify({"ok": False, "error": "Failed to split job."}), 500
    finally:
        conn.close()




@app.route("/workstations")
@permission_required("view_workstations")
def workstations():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            w.id,
            w.name,
            w.description,
            w.hours_per_shift,
            w.shifts_per_day,
            w.working_days_per_month,
            w.color,
            (w.hours_per_shift * w.shifts_per_day * w.working_days_per_month) AS monthly_capacity,
            COALESCE(w.cost_per_hour, 0),
            COALESCE((
                SELECT SUM(
                    oj.estimated_hours *
                    CASE
                        WHEN (oj.planned_quantity - oj.completed_quantity) < 0 THEN 0
                        ELSE (oj.planned_quantity - oj.completed_quantity)
                    END
                )
                FROM order_jobs oj
                WHERE oj.workstation_id = w.id
                  AND oj.company_id = w.company_id
                  AND oj.status != 'Done'
                  AND (
                        oj.status = 'Waiting'
                        OR oj.status = 'Ongoing'
                        OR oj.status = 'Paused'
                        OR oj.status = 'Delayed'
                  )
                  AND (
                        oj.is_split_child = 1
                        OR NOT EXISTS (
                            SELECT 1
                            FROM order_jobs child
                            WHERE child.parent_job_id = oj.id
                              AND child.is_split_child = 1
                              AND child.company_id = oj.company_id
                        )
                  )
            ), 0) AS used_load,
            COALESCE(wg.name, ''),
            COALESCE(wg.color, '')
        FROM workstations w
        LEFT JOIN workstation_groups wg
          ON wg.id = w.group_id
         AND wg.company_id = w.company_id
        WHERE w.company_id = ?
        ORDER BY COALESCE(wg.name, 'ZZZ'), w.name ASC
    """, (company_id,))
    rows = cursor.fetchall()

    workstation_groups = fetch_workstation_groups(cursor, company_id)
    conn.close()

    workstations = []
    for row in rows:
        monthly_capacity = float(row[7] or 0)
        used_load = float(row[9] or 0)
        free_load = monthly_capacity - used_load

        workstations.append({
            "id": row[0],
            "name": row[1],
            "description": row[2],
            "hours_per_shift": row[3],
            "shifts_per_day": row[4],
            "working_days_per_month": row[5],
            "color": row[6],
            "monthly_capacity": monthly_capacity,
            "cost_per_hour": float(row[8] or 0),
            "used_load": used_load,
            "free_load": free_load,
            "group_name": row[10] or "No group",
            "group_color": row[11] or "#6366f1",
        })

    return render_template(
        "workstations.html",
        workstations=workstations,
        workstation_groups=workstation_groups,
        active_page="workstations"
    )


@app.route("/workstation-groups/new", methods=["POST"])
@permission_required("view_workstations")
def create_workstation_group():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    name = request.form.get("name", "").strip()
    description = request.form.get("description", "").strip()
    color = request.form.get("color", "#6366f1").strip() or "#6366f1"

    if not name:
        flash("Workstation group name is required.", "error")
        return redirect(url_for("workstations"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM workstation_groups
        WHERE company_id = ?
          AND LOWER(name) = LOWER(?)
        LIMIT 1
    """, (company_id, name))
    existing = cursor.fetchone()

    if existing:
        conn.close()
        flash("Workstation group with this name already exists.", "error")
        return redirect(url_for("workstations"))

    cursor.execute("""
        INSERT INTO workstation_groups (
            company_id,
            name,
            description,
            color
        )
        VALUES (?, ?, ?, ?)
    """, (
        company_id,
        name,
        description or None,
        color
    ))

    conn.commit()
    conn.close()

    flash("Workstation group created successfully.", "success")
    return redirect(url_for("workstations"))


@app.route("/workstation-groups/delete/<int:group_id>", methods=["POST"])
def delete_workstation_group(group_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*)
        FROM workstations
        WHERE group_id = ?
          AND company_id = ?
    """, (group_id, company_id))
    linked_count = int(cursor.fetchone()[0] or 0)

    if linked_count > 0:
        conn.close()
        flash("Cannot delete this workstation group while workstations are still assigned to it.", "error")
        return redirect(url_for("workstations"))

    cursor.execute("""
        DELETE FROM workstation_groups
        WHERE id = ?
          AND company_id = ?
    """, (group_id, company_id))

    conn.commit()
    conn.close()

    flash("Workstation group deleted.", "success")
    return redirect(url_for("workstations"))


@app.route("/users")
@permission_required("manage_users")
def users():
    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, full_name, email, role, COALESCE(is_active, 1)
        FROM users
        WHERE company_id = ?
        ORDER BY id ASC
    """, (company_id,))
    rows = cursor.fetchall()
    conn.close()

    users = []
    for row in rows:
        user_id = row[0]
        role = row[3] or "worker"
        effective_permissions = get_effective_permissions(user_id=user_id, role=role)

        users.append({
            "id": user_id,
            "full_name": row[1],
            "email": row[2],
            "role": role,
            "is_active": int(row[4] or 0),
            "permissions": sorted(effective_permissions),
        })

    return render_template(
        "users.html",
        users=users,
        permission_keys=ALL_PERMISSION_KEYS,
        active_page="users"
    )

@app.route("/users/new", methods=["GET", "POST"])
@permission_required("manage_users")
def new_user():
    company_id = get_company_id()

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        role = request.form.get("role", "worker").strip().lower()

        # basic validation
        if not full_name or not email or not password:
            flash("All fields are required.", "error")
            return redirect(url_for("new_user"))

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return redirect(url_for("new_user"))

        if role not in ROLE_DEFAULT_PERMISSIONS:
            flash("Invalid role.", "error")
            return redirect(url_for("new_user"))

        conn = get_connection()
        cursor = conn.cursor()

        try:
            # 🔍 check duplicate email in same company
            cursor.execute("""
                SELECT id FROM users
                WHERE email = ? AND company_id = ?
            """, (email, company_id))

            if cursor.fetchone():
                flash("User with this email already exists.", "error")
                return redirect(url_for("new_user"))

            # 🔐 hash password
            hashed_password = generate_password_hash(password)

            # 🏢 get company name safely
            cursor.execute("SELECT name FROM companies WHERE id = ?", (company_id,))
            company_row = cursor.fetchone()
            company_name = company_row[0] if company_row else ""

            # 🔍 check if is_active column exists
            cursor.execute("PRAGMA table_info(users)")
            columns = [row[1] for row in cursor.fetchall()]

            if "is_active" in columns:
                cursor.execute("""
                    INSERT INTO users (full_name, company, email, password, company_id, role, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (full_name, company_name, email, hashed_password, company_id, role))
            else:
                # fallback jei DB sena
                cursor.execute("""
                    INSERT INTO users (full_name, company, email, password, company_id, role)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (full_name, company_name, email, hashed_password, company_id, role))

            conn.commit()

            flash("User created successfully.", "success")
            return redirect(url_for("account"))

        except sqlite3.IntegrityError as e:
            conn.rollback()
            flash(f"Integrity error: {str(e)}", "error")
            return redirect(url_for("new_user"))

        except sqlite3.OperationalError as e:
            conn.rollback()
            flash(f"Database error: {str(e)}", "error")
            return redirect(url_for("new_user"))

        except Exception as e:
            conn.rollback()
            flash(f"Unexpected error: {str(e)}", "error")
            return redirect(url_for("new_user"))

        finally:
            conn.close()

    return render_template("new_user.html")


@app.route("/users/<int:user_id>/role", methods=["POST"])
@permission_required("manage_users")
def update_user_role(user_id):
    company_id = get_company_id()
    current_user_id = session.get("user_id")
    new_role = request.form.get("role", "worker").strip().lower()

    if new_role not in ROLE_DEFAULT_PERMISSIONS:
        flash("Invalid role.", "error")
        return redirect(url_for("users"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM users
        WHERE id = ? AND company_id = ?
    """, (user_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("User not found.", "error")
        return redirect(url_for("users"))

    if user_id == current_user_id and new_role != "admin":
        conn.close()
        flash("You cannot remove admin role from your own account.", "error")
        return redirect(url_for("users"))

    cursor.execute("""
        UPDATE users
        SET role = ?
        WHERE id = ? AND company_id = ?
    """, (new_role, user_id, company_id))

    conn.commit()
    conn.close()

    flash("Role updated.", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:user_id>/permissions", methods=["POST"])
@permission_required("manage_users")
def update_user_permissions(user_id):
    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT role
        FROM users
        WHERE id = ? AND company_id = ?
    """, (user_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("User not found.", "error")
        return redirect(url_for("users"))

    role = (row[0] or "worker").lower()
    defaults = get_role_default_permissions(role)

    cursor.execute("DELETE FROM user_permissions WHERE user_id = ? AND (company_id = ? OR company_id IS NULL)", (user_id, company_id))

    for permission_key in ALL_PERMISSION_KEYS:
        selected_has = request.form.get(f"perm_{permission_key}") == "1"
        default_has = permission_key in defaults

        if selected_has != default_has:
            cursor.execute("""
                INSERT INTO user_permissions (user_id, company_id, permission_key, allowed)
                VALUES (?, ?, ?, ?)
            """, (user_id, company_id, permission_key, 1 if selected_has else 0))

    conn.commit()
    conn.close()

    flash("Permissions updated.", "success")
    return redirect(url_for("users"))
    

@app.route("/account")
@permission_required("view_dashboard")
def account():
    user_id = session.get("user_id")
    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, full_name, email, role, COALESCE(is_active, 1)
        FROM users
        WHERE id = ? AND company_id = ?
    """, (user_id, company_id))
    row = cursor.fetchone()
    conn.close()

    if row is None:
        flash("Account not found.", "error")
        return redirect(url_for("dashboard"))

    account_user = {
        "id": row[0],
        "full_name": row[1],
        "email": row[2],
        "role": row[3],
        "is_active": int(row[4] or 0),
        "permissions": sorted(get_effective_permissions(user_id=row[0], role=row[3])),
    }

    permission_groups = {
        "Orders": ["view_orders", "manage_orders"],
        "Jobs": ["view_jobs", "update_job_progress", "manage_jobs"],
        "Inventory": ["view_inventory", "manage_inventory"],
        "Products & Items": ["view_products", "manage_products", "view_items", "manage_items"],
        "Workstations": ["view_workstations", "manage_workstations"],
        "Reports & Export": ["view_reports", "export_data"],
        "Administration": ["manage_users"],
        "Procurement": ["manage_procurement"],
    }

    return render_template(
        "account.html",
        account_user=account_user,
        permission_groups=permission_groups,
        active_page="account"
    )   


@app.route("/account/change-password", methods=["POST"])
@permission_required("view_dashboard")
def change_password():
    user_id = session.get("user_id")
    company_id = get_company_id()

    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    if not current_password or not new_password or not confirm_password:
        flash("All password fields are required.", "error")
        return redirect(url_for("account"))

    if new_password != confirm_password:
        flash("New passwords do not match.", "error")
        return redirect(url_for("account"))

    if len(new_password) < 6:
        flash("New password must be at least 6 characters.", "error")
        return redirect(url_for("account"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT password
        FROM users
        WHERE id = ? AND company_id = ?
    """, (user_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Account not found.", "error")
        return redirect(url_for("dashboard"))

    current_password_hash = row[0]

    if not check_password_hash(current_password_hash, current_password):
        conn.close()
        flash("Current password is incorrect.", "error")
        return redirect(url_for("account"))

    new_password_hash = generate_password_hash(new_password)

    cursor.execute("""
        UPDATE users
        SET password = ?
        WHERE id = ? AND company_id = ?
    """, (new_password_hash, user_id, company_id))

    conn.commit()
    conn.close()

    flash("Password updated successfully.", "success")
    return redirect(url_for("account"))
# ---------------------------
# SUPPLIERS LIST
# ---------------------------
@app.route("/suppliers")
@permission_required("manage_procurement")
def suppliers():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip().lower()

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT
            id,
            company_id,
            name,
            supplier_code,
            contact_person,
            email,
            phone,
            address,
            notes,
            is_active,
            created_at
        FROM suppliers
        WHERE company_id = ?
    """
    params = [company_id]

    if search:
        sql += """
          AND (
                LOWER(COALESCE(name, '')) LIKE ?
             OR LOWER(COALESCE(supplier_code, '')) LIKE ?
             OR LOWER(COALESCE(contact_person, '')) LIKE ?
             OR LOWER(COALESCE(email, '')) LIKE ?
             OR LOWER(COALESCE(phone, '')) LIKE ?
          )
        """
        like_value = f"%{search.lower()}%"
        params.extend([like_value, like_value, like_value, like_value, like_value])

    if status_filter == "active":
        sql += " AND COALESCE(is_active, 0) = 1"
    elif status_filter == "inactive":
        sql += " AND COALESCE(is_active, 0) = 0"

    sql += " ORDER BY created_at DESC, id DESC"

    cursor.execute(sql, tuple(params))
    suppliers = cursor.fetchall()
    conn.close()

    return render_template(
        "suppliers.html",
        suppliers=suppliers,
        active_page="suppliers",
        search=search,
        status_filter=status_filter
    )


@app.route("/suppliers/new", methods=["GET", "POST"])
@permission_required("manage_suppliers")
def new_supplier():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        supplier_code = request.form.get("supplier_code", "").strip()
        contact_person = request.form.get("contact_person", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        notes = request.form.get("notes", "").strip()

        if not name:
            conn.close()
            flash("Supplier name is required.", "error")
            return redirect(url_for("new_supplier"))

        cursor.execute("""
            INSERT INTO suppliers (
                company_id,
                name,
                supplier_code,
                contact_person,
                email,
                phone,
                address,
                notes,
                is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
        """, (
            company_id,
            name,
            supplier_code or None,
            contact_person or None,
            email or None,
            phone or None,
            address or None,
            notes or None
        ))

        conn.commit()
        conn.close()

        flash("Supplier created successfully.", "success")
        return redirect(url_for("suppliers"))

    conn.close()
    return render_template(
        "new_supplier.html",
        active_page="suppliers"
    )

@app.route("/suppliers/<int:supplier_id>/edit", methods=["GET", "POST"])
@permission_required("manage_suppliers")
def edit_supplier(supplier_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            company_id,
            name,
            supplier_code,
            contact_person,
            email,
            phone,
            address,
            notes,
            is_active,
            created_at
        FROM suppliers
        WHERE id = ? AND company_id = ?
    """, (supplier_id, company_id))

    supplier = cursor.fetchone()

    if supplier is None:
        conn.close()
        flash("Supplier not found.", "error")
        return redirect(url_for("suppliers"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        supplier_code = request.form.get("supplier_code", "").strip()
        contact_person = request.form.get("contact_person", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        notes = request.form.get("notes", "").strip()
        is_active = 1 if request.form.get("is_active") == "on" else 0

        if not name:
            conn.close()
            flash("Supplier name is required.", "error")
            return redirect(url_for("edit_supplier", supplier_id=supplier_id))

        cursor.execute("""
            UPDATE suppliers
            SET
                name = ?,
                supplier_code = ?,
                contact_person = ?,
                email = ?,
                phone = ?,
                address = ?,
                notes = ?,
                is_active = ?
            WHERE id = ? AND company_id = ?
        """, (
            name,
            supplier_code or None,
            contact_person or None,
            email or None,
            phone or None,
            address or None,
            notes or None,
            is_active,
            supplier_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Supplier updated successfully.", "success")
        return redirect(url_for("suppliers"))

    conn.close()
    return render_template(
        "edit_supplier.html",
        supplier=supplier,
        active_page="suppliers"
    )


@app.route("/procurement/requests")
@permission_required("view_procurement")
def purchase_requests():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    search = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip().lower()
    priority_filter = request.args.get("priority", "").strip().lower()
    supplier_filter = request.args.get("supplier_id", "").strip()

    history_search = request.args.get("history_search", "").strip()
    history_status_filter = request.args.get("history_status", "").strip().lower()
    history_priority_filter = request.args.get("history_priority", "").strip().lower()
    history_supplier_filter = request.args.get("history_supplier_id", "").strip()

    show_history = request.args.get("show_history", "0") == "1"

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    active_sql = """
        SELECT
            pr.id,
            pr.company_id,
            pr.request_number,
            pr.item_id,
            pr.supplier_id,
            pr.title,
            pr.description,
            pr.quantity,
            pr.unit,
            pr.status,
            pr.priority,
            pr.needed_by,
            pr.requested_by,
            pr.approved_by,
            pr.ordered_by,
            pr.notes,
            pr.created_at,
            pr.updated_at,
            s.name AS supplier_name,
            u.full_name AS requester_name,
            i.item_name AS item_name,
            pr.order_id,
            pr.source_type,
            o.order_number
        FROM purchase_requests pr
        LEFT JOIN suppliers s
          ON pr.supplier_id = s.id
         AND s.company_id = pr.company_id
        LEFT JOIN users u
          ON pr.requested_by = u.id
         AND u.company_id = pr.company_id
        LEFT JOIN items i
          ON pr.item_id = i.id
         AND i.company_id = pr.company_id
        LEFT JOIN orders o
          ON pr.order_id = o.id
         AND o.company_id = pr.company_id
        WHERE pr.company_id = ?
          AND COALESCE(pr.status, 'draft') NOT IN ('received', 'cancelled', 'rejected')
    """
    active_params = [company_id]

    if search:
        like_value = f"%{search.lower()}%"
        active_sql += """
          AND (
                LOWER(COALESCE(pr.request_number, '')) LIKE ?
             OR LOWER(COALESCE(pr.title, '')) LIKE ?
             OR LOWER(COALESCE(pr.description, '')) LIKE ?
             OR LOWER(COALESCE(s.name, '')) LIKE ?
             OR LOWER(COALESCE(i.item_name, '')) LIKE ?
             OR LOWER(COALESCE(o.order_number, '')) LIKE ?
          )
        """
        active_params.extend([
            like_value,
            like_value,
            like_value,
            like_value,
            like_value,
            like_value,
        ])

    if status_filter in ("draft", "submitted", "approved", "ordered"):
        active_sql += " AND LOWER(COALESCE(pr.status, 'draft')) = ?"
        active_params.append(status_filter)

    if priority_filter in ("low", "normal", "high"):
        active_sql += " AND LOWER(COALESCE(pr.priority, 'normal')) = ?"
        active_params.append(priority_filter)

    if supplier_filter:
        try:
            active_sql += " AND pr.supplier_id = ?"
            active_params.append(int(supplier_filter))
        except ValueError:
            pass

    active_sql += " ORDER BY pr.created_at DESC, pr.id DESC"
    cursor.execute(active_sql, tuple(active_params))
    requests = [dict(row) for row in cursor.fetchall()]

    archive_sql = """
        SELECT
            pr.id,
            pr.company_id,
            pr.request_number,
            pr.item_id,
            pr.supplier_id,
            pr.title,
            pr.description,
            pr.quantity,
            pr.unit,
            pr.status,
            pr.priority,
            pr.needed_by,
            pr.requested_by,
            pr.approved_by,
            pr.ordered_by,
            pr.notes,
            pr.created_at,
            pr.updated_at,
            s.name AS supplier_name,
            u.full_name AS requester_name,
            i.item_name AS item_name,
            pr.order_id,
            pr.source_type,
            o.order_number
        FROM purchase_requests pr
        LEFT JOIN suppliers s
          ON pr.supplier_id = s.id
         AND s.company_id = pr.company_id
        LEFT JOIN users u
          ON pr.requested_by = u.id
         AND u.company_id = pr.company_id
        LEFT JOIN items i
          ON pr.item_id = i.id
         AND i.company_id = pr.company_id
        LEFT JOIN orders o
          ON pr.order_id = o.id
         AND o.company_id = pr.company_id
        WHERE pr.company_id = ?
          AND COALESCE(pr.status, 'draft') IN ('received', 'cancelled', 'rejected')
    """
    archive_params = [company_id]

    if history_search:
        history_like_value = f"%{history_search.lower()}%"
        archive_sql += """
          AND (
                LOWER(COALESCE(pr.request_number, '')) LIKE ?
             OR LOWER(COALESCE(pr.title, '')) LIKE ?
             OR LOWER(COALESCE(pr.description, '')) LIKE ?
             OR LOWER(COALESCE(s.name, '')) LIKE ?
             OR LOWER(COALESCE(i.item_name, '')) LIKE ?
             OR LOWER(COALESCE(o.order_number, '')) LIKE ?
          )
        """
        archive_params.extend([
            history_like_value,
            history_like_value,
            history_like_value,
            history_like_value,
            history_like_value,
            history_like_value,
        ])

    if history_status_filter in ("received", "cancelled", "rejected"):
        archive_sql += " AND LOWER(COALESCE(pr.status, 'draft')) = ?"
        archive_params.append(history_status_filter)

    if history_priority_filter in ("low", "normal", "high"):
        archive_sql += " AND LOWER(COALESCE(pr.priority, 'normal')) = ?"
        archive_params.append(history_priority_filter)

    if history_supplier_filter:
        try:
            archive_sql += " AND pr.supplier_id = ?"
            archive_params.append(int(history_supplier_filter))
        except ValueError:
            pass

    archive_sql += " ORDER BY COALESCE(pr.updated_at, pr.created_at) DESC, pr.id DESC"
    cursor.execute(archive_sql, tuple(archive_params))
    archived_requests = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
        SELECT id, name
        FROM suppliers
        WHERE company_id = ?
        ORDER BY name ASC
    """, (company_id,))
    supplier_options = [dict(row) for row in cursor.fetchall()]

    conn.close()

    return render_template(
        "purchase_requests.html",
        requests=requests,
        archived_requests=archived_requests,
        supplier_options=supplier_options,
        search=search,
        status_filter=status_filter,
        priority_filter=priority_filter,
        supplier_filter=supplier_filter,
        history_search=history_search,
        history_status_filter=history_status_filter,
        history_priority_filter=history_priority_filter,
        history_supplier_filter=history_supplier_filter,
        show_history=show_history,
        active_page="purchase_requests",
    )


@app.route("/procurement/requests/new", methods=["GET", "POST"])
@permission_required("manage_procurement")
def new_purchase_request():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        item_id_raw = request.form.get("item_id", "").strip()
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        quantity_raw = request.form.get("quantity", "").strip()
        unit = request.form.get("unit", "").strip()
        supplier_id_raw = request.form.get("supplier_id", "").strip()
        priority = request.form.get("priority", "normal").strip().lower()
        needed_by = request.form.get("needed_by", "").strip()
        notes = request.form.get("notes", "").strip()

        item_id = None
        selected_item_name = None
        selected_item_unit = None
        selected_item_supplier_id = None

        if item_id_raw:
            try:
                item_id = int(item_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid item selected.", "error")
                return redirect(url_for("new_purchase_request"))

            cursor.execute("""
                SELECT id, item_name, measurement_unit, supplier_id
                FROM items
                WHERE id = ? AND company_id = ?
            """, (item_id, company_id))
            item_row = cursor.fetchone()

            if item_row is None:
                conn.close()
                flash("Selected item was not found.", "error")
                return redirect(url_for("new_purchase_request"))

            selected_item_name = item_row[1]
            selected_item_unit = item_row[2]
            selected_item_supplier_id = item_row[3]

        if not title:
            title = selected_item_name or ""

        if not title:
            conn.close()
            flash("Request title is required.", "error")
            return redirect(url_for("new_purchase_request"))

        if not quantity_raw:
            conn.close()
            flash("Quantity is required.", "error")
            return redirect(url_for("new_purchase_request"))

        try:
            quantity = float(quantity_raw)
        except ValueError:
            conn.close()
            flash("Quantity must be a valid number.", "error")
            return redirect(url_for("new_purchase_request"))

        if quantity <= 0:
            conn.close()
            flash("Quantity must be greater than zero.", "error")
            return redirect(url_for("new_purchase_request"))

        if not unit and selected_item_unit:
            unit = selected_item_unit

        if not unit:
            conn.close()
            flash("Unit is required.", "error")
            return redirect(url_for("new_purchase_request"))

        if priority not in ("low", "normal", "high"):
            priority = "normal"

        if not supplier_id_raw and selected_item_supplier_id:
            supplier_id_raw = str(selected_item_supplier_id)

        supplier_id = None
        if supplier_id_raw:
            try:
                supplier_id = int(supplier_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid supplier selected.", "error")
                return redirect(url_for("new_purchase_request"))

            cursor.execute("""
                SELECT id
                FROM suppliers
                WHERE id = ? AND company_id = ?
            """, (supplier_id, company_id))
            supplier_row = cursor.fetchone()

            if supplier_row is None:
                conn.close()
                flash("Selected supplier was not found.", "error")
                return redirect(url_for("new_purchase_request"))

        cursor.execute("""
            INSERT INTO purchase_requests (
                company_id,
                request_number,
                item_id,
                supplier_id,
                title,
                description,
                quantity,
                unit,
                status,
                priority,
                needed_by,
                requested_by,
                approved_by,
                ordered_by,
                notes,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (
            company_id,
            None,
            item_id,
            supplier_id,
            title,
            description or None,
            quantity,
            unit,
            "draft",
            priority,
            needed_by or None,
            user_id,
            None,
            None,
            notes or None
        ))

        new_request_id = cursor.lastrowid
        request_number = f"PR-{new_request_id:05d}"

        cursor.execute("""
            UPDATE purchase_requests
            SET request_number = ?
            WHERE id = ? AND company_id = ?
        """, (request_number, new_request_id, company_id))

        conn.commit()
        conn.close()

        flash("Purchase request created successfully.", "success")
        return redirect(url_for("purchase_requests"))

    cursor.execute("""
        SELECT id, name
        FROM suppliers
        WHERE company_id = ? AND is_active = 1
        ORDER BY name ASC
    """, (company_id,))
    suppliers = cursor.fetchall()

    cursor.execute("""
        SELECT
            i.id,
            i.item_name,
            i.measurement_unit,
            i.supplier_id,
            s.name
        FROM items i
        LEFT JOIN suppliers s
          ON i.supplier_id = s.id
         AND s.company_id = i.company_id
        WHERE i.company_id = ?
        ORDER BY i.item_name ASC
    """, (company_id,))
    items = cursor.fetchall()

    conn.close()

    return render_template(
        "new_purchase_request.html",
        suppliers=suppliers,
        items=items,
        active_page="purchase_requests"
    )


@app.route("/procurement/requests/<int:request_id>/status/<status>")
@permission_required("manage_procurement")
def update_request_status(request_id, status):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")
    status = (status or "").strip().lower()

    allowed_statuses = {"submitted", "approved", "rejected", "ordered", "cancelled"}

    if status not in allowed_statuses:
        flash("Invalid status.", "error")
        return redirect(url_for("purchase_requests"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, status
        FROM purchase_requests
        WHERE id = ? AND company_id = ?
    """, (request_id, company_id))

    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Purchase request not found.", "error")
        return redirect(url_for("purchase_requests"))

    current_status = (row[1] or "").lower()

    allowed_transitions = {
        "draft": {"submitted", "cancelled"},
        "submitted": {"approved", "rejected", "cancelled"},
        "approved": {"ordered", "cancelled"},
        "ordered": set(),
        "rejected": set(),
        "cancelled": set()
    }

    if status not in allowed_transitions.get(current_status, set()):
        conn.close()
        flash(f"Cannot change status from {current_status} to {status}.", "error")
        return redirect(url_for("purchase_requests"))

    approved_by = None
    ordered_by = None

    if status == "approved":
        approved_by = user_id

    if status == "ordered":
        ordered_by = user_id

    cursor.execute("""
        UPDATE purchase_requests
        SET
            status = ?,
            approved_by = COALESCE(?, approved_by),
            ordered_by = COALESCE(?, ordered_by),
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND company_id = ?
    """, (
        status,
        approved_by,
        ordered_by,
        request_id,
        company_id
    ))

    conn.commit()
    conn.close()

    flash(f"Request marked as {status}.", "success")
    return redirect(url_for("purchase_requests"))


@app.route("/procurement/requests/<int:request_id>/edit", methods=["GET", "POST"])
@permission_required("manage_procurement")
def edit_purchase_request(request_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            company_id,
            request_number,
            item_id,
            supplier_id,
            title,
            description,
            quantity,
            unit,
            status,
            priority,
            needed_by,
            requested_by,
            approved_by,
            ordered_by,
            notes,
            created_at,
            updated_at
        FROM purchase_requests
        WHERE id = ? AND company_id = ?
    """, (request_id, company_id))

    purchase_request = cursor.fetchone()

    if purchase_request is None:
        conn.close()
        flash("Purchase request not found.", "error")
        return redirect(url_for("purchase_requests"))

    current_status = (purchase_request[9] or "").lower()

    if current_status not in ("draft", "submitted"):
        conn.close()
        flash("Only draft or submitted requests can be edited.", "error")
        return redirect(url_for("purchase_requests"))

    if request.method == "POST":
        item_id_raw = request.form.get("item_id", "").strip()
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        quantity_raw = request.form.get("quantity", "").strip()
        unit = request.form.get("unit", "").strip()
        supplier_id_raw = request.form.get("supplier_id", "").strip()
        priority = request.form.get("priority", "normal").strip().lower()
        needed_by = request.form.get("needed_by", "").strip()
        notes = request.form.get("notes", "").strip()

        item_id = None
        selected_item_name = None
        selected_item_unit = None
        selected_item_supplier_id = None

        if item_id_raw:
            try:
                item_id = int(item_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid item selected.", "error")
                return redirect(url_for("edit_purchase_request", request_id=request_id))

            cursor.execute("""
                SELECT id, item_name, measurement_unit, supplier_id
                FROM items
                WHERE id = ? AND company_id = ?
            """, (item_id, company_id))
            item_row = cursor.fetchone()

            if item_row is None:
                conn.close()
                flash("Selected item was not found.", "error")
                return redirect(url_for("edit_purchase_request", request_id=request_id))

            selected_item_name = item_row[1]
            selected_item_unit = item_row[2]
            selected_item_supplier_id = item_row[3]

        if not title:
            title = selected_item_name or ""

        if not title:
            conn.close()
            flash("Request title is required.", "error")
            return redirect(url_for("edit_purchase_request", request_id=request_id))

        if not quantity_raw:
            conn.close()
            flash("Quantity is required.", "error")
            return redirect(url_for("edit_purchase_request", request_id=request_id))

        try:
            quantity = float(quantity_raw)
        except ValueError:
            conn.close()
            flash("Quantity must be a valid number.", "error")
            return redirect(url_for("edit_purchase_request", request_id=request_id))

        if quantity <= 0:
            conn.close()
            flash("Quantity must be greater than zero.", "error")
            return redirect(url_for("edit_purchase_request", request_id=request_id))

        if not unit and selected_item_unit:
            unit = selected_item_unit

        if not unit:
            conn.close()
            flash("Unit is required.", "error")
            return redirect(url_for("edit_purchase_request", request_id=request_id))

        if priority not in ("low", "normal", "high"):
            priority = "normal"

        if not supplier_id_raw and selected_item_supplier_id:
            supplier_id_raw = str(selected_item_supplier_id)

        supplier_id = None
        if supplier_id_raw:
            try:
                supplier_id = int(supplier_id_raw)
            except ValueError:
                conn.close()
                flash("Invalid supplier selected.", "error")
                return redirect(url_for("edit_purchase_request", request_id=request_id))

            cursor.execute("""
                SELECT id
                FROM suppliers
                WHERE id = ? AND company_id = ?
            """, (supplier_id, company_id))

            supplier_row = cursor.fetchone()

            if supplier_row is None:
                conn.close()
                flash("Selected supplier was not found.", "error")
                return redirect(url_for("edit_purchase_request", request_id=request_id))

        cursor.execute("""
            UPDATE purchase_requests
            SET
                item_id = ?,
                supplier_id = ?,
                title = ?,
                description = ?,
                quantity = ?,
                unit = ?,
                priority = ?,
                needed_by = ?,
                notes = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ? AND company_id = ?
        """, (
            item_id,
            supplier_id,
            title,
            description or None,
            quantity,
            unit,
            priority,
            needed_by or None,
            notes or None,
            request_id,
            company_id
        ))

        conn.commit()
        conn.close()

        flash("Purchase request updated successfully.", "success")
        return redirect(url_for("purchase_requests"))

    cursor.execute("""
        SELECT id, name
        FROM suppliers
        WHERE company_id = ? AND is_active = 1
        ORDER BY name ASC
    """, (company_id,))
    suppliers = cursor.fetchall()

    cursor.execute("""
        SELECT
            i.id,
            i.item_name,
            i.measurement_unit,
            i.supplier_id,
            s.name
        FROM items i
        LEFT JOIN suppliers s
          ON i.supplier_id = s.id
         AND s.company_id = i.company_id
        WHERE i.company_id = ?
        ORDER BY i.item_name ASC
    """, (company_id,))
    items = cursor.fetchall()

    conn.close()

    return render_template(
        "edit_purchase_request.html",
        purchase_request=purchase_request,
        suppliers=suppliers,
        items=items,
        active_page="purchase_requests"
    )


@app.route("/procurement/requests/<int:request_id>/receive", methods=["GET", "POST"])
@permission_required("manage_procurement")
def receive_purchase_request(request_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            item_id,
            quantity,
            unit,
            status,
            order_id,
            source_type
        FROM purchase_requests
        WHERE id = ?
          AND company_id = ?
    """, (request_id, company_id))
    pr = cursor.fetchone()

    if pr is None:
        conn.close()
        flash("Purchase request not found.", "error")
        return redirect(url_for("purchase_requests"))

    pr_id, item_id, quantity, unit, status, order_id, source_type = pr
    receive_qty = float(quantity or 0)

    if not item_id:
        conn.close()
        flash("This request has no linked item.", "error")
        return redirect(url_for("purchase_requests"))

    cursor.execute("""
        UPDATE items
        SET stock_quantity = COALESCE(stock_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (receive_qty, item_id, company_id))

    cursor.execute("""
        UPDATE purchase_requests
        SET status = 'received',
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
          AND company_id = ?
    """, (request_id, company_id))

    recalculate_shortages_for_item(cursor, item_id, company_id=company_id)
    rebuild_company_reserved_quantities(cursor, company_id=company_id)

    conn.commit()
    conn.close()

    flash("Purchase request received and shortages updated.", "success")
    return redirect(url_for("purchase_requests"))



@app.route("/materials-shortage/<int:item_id>/create-request", methods=["POST"])
@permission_required("manage_procurement")
def create_request_from_shortage(item_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            item_name,
            measurement_unit,
            supplier_id,
            COALESCE(stock_quantity, 0),
            COALESCE(min_stock, 0)
        FROM items
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Item not found.", "error")
        return redirect(url_for("materials_shortage"))

    stock_quantity = float(row[4] or 0)
    min_stock = float(row[5] or 0)
    required_to_order = max(0, min_stock - stock_quantity)

    if required_to_order <= 0:
        conn.close()
        flash("This item is no longer below minimum stock.", "info")
        return redirect(url_for("materials_shortage"))

    cursor.execute("""
        INSERT INTO purchase_requests (
            company_id,
            request_number,
            item_id,
            supplier_id,
            title,
            description,
            quantity,
            unit,
            status,
            priority,
            needed_by,
            requested_by,
            approved_by,
            ordered_by,
            notes,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
    """, (
        company_id,
        None,
        row[0],
        row[3],
        row[1],
        f"Auto-created from shortage. Current stock: {stock_quantity:g}, minimum stock: {min_stock:g}.",
        required_to_order,
        row[2] or "",
        "draft",
        "high",
        None,
        user_id,
        None,
        None,
        "Created from shortage screen."
    ))

    new_request_id = cursor.lastrowid
    request_number = f"PR-{new_request_id:05d}"

    cursor.execute("""
        UPDATE purchase_requests
        SET request_number = ?
        WHERE id = ?
          AND company_id = ?
    """, (request_number, new_request_id, company_id))

    conn.commit()
    conn.close()

    flash("Purchase request created from shortage.", "success")
    return redirect(url_for("purchase_requests"))





@app.route("/dashboard/save-layout", methods=["POST"])
def save_dashboard_layout():
    if not is_logged_in():
        return {"ok": False, "error": "Unauthorized"}, 401

    company_id = get_company_id()
    user_id = session.get("user_id")

    data = request.get_json(silent=True) or {}
    layout = data.get("layout", [])

    if not isinstance(layout, list):
        return {"ok": False, "error": "Invalid payload"}, 400

    clean_layout = []

    for item in layout:
        if not isinstance(item, dict):
            continue

        widget_id = str(item.get("id", "")).strip()
        if not widget_id:
            continue

        try:
            x = float(item.get("x", 0))
            y = float(item.get("y", 0))
            w = float(item.get("w", 320))
            h = float(item.get("h", 180))
        except (TypeError, ValueError):
            continue

        clean_layout.append({
            "id": widget_id,
            "x": max(0, x),
            "y": max(0, y),
            "w": max(240, w),
            "h": max(100, h)
        })

    conn = get_connection()
    cursor = conn.cursor()

    save_dashboard_layout_record(cursor, user_id, company_id, clean_layout, "dashboard")

    conn.commit()
    conn.close()

    return {"ok": True}




@app.route("/reports")
@permission_required("view_reports")
def reports():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()

    report_type = request.args.get("report_type", "").strip()
    job_search = request.args.get("job_search", "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            pr.id,
            pr.report_type,
            pr.quantity,
            COALESCE(pr.unit, 'pcs') AS unit,
            pr.notes,
            pr.created_at,
            oj.id,
            oj.job_name,
            oj.status,
            o.order_number,
            COALESCE(p.product_name, '-') AS product_name,
            COALESCE(w.name, '-') AS workstation_name,
            COALESCE(u.full_name, 'System') AS reported_by_name
        FROM production_reports pr
        JOIN order_jobs oj
          ON pr.job_id = oj.id
         AND oj.company_id = pr.company_id
        LEFT JOIN orders o
          ON pr.order_id = o.id
         AND o.company_id = pr.company_id
        LEFT JOIN products p
          ON pr.product_id = p.id
         AND p.company_id = pr.company_id
        LEFT JOIN workstations w
          ON pr.workstation_id = w.id
         AND w.company_id = pr.company_id
        LEFT JOIN users u
          ON pr.reported_by = u.id
        WHERE pr.company_id = ?
    """
    params = [company_id]

    if report_type:
        query += " AND pr.report_type = ?"
        params.append(report_type)

    if job_search:
        query += """
          AND (
                o.order_number LIKE ?
                OR oj.job_name LIKE ?
                OR p.product_name LIKE ?
                OR w.name LIKE ?
          )
        """
        like_value = f"%{job_search}%"
        params.extend([like_value, like_value, like_value, like_value])

    query += " ORDER BY pr.created_at DESC, pr.id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    reports = []
    total_scrap = 0
    total_waste = 0
    total_defect = 0

    for row in rows:
        report = {
            "id": row[0],
            "report_type": row[1],
            "quantity": float(row[2] or 0),
            "unit": row[3] or "pcs",
            "notes": row[4],
            "created_at": row[5],
            "job_id": row[6],
            "job_name": row[7],
            "job_status": row[8],
            "order_number": row[9],
            "product_name": row[10],
            "workstation_name": row[11],
            "reported_by_name": row[12]
        }
        reports.append(report)

        if report["report_type"] == "scrap":
            total_scrap += report["quantity"]
        elif report["report_type"] == "waste":
            total_waste += report["quantity"]
        elif report["report_type"] == "defect":
            total_defect += report["quantity"]

    conn.close()

    return render_template(
        "reports.html",
        reports=reports,
        total_scrap=total_scrap,
        total_waste=total_waste,
        total_defect=total_defect,
        filters={
            "report_type": report_type,
            "job_search": job_search
        },
        active_page="reports"
    )


@app.route("/reports/new", methods=["GET", "POST"])
@permission_required("view_reports")
def new_report():
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    cursor = conn.cursor()

    if request.method == "POST":
        job_id_raw = request.form.get("job_id", "").strip()
        report_type = request.form.get("report_type", "").strip().lower()
        quantity_raw = request.form.get("quantity", "").strip()
        unit_raw = request.form.get("unit", "").strip()
        custom_unit_raw = request.form.get("custom_unit", "").strip()
        notes = request.form.get("notes", "").strip()

        if not job_id_raw or not report_type or not quantity_raw:
            conn.close()
            flash("Job, report type and quantity are required.", "error")
            return redirect(url_for("new_report"))

        if report_type not in {"scrap", "waste", "defect", "note"}:
            conn.close()
            flash("Invalid report type.", "error")
            return redirect(url_for("new_report"))

        try:
            job_id = int(job_id_raw)
            quantity = float(quantity_raw)
        except ValueError:
            conn.close()
            flash("Invalid job or quantity.", "error")
            return redirect(url_for("new_report"))

        if quantity < 0:
            conn.close()
            flash("Quantity cannot be negative.", "error")
            return redirect(url_for("new_report"))

        unit = unit_raw or "pcs"
        if unit == "custom":
            unit = custom_unit_raw.strip()

        if not unit:
            unit = "pcs"

        cursor.execute("""
            SELECT
                oj.id,
                oj.order_id,
                oj.job_product_id,
                oj.workstation_id,
                oj.status
            FROM order_jobs oj
            WHERE oj.id = ?
              AND oj.company_id = ?
        """, (job_id, company_id))
        job_row = cursor.fetchone()

        if job_row is None:
            conn.close()
            flash("Selected job not found.", "error")
            return redirect(url_for("new_report"))

        if job_row[4] not in ("Waiting", "Ongoing", "Paused", "Delayed"):
            conn.close()
            flash("Reports can only be created for active jobs.", "error")
            return redirect(url_for("new_report"))

        cursor.execute("""
            INSERT INTO production_reports (
                company_id,
                job_id,
                order_id,
                product_id,
                workstation_id,
                report_type,
                quantity,
                unit,
                notes,
                reported_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            company_id,
            job_row[0],
            job_row[1],
            job_row[2],
            job_row[3],
            report_type,
            quantity,
            unit,
            notes,
            user_id
        ))

        conn.commit()
        conn.close()

        flash("Production report created successfully.", "success")
        return redirect(url_for("reports"))

    active_jobs = get_active_jobs_for_reports(cursor, company_id)
    conn.close()

    return render_template(
        "new_report.html",
        active_jobs=active_jobs,
        active_page="reports"
    )


@app.route("/reports/new/<int:job_id>")
@permission_required("view_reports")
def new_report_for_job(job_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    return redirect(url_for("new_report", job_id=job_id))





@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()

        if not email:
            flash("Email is required.", "error")
            return render_template("forgot_password.html")

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, email
            FROM users
            WHERE LOWER(email) = ?
            LIMIT 1
        """, (email,))
        user_row = cursor.fetchone()

        # Svarbu: neatskleidžiam ar email egzistuoja
        if user_row:
            user_id = user_row[0]
            user_email = user_row[1]

            raw_token = secrets.token_urlsafe(32)
            token_hash = hash_reset_token(raw_token)
            expires_at = (datetime.utcnow() + timedelta(minutes=45)).isoformat()

            cursor.execute("""
                INSERT INTO password_reset_tokens (user_id, token_hash, expires_at)
                VALUES (?, ?, ?)
            """, (user_id, token_hash, expires_at))

            conn.commit()

            reset_link = build_reset_link(raw_token)

            try:
                send_password_reset_email(user_email, reset_link)
            except Exception as e:
                conn.rollback()
                conn.close()
                print(f"[SMTP ERROR] {e}")
                flash("Failed to send reset email. Please try again later.", "error")
                return redirect(url_for("forgot_password"))

        conn.close()

        flash("If that email exists, a password reset link has been sent.", "success")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")
@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    token_hash = hash_reset_token(token)

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT prt.id, prt.user_id, prt.expires_at, prt.used_at, u.email
        FROM password_reset_tokens prt
        JOIN users u ON prt.user_id = u.id
        WHERE prt.token_hash = ?
        LIMIT 1
    """, (token_hash,))
    row = cursor.fetchone()

    if row is None:
        conn.close()
        flash("Invalid or expired reset link.", "error")
        return redirect(url_for("login"))

    token_id, user_id, expires_at, used_at, user_email = row

    if used_at:
        conn.close()
        flash("This reset link has already been used.", "error")
        return redirect(url_for("login"))

    try:
        expires_dt = datetime.fromisoformat(expires_at)
    except ValueError:
        conn.close()
        flash("Invalid or expired reset link.", "error")
        return redirect(url_for("login"))

    if datetime.utcnow() > expires_dt:
        conn.close()
        flash("This reset link has expired.", "error")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not password or not confirm_password:
            conn.close()
            flash("Both password fields are required.", "error")
            return render_template("reset_password.html", token=token, email=user_email)

        if password != confirm_password:
            conn.close()
            flash("Passwords do not match.", "error")
            return render_template("reset_password.html", token=token, email=user_email)

        if len(password) < 6:
            conn.close()
            flash("Password must be at least 6 characters.", "error")
            return render_template("reset_password.html", token=token, email=user_email)

        new_password_hash = generate_password_hash(password)

        cursor.execute("""
            UPDATE users
            SET password = ?
            WHERE id = ?
        """, (new_password_hash, user_id))

        cursor.execute("""
            UPDATE password_reset_tokens
            SET used_at = ?
            WHERE id = ?
        """, (datetime.utcnow().isoformat(), token_id))

        conn.commit()
        conn.close()

        flash("Password reset successfully. You can now log in.", "success")
        return redirect(url_for("login"))

    conn.close()
    return render_template("reset_password.html", token=token, email=user_email)



@app.route("/admin/backup/download")
@permission_required("manage_users")
def download_backup():
    backup_path = create_database_backup("manual")
    download_name = os.path.basename(backup_path)
    mimetype = "application/sql" if backup_path.endswith(".sql") else "application/x-sqlite3"
    return send_file(
        backup_path,
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype
    )


@app.route("/exports/orders.xlsx")
@permission_required("export_data")
def export_orders_excel():
    rows = fetch_orders_export_rows(get_company_id(), request.args)
    return excel_response(
        "orders_export.xlsx",
        [("Orders", ["Order", "Product", "Quantity", "Status", "Priority", "Due Date"], rows)]
    )


@app.route("/exports/orders.pdf")
@permission_required("export_data")
def export_orders_pdf():
    rows = fetch_orders_export_rows(get_company_id(), request.args)
    return pdf_response(
        "orders_export.pdf",
        "Orders Export",
        ["Order", "Product", "Qty", "Status", "Priority", "Due Date"],
        rows
    )


@app.route("/exports/jobs.xlsx")
@permission_required("export_data")
def export_jobs_excel():
    rows = fetch_jobs_export_rows(get_company_id(), request.args)
    return excel_response(
        "jobs_export.xlsx",
        [("Jobs", ["Order", "Product", "Job", "Workstation", "Sequence", "Planned Qty", "Done Qty", "Est. Hours", "Status", "Due Date", "Planned Start", "Planned End"], rows)]
    )


@app.route("/exports/jobs.pdf")
@permission_required("export_data")
def export_jobs_pdf():
    rows = fetch_jobs_export_rows(get_company_id(), request.args)
    return pdf_response(
        "jobs_export.pdf",
        "Jobs Export",
        ["Order", "Product", "Job", "Workstation", "Seq", "Planned", "Done", "Hours", "Status", "Due", "Start", "End"],
        rows
    )



@app.route("/exports/suppliers.xlsx")
@permission_required("export_data")
def export_suppliers_excel():
    rows = fetch_suppliers_export_rows(get_company_id(), request.args)
    return excel_response(
        "suppliers_export.xlsx",
        [("Suppliers", ["Supplier", "Code", "Contact", "Email", "Phone", "Address", "Status", "Created At"], rows)]
    )


@app.route("/exports/suppliers.pdf")
@permission_required("export_data")
def export_suppliers_pdf():
    rows = fetch_suppliers_export_rows(get_company_id(), request.args)
    return pdf_response(
        "suppliers_export.pdf",
        "Suppliers Export",
        ["Supplier", "Code", "Contact", "Email", "Phone", "Address", "Status", "Created"],
        rows
    )


@app.route("/exports/shortages.xlsx")
@permission_required("export_data")
def export_shortages_excel():
    rows = fetch_shortage_export_rows(get_company_id())
    return excel_response(
        "shortages_export.xlsx",
        [("Shortages", ["Item Code", "Item Name", "Unit", "In Stock", "Minimum", "Required To Order", "Supplier"], rows)]
    )


@app.route("/exports/shortages.pdf")
@permission_required("export_data")
def export_shortages_pdf():
    rows = fetch_shortage_export_rows(get_company_id())
    return pdf_response(
        "shortages_export.pdf",
        "Shortages Export",
        ["Item Code", "Item Name", "Unit", "In Stock", "Minimum", "Need", "Supplier"],
        rows
    )


@app.route("/exports/reports.xlsx")
@permission_required("export_data")
def export_reports_excel():
    rows = fetch_reports_export_rows(get_company_id(), request.args)
    return excel_response(
        "reports_export.xlsx",
        [("Reports", ["Type", "Quantity", "Unit", "Notes", "Created", "Job", "Job Status", "Order", "Product", "Workstation", "Reported By"], rows)]
    )


@app.route("/exports/reports.pdf")
@permission_required("export_data")
def export_reports_pdf():
    rows = fetch_reports_export_rows(get_company_id(), request.args)
    return pdf_response(
        "reports_export.pdf",
        "Reports Export",
        ["Type", "Qty", "Unit", "Notes", "Created", "Job", "Job Status", "Order", "Product", "Workstation", "User"],
        rows
    )



@app.route("/exports/purchase-requests.xlsx")
@permission_required("export_data")
def export_purchase_requests_excel():
    company_id = get_company_id()
    active_rows = fetch_purchase_request_export_rows(company_id, request.args, history=False)
    history_rows = fetch_purchase_request_export_rows(company_id, request.args, history=True)

    return excel_response(
        "purchase_requests_export.xlsx",
        [
            ("Active Requests", ["Request #", "Title", "Item", "Supplier", "Qty", "Unit", "Priority", "Status", "Needed By", "Requester", "Created", "Updated", "Notes"], active_rows),
            ("History", ["Request #", "Title", "Item", "Supplier", "Qty", "Unit", "Priority", "Status", "Needed By", "Requester", "Created", "Updated", "Notes"], history_rows),
        ]
    )


@app.route("/exports/purchase-requests.pdf")
@permission_required("export_data")
def export_purchase_requests_pdf():
    show_history = request.args.get("show_history", "") == "1"
    rows = fetch_purchase_request_export_rows(get_company_id(), request.args, history=show_history)
    title = "Purchase Request History Export" if show_history else "Purchase Requests Export"

    return pdf_response(
        "purchase_requests_export.pdf",
        title,
        ["Request #", "Title", "Item", "Supplier", "Qty", "Unit", "Priority", "Status", "Needed By", "Requester", "Created", "Updated", "Notes"],
        rows
    )

@app.route("/exports/inventory.xlsx")
@permission_required("export_data")
def export_inventory_excel():
    rows = fetch_inventory_export_rows(get_company_id())
    return excel_response(
        "inventory_export.xlsx",
        [("Inventory", ["Type", "Code", "Name", "Unit", "Stock Qty", "Unit Price", "Stock Value"], rows)]
    )


@app.route("/exports/inventory.pdf")
@permission_required("export_data")
def export_inventory_pdf():
    rows = fetch_inventory_export_rows(get_company_id())
    return pdf_response(
        "inventory_export.pdf",
        "Inventory Export",
        ["Type", "Code", "Name", "Unit", "Stock Qty", "Unit Price", "Stock Value"],
        rows
    )



@app.route("/exports/products.xlsx")
@permission_required("export_data")
def export_products_excel():
    rows = fetch_products_export_rows(get_company_id())
    return excel_response(
        "products_export.xlsx",
        [("Products", ["Code", "Name", "Description", "Unit", "Time / Unit", "Stock Qty"], rows)]
    )


@app.route("/exports/items.xlsx")
@permission_required("export_data")
def export_items_excel():
    rows = fetch_items_export_rows(get_company_id())
    return excel_response(
        "items_export.xlsx",
        [("Items", ["Code", "Name", "Description", "Unit", "Unit Price", "Stock Qty", "Min Stock"], rows)]
    )
@app.route("/exports/planner.xlsx")
@permission_required("export_data")
def export_planner_excel():
    today = datetime.today()
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int) or today.month

    scheduled, unscheduled = fetch_planner_export_rows(get_company_id(), year, month)

    return excel_response(
        "planner_export.xlsx",
        [
            ("Scheduled", ["Order", "Product", "Job", "Workstation", "Status", "Planned Start", "Planned End", "Planned Qty", "Done Qty", "Est. Hours"], scheduled),
            ("Unscheduled", ["Order", "Product", "Job", "Workstation", "Status", "Planned Start", "Planned End", "Planned Qty", "Done Qty", "Est. Hours"], unscheduled),
        ]
    )


@app.route("/exports/planner.pdf")
@permission_required("export_data")
def export_planner_pdf():
    today = datetime.today()
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int) or today.month

    scheduled, _ = fetch_planner_export_rows(get_company_id(), year, month)

    return pdf_response(
        "planner_export.pdf",
        f"Planner Export - {calendar.month_name[month]} {year}",
        ["Order", "Product", "Job", "Workstation", "Status", "Planned Start", "Planned End", "Planned Qty", "Done Qty", "Est. Hours"],
        scheduled
    )
@app.route("/exports/order-materials/<int:order_id>.xlsx")
@permission_required("export_data")
def export_order_materials_excel(order_id):
    order, rows = fetch_order_material_rows(get_company_id(), order_id)

    if not order:
        flash("Order not found.", "error")
        return redirect(url_for("orders"))

    return excel_response(
        f"{order['order_number']}_materials.xlsx",
        [("Materials", ["Item Code", "Item Name", "Unit", "BOM / 1 Unit", "Total Needed", "Unit Price", "Total Cost"], rows)]
    )


@app.route("/exports/order-materials/<int:order_id>.pdf")
@permission_required("export_data")
def export_order_materials_pdf(order_id):
    order, rows = fetch_order_material_rows(get_company_id(), order_id)

    if not order:
        flash("Order not found.", "error")
        return redirect(url_for("orders"))

    subtitle = f"Order {order['order_number']} · {order['product_name']} · Qty {order['quantity']}"

    return pdf_response(
        f"{order['order_number']}_materials.pdf",
        "Order Materials Export",
        ["Item Code", "Item Name", "Unit", "BOM / 1 Unit", "Total Needed", "Unit Price", "Total Cost"],
        rows,
        subtitle
    )


@app.route("/orders/<int:order_id>/shortages")
@permission_required("view_orders")
def order_shortages(order_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT order_number
        FROM orders
        WHERE id = ?
          AND company_id = ?
    """, (order_id, company_id))
    order_row = cursor.fetchone()

    if order_row is None:
        conn.close()
        flash("Order not found.", "error")
        return redirect(url_for("orders"))

    cursor.execute("""
        SELECT
            s.id,
            i.item_code,
            i.item_name,
            i.measurement_unit,
            COALESCE(s.required_qty, 0),
            COALESCE(s.reserved_qty, 0),
            COALESCE(s.missing_qty, 0),
            COALESCE(s.covered, 0)
        FROM shortages s
        JOIN items i
          ON s.item_id = i.id
         AND i.company_id = s.company_id
        WHERE s.order_id = ?
          AND s.company_id = ?
          AND COALESCE(s.missing_qty, 0) > 0
          AND COALESCE(s.covered, 0) = 0
        ORDER BY i.item_name ASC
    """, (order_id, company_id))
    rows = cursor.fetchall()
    conn.close()

    shortages = []
    for row in rows:
        shortages.append({
            "id": row[0],
            "item_code": row[1],
            "item_name": row[2],
            "unit": row[3],
            "required_qty": float(row[4] or 0),
            "reserved_qty": float(row[5] or 0),
            "missing_qty": float(row[6] or 0),
            "covered": int(row[7] or 0)
        })

    return render_template(
        "order_shortages.html",
        order_id=order_id,
        order_number=order_row[0],
        shortages=shortages,
        active_page="orders"
    )


@app.route("/shortages/<int:shortage_id>/create-pr", methods=["POST"])
def create_purchase_request_from_shortage(shortage_id):
    if not is_logged_in():
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "Please log in again."}), 401
        return redirect(url_for("login"))

    if not (has_permission("manage_orders") or has_permission("manage_procurement")):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "You do not have permission to create purchase requests."}), 403
        flash("You do not have permission to create purchase requests.", "error")
        return redirect(url_for("dashboard"))

    company_id = get_company_id()
    user_id = session.get("user_id")

    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                s.id,
                s.order_id,
                s.item_id,
                COALESCE(s.required_qty, 0) AS required_qty,
                COALESCE(s.missing_qty, 0) AS missing_qty,
                COALESCE(s.covered, 0) AS covered,
                i.item_code,
                i.item_name,
                COALESCE(i.measurement_unit, 'pcs') AS measurement_unit,
                i.supplier_id,
                o.order_number,
                o.product_id
            FROM shortages s
            JOIN items i
              ON i.id = s.item_id
             AND i.company_id = s.company_id
            JOIN orders o
              ON o.id = s.order_id
             AND o.company_id = s.company_id
            WHERE s.id = ?
              AND s.company_id = ?
            LIMIT 1
        """, (shortage_id, company_id))
        shortage = cursor.fetchone()

        if not shortage:
            return jsonify({"ok": False, "error": "Shortage not found."}), 404

        if float(shortage["missing_qty"] or 0) <= 0:
            return jsonify({"ok": False, "error": "This shortage no longer has missing quantity."}), 400

        if int(shortage["covered"] or 0) == 1:
            return jsonify({"ok": False, "error": "This shortage is already covered."}), 409

        cursor.execute("PRAGMA table_info(purchase_requests)")
        pr_columns_info = cursor.fetchall()
        pr_columns = {row["name"] for row in pr_columns_info}

        if not pr_columns:
            return jsonify({"ok": False, "error": "purchase_requests table is missing."}), 500

        active_statuses = ("draft", "ordered", "waiting_for_receiving", "pending", "open")

        existing_query = """
            SELECT id
            FROM purchase_requests
            WHERE company_id = ?
              AND item_id = ?
              AND order_id = ?
        """
        existing_params = [company_id, shortage["item_id"], shortage["order_id"]]

        if "source_type" in pr_columns:
            existing_query += " AND source_type = ?"
            existing_params.append("shortage")

        if "status" in pr_columns:
            placeholders = ",".join("?" for _ in active_statuses)
            existing_query += f" AND status IN ({placeholders})"
            existing_params.extend(active_statuses)

        existing_query += " LIMIT 1"

        cursor.execute(existing_query, existing_params)
        existing = cursor.fetchone()
        if existing:
            return jsonify({"ok": False, "error": "An active purchase request already exists for this shortage."}), 409

        request_number = None
        if "request_number" in pr_columns:
            cursor.execute("""
                SELECT COUNT(*)
                FROM purchase_requests
                WHERE company_id = ?
            """, (company_id,))
            pr_count = cursor.fetchone()[0] or 0
            request_number = f"PR-{int(company_id):03d}-{int(pr_count) + 1:05d}"

        insert_map = {}

        if "company_id" in pr_columns:
            insert_map["company_id"] = company_id
        if "request_number" in pr_columns:
            insert_map["request_number"] = request_number
        if "item_id" in pr_columns:
            insert_map["item_id"] = shortage["item_id"]
        if "supplier_id" in pr_columns:
            insert_map["supplier_id"] = shortage["supplier_id"]
        if "title" in pr_columns:
            insert_map["title"] = f"Order shortage - {shortage['item_name']}"
        if "description" in pr_columns:
            insert_map["description"] = f"Created from shortage for order {shortage['order_number']}"
        if "quantity" in pr_columns:
            insert_map["quantity"] = float(shortage["missing_qty"] or 0)
        if "unit" in pr_columns:
            insert_map["unit"] = shortage["measurement_unit"]
        if "status" in pr_columns:
            insert_map["status"] = "draft"
        if "priority" in pr_columns:
            insert_map["priority"] = "Medium"
        if "requested_by" in pr_columns:
            insert_map["requested_by"] = user_id
        if "order_id" in pr_columns:
            insert_map["order_id"] = shortage["order_id"]
        if "source_type" in pr_columns:
            insert_map["source_type"] = "shortage"
        if "linked_product_id" in pr_columns:
            insert_map["linked_product_id"] = shortage["product_id"]
        if "created_at" in pr_columns:
            insert_map["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        required_minimum = {"company_id", "item_id"}
        if not required_minimum.issubset(pr_columns):
            missing_required = ", ".join(sorted(required_minimum - pr_columns))
            return jsonify({"ok": False, "error": f"purchase_requests table is missing required columns: {missing_required}"}), 500

        columns_sql = ", ".join(insert_map.keys())
        placeholders_sql = ", ".join(["?"] * len(insert_map))
        values = list(insert_map.values())

        cursor.execute(
            f"INSERT INTO purchase_requests ({columns_sql}) VALUES ({placeholders_sql})",
            values
        )

        cursor.execute("PRAGMA table_info(shortages)")
        shortage_columns_info = cursor.fetchall()
        shortage_columns = {row["name"] for row in shortage_columns_info}

        if "covered" in shortage_columns:
            cursor.execute("""
                UPDATE shortages
                SET covered = 1
                WHERE id = ?
                  AND company_id = ?
            """, (shortage_id, company_id))

        conn.commit()

        return jsonify({
            "ok": True,
            "message": f"Purchase request created for {shortage['item_name']}."
        })

    except Exception as e:
        conn.rollback()
        return jsonify({
            "ok": False,
            "error": f"Create PR failed: {str(e)}"
        }), 500

    finally:
        conn.close()


@app.route("/inventory/items/<int:item_id>/add-stock", methods=["POST"])
@permission_required("manage_inventory")
def add_item_stock(item_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    add_quantity = float(request.form.get("add_quantity", 0) or 0)

    if add_quantity <= 0:
        flash("Add quantity must be greater than 0.", "error")
        return redirect(request.referrer or url_for("inventory"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM items
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))
    item = cursor.fetchone()

    if item is None:
        conn.close()
        flash("Item not found.", "error")
        return redirect(request.referrer or url_for("inventory"))

    cursor.execute("""
        UPDATE items
        SET stock_quantity = COALESCE(stock_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (add_quantity, item_id, company_id))

    cursor.execute("""
        UPDATE items
        SET available_quantity = COALESCE(stock_quantity, 0) - COALESCE(reserved_quantity, 0)
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))

    cursor.execute("""
        SELECT id, required_qty, reserved_qty, missing_qty
        FROM shortages
        WHERE item_id = ?
          AND company_id = ?
          AND covered = 0
        ORDER BY created_at ASC, id ASC
    """, (item_id, company_id))
    shortage_rows = cursor.fetchall()

    remaining_receipt_qty = add_quantity

    for shortage_id, required_qty, reserved_qty, missing_qty in shortage_rows:
        if remaining_receipt_qty <= 0:
            break

        missing_qty = float(missing_qty or 0)
        reserved_qty = float(reserved_qty or 0)

        if missing_qty <= 0:
            cursor.execute("""
                UPDATE shortages
                SET covered = 1
                WHERE id = ?
                  AND company_id = ?
            """, (shortage_id, company_id))
            continue

        allocate_qty = min(remaining_receipt_qty, missing_qty)

        cursor.execute("""
            UPDATE items
            SET reserved_quantity = COALESCE(reserved_quantity, 0) + ?
            WHERE id = ?
              AND company_id = ?
        """, (allocate_qty, item_id, company_id))

        new_reserved = reserved_qty + allocate_qty
        new_missing = max(0, missing_qty - allocate_qty)

        cursor.execute("""
            UPDATE shortages
            SET reserved_qty = ?,
                missing_qty = ?,
                covered = ?
            WHERE id = ?
              AND company_id = ?
        """, (
            new_reserved,
            new_missing,
            1 if new_missing <= 0 else 0,
            shortage_id,
            company_id
        ))

        remaining_receipt_qty -= allocate_qty

    cursor.execute("""
        UPDATE items
        SET available_quantity = COALESCE(stock_quantity, 0) - COALESCE(reserved_quantity, 0)
        WHERE id = ?
          AND company_id = ?
    """, (item_id, company_id))

    rebuild_company_reserved_quantities(cursor, company_id=company_id)

    conn.commit()
    conn.close()

    flash("Item stock received and shortages updated.", "success")
    return redirect(request.referrer or url_for("inventory"))


@app.route("/inventory/products/<int:product_id>/add-stock", methods=["POST"])
@permission_required("manage_inventory")
def add_product_stock(product_id):
    if not is_logged_in():
        return redirect(url_for("login"))

    company_id = get_company_id()
    add_quantity = float(request.form.get("add_quantity", 0) or 0)

    if add_quantity <= 0:
        flash("Add quantity must be greater than 0.", "error")
        return redirect(request.referrer or url_for("inventory"))

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id
        FROM products
        WHERE id = ?
          AND company_id = ?
    """, (product_id, company_id))
    product = cursor.fetchone()

    if product is None:
        conn.close()
        flash("Product not found.", "error")
        return redirect(request.referrer or url_for("inventory"))

    cursor.execute("""
        UPDATE products
        SET stock_quantity = COALESCE(stock_quantity, 0) + ?
        WHERE id = ?
          AND company_id = ?
    """, (add_quantity, product_id, company_id))

    conn.commit()
    conn.close()

    flash("Product stock added.", "success")
    return redirect(request.referrer or url_for("inventory"))


if __name__ == "__main__":
    app.run(debug=True)
